from __future__ import annotations

from collections import defaultdict
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
import hashlib
import json
from pathlib import Path
import re
import unicodedata

from django.contrib.auth import get_user_model
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel

from assets.models import (
    AssetManagerRole,
    Contract,
    ContractType,
    ModuleToggle,
    Office,
    Supplier,
    Vehicle,
)


User = get_user_model()
EMPTY_MARKERS = {"", "/", "-", "?", "？", "无", "none", "null"}
TERMINATED_WORDS = ("已解除", "解除合作", "已取消", "取消合作", "不续租", "不续签", "终止合作")


def clean_text(value) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return unicodedata.normalize("NFKC", str(value)).strip()


def meaningful(value) -> str:
    value = clean_text(value)
    return "" if value.casefold() in EMPTY_MARKERS else value


def normalize_name(value) -> str:
    return re.sub(r"\s+", "", meaningful(value))


def stable_id(namespace: str, value: str) -> str:
    digest = hashlib.sha1(value.encode("utf-8")).hexdigest()[:16]
    return f"admin-ledger:{namespace}:{digest}"


def stable_code(prefix: str, value: str) -> str:
    digest = hashlib.sha1(value.encode("utf-8")).hexdigest()[:8].upper()
    return f"{prefix}-{digest}"


def as_date(value):
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)) and 1_000 < value < 100_000:
        try:
            converted = from_excel(value)
            return converted.date() if isinstance(converted, datetime) else converted
        except (TypeError, ValueError, OverflowError):
            return None
    text = clean_text(value)
    for pattern in (r"^(\d{4})[-/.年](\d{1,2})[-/.月](\d{1,2})日?$",):
        matched = re.match(pattern, text)
        if matched:
            try:
                return date(*(int(item) for item in matched.groups()))
            except ValueError:
                return None
    return None


def pure_decimal(value):
    if value in (None, "") or isinstance(value, bool):
        return None
    if isinstance(value, (int, float, Decimal)):
        try:
            return Decimal(str(value)).quantize(Decimal("0.01"))
        except (InvalidOperation, ValueError):
            return None
    text = meaningful(value).replace(",", "").replace("，", "")
    if not re.fullmatch(r"\d+(?:\.\d+)?", text):
        return None
    try:
        return Decimal(text).quantize(Decimal("0.01"))
    except InvalidOperation:
        return None


def period_dates(value):
    text = clean_text(value)
    matches = re.findall(r"(20\d{2})[./-](\d{1,2})[./-](\d{1,2})", text)
    parsed = []
    for year, month, day in matches:
        try:
            parsed.append(date(int(year), int(month), int(day)))
        except ValueError:
            continue
    return (parsed[0], parsed[-1]) if len(parsed) >= 2 else (None, None)


def join_notes(*parts) -> str:
    result = []
    for part in parts:
        value = meaningful(part)
        if value and value not in result:
            result.append(value)
    return "\n".join(result)


def contract_status(start_date, end_date, notes, *, legacy=False):
    if any(word in notes for word in TERMINATED_WORDS):
        return Contract.Status.TERMINATED
    if legacy:
        return Contract.Status.COMPLETED
    today = date.today()
    if start_date and start_date > today:
        return Contract.Status.DRAFT
    if end_date and end_date >= today:
        return Contract.Status.ACTIVE
    if end_date and end_date < today:
        return Contract.Status.COMPLETED
    return Contract.Status.DRAFT


class Command(BaseCommand):
    help = "幂等导入行政供应商、合同、车辆和办事处台账"

    def add_arguments(self, parser):
        parser.add_argument("--supplier-workbook", required=True)
        parser.add_argument("--office-workbook", required=True)
        parser.add_argument("--owner", required=True, help="无法匹配责任人时使用的系统用户名")
        parser.add_argument("--dry-run", action="store_true", help="完成校验和统计后回滚")

    def handle(self, *args, **options):
        supplier_path = Path(options["supplier_workbook"]).expanduser().resolve()
        office_path = Path(options["office_workbook"]).expanduser().resolve()
        for path in (supplier_path, office_path):
            if not path.is_file():
                raise CommandError(f"找不到工作簿：{path}")
        owner = User.objects.filter(username=options["owner"], is_active=True).first()
        if not owner:
            raise CommandError(f"找不到可用账号：{options['owner']}")

        self.stats = defaultdict(lambda: defaultdict(int))
        self.warnings = []
        self.users = list(User.objects.filter(is_active=True))

        with transaction.atomic():
            supplier_book = load_workbook(supplier_path, data_only=True)
            office_book = load_workbook(office_path, data_only=True)
            supplier_records = self._collect_suppliers(supplier_book)
            suppliers = self._import_suppliers(supplier_records)
            self._import_supplier_contracts(supplier_book, suppliers, owner)
            self._import_vehicles(supplier_book)
            self._import_offices(office_book, owner)
            self._enable_modules()
            if options["dry_run"]:
                transaction.set_rollback(True)

        summary = {
            key: dict(value)
            for key, value in sorted(self.stats.items())
        }
        summary["warnings"] = len(self.warnings)
        summary["dry_run"] = bool(options["dry_run"])
        self.stdout.write(f"IMPORT_SUMMARY {json.dumps(summary, ensure_ascii=False, sort_keys=True)}")
        for warning in self.warnings:
            self.stdout.write(self.style.WARNING(f"WARN {warning}"))

    def _record(self, section, outcome):
        self.stats[section][outcome] += 1

    def _set_fields(self, instance, defaults):
        changed = []
        for field, value in defaults.items():
            current = getattr(instance, f"{field}_id", None) if hasattr(value, "pk") else getattr(instance, field)
            wanted = value.pk if hasattr(value, "pk") else value
            if current != wanted:
                setattr(instance, field, value)
                changed.append(field)
        if changed:
            instance.save(update_fields=[*changed, "updated_at"])
        return bool(changed)

    def _find_owner(self, source_contact, fallback):
        contact = normalize_name(source_contact)
        if not contact:
            return fallback
        candidates = []
        for user in self.users:
            for label in (user.get_full_name(), f"{user.last_name}{user.first_name}", user.username):
                alias = normalize_name(label)
                if alias and len(alias) >= 2 and alias in contact:
                    candidates.append((len(alias), user))
        return max(candidates, key=lambda item: item[0])[1] if candidates else fallback

    def _collect_suppliers(self, workbook):
        records = {}

        def add(name, *, source, category="", brand="", scope="", cooperation="", evaluation="",
                started="", channel="cooperative", contact="", phone="", address="", license_flag="",
                active=None, note=""):
            normalized = normalize_name(name)
            if not normalized:
                return
            record = records.setdefault(normalized, {
                "name": normalized,
                "categories": [], "brands": [], "scopes": [], "cooperation": [],
                "evaluations": [], "started": [], "channels": [], "contacts": [],
                "phones": [], "addresses": [], "licenses": [], "active": [], "sources": [],
                "notes": [],
            })
            for key, value in [
                ("categories", category), ("brands", brand), ("scopes", scope),
                ("cooperation", cooperation), ("evaluations", evaluation), ("started", started),
                ("channels", channel), ("contacts", contact), ("phones", phone),
                ("addresses", address), ("licenses", license_flag), ("notes", note),
            ]:
                value = meaningful(value)
                if value and value not in record[key]:
                    record[key].append(value)
            if active is not None:
                record["active"].append(bool(active))
            record["sources"].append(source)

        ws = workbook["01 行政模块供应商合同档案"]
        for row in range(4, 39):
            end = as_date(ws.cell(row, 16).value)
            notes = clean_text(ws.cell(row, 26).value)
            add(
                ws.cell(row, 11).value,
                source=f"行政合同表第{row}行",
                category=ws.cell(row, 5).value,
                scope=ws.cell(row, 25).value,
                cooperation=ws.cell(row, 6).value,
                channel="cooperative",
                contact=ws.cell(row, 12).value,
                address=ws.cell(row, 13).value,
                license_flag=ws.cell(row, 14).value,
                active=bool(end and end >= date.today() and not any(word in notes for word in TERMINATED_WORDS)),
            )

        ws = workbook["02 其他模块行政供应商合同档案"]
        for row in range(4, 217):
            add(
                ws.cell(row, 10).value,
                source=f"其他合同表第{row}行",
                category=ws.cell(row, 4).value,
                scope=ws.cell(row, 20).value,
                cooperation=ws.cell(row, 5).value,
                channel="cooperative",
                contact=ws.cell(row, 11).value,
                address=ws.cell(row, 12).value,
                license_flag=ws.cell(row, 13).value,
                active=False,
            )

        ws = workbook["03 线下供应商名单"]
        for row in range(3, 104):
            name = ws.cell(row, 8).value
            if not normalize_name(name):
                continue
            cooperation = meaningful(ws.cell(row, 5).value)
            eliminated = meaningful(ws.cell(row, 15).value)
            raw_channel = meaningful(ws.cell(row, 7).value)
            channel = "cooperative" if "合同" in raw_channel else "ecommerce" if "电商" in raw_channel else "other"
            active = cooperation == "是" and "淘汰" not in eliminated
            add(
                name,
                source=f"线下供应商表第{row}行",
                category=ws.cell(row, 3).value,
                brand=ws.cell(row, 9).value,
                scope=ws.cell(row, 10).value,
                cooperation=cooperation,
                evaluation=ws.cell(row, 6).value,
                started=ws.cell(row, 4).value,
                channel=channel,
                contact=ws.cell(row, 11).value,
                phone=ws.cell(row, 12).value,
                address=ws.cell(row, 13).value,
                active=active,
                note=join_notes(ws.cell(row, 14).value, eliminated),
            )
        return records

    def _import_suppliers(self, records):
        result = {}
        for normalized, record in sorted(records.items()):
            external_id = stable_id("supplier", normalized)
            licenses = record["licenses"]
            has_license = any(value.startswith("有") for value in licenses)
            missing_license = any(value.startswith("无") for value in licenses)
            license_status = "registered" if has_license else "missing" if missing_license else "unknown"
            conflict_note = "营业执照状态来源存在冲突，按“已登记”保留，实际文件待补。" if has_license and missing_license else ""
            contact_pairs = [
                " / ".join(item for item in pair if item)
                for pair in zip(record["contacts"], record["phones"] + [""] * len(record["contacts"]))
            ]
            defaults = {
                "code": stable_code("SUP", normalized),
                "name": record["name"],
                "category": "；".join(record["categories"])[:120],
                "brand_name": "；".join(record["brands"])[:120],
                "business_scope": "；".join(record["scopes"]),
                "cooperation_status": "；".join(record["cooperation"])[:40],
                "evaluation": "；".join(record["evaluations"])[:40],
                "cooperation_started": "；".join(record["started"])[:40],
                "channel": record["channels"][0] if record["channels"] else Supplier.Channel.COOPERATIVE,
                "contact_name": record["contacts"][0][:80] if record["contacts"] else "",
                "contact_phone": record["phones"][0][:40] if record["phones"] else "",
                "address": record["addresses"][0][:255] if record["addresses"] else "",
                "business_license_status": license_status,
                "external_id": external_id,
                "notes": join_notes(
                    conflict_note,
                    "来源：" + "；".join(record["sources"]),
                    "其他联系人：" + "；".join(contact_pairs[1:]) if len(contact_pairs) > 1 else "",
                    *record["notes"],
                ),
                "is_active": any(record["active"]) if record["active"] else True,
            }
            instance = Supplier.objects.filter(external_id=external_id).first()
            if not instance:
                instance = Supplier.objects.filter(name=record["name"]).first()
            if instance:
                self._record("suppliers", "updated" if self._set_fields(instance, defaults) else "unchanged")
            else:
                instance = Supplier.objects.create(**defaults)
                self._record("suppliers", "created")
            result[normalized] = instance
        return result

    def _contract_type(self, name, fallback="历史行政合同"):
        name = meaningful(name) or fallback
        contract_type, _ = ContractType.objects.get_or_create(
            name=name[:80],
            defaults={"code": stable_code("CT", name), "is_active": True},
        )
        return contract_type

    def _unique_contract_no(self, preferred, external_id):
        preferred = re.sub(r"\s+", "", meaningful(preferred))[:64] or stable_code("LEG", external_id)
        existing = Contract.objects.filter(external_id=external_id).first()
        if existing:
            return existing.contract_no
        candidate = preferred
        index = 2
        while Contract.objects.filter(contract_no=candidate).exclude(external_id=external_id).exists():
            suffix = f"-{index:02d}"
            candidate = f"{preferred[:64-len(suffix)]}{suffix}"
            index += 1
        return candidate

    def _upsert_contract(self, external_id, defaults):
        instance = Contract.objects.filter(external_id=external_id).first()
        defaults["contract_no"] = self._unique_contract_no(defaults["contract_no"], external_id)
        defaults["external_id"] = external_id
        if instance:
            self._record("contracts", "updated" if self._set_fields(instance, defaults) else "unchanged")
        else:
            instance = Contract.objects.create(**defaults)
            self._record("contracts", "created")
        return instance

    def _import_supplier_contracts(self, workbook, suppliers, fallback_owner):
        ws = workbook["01 行政模块供应商合同档案"]
        seen_numbers = defaultdict(int)
        for row in range(4, 39):
            name = meaningful(ws.cell(row, 3).value)
            if not name:
                continue
            external_id = f"admin-ledger:contract:administrative:{row}"
            raw_no = re.sub(r"\s+", "", meaningful(ws.cell(row, 2).value))
            seen_numbers[raw_no] += 1
            contract_no = raw_no or f"LEG-ADM-{row:04d}"
            if seen_numbers[raw_no] > 1:
                contract_no = f"{contract_no}-{seen_numbers[raw_no]:02d}"
            start = as_date(ws.cell(row, 15).value)
            end = as_date(ws.cell(row, 16).value)
            warning = ""
            if start and end and end < start:
                warning = f"源表结束日期 {end} 早于开始日期 {start}，未导入结束日期。"
                end = None
                self.warnings.append(f"行政合同表第{row}行日期倒置")
            notes = join_notes(ws.cell(row, 26).value, warning)
            supplier = suppliers.get(normalize_name(ws.cell(row, 11).value))
            self._upsert_contract(external_id, {
                "contract_no": contract_no,
                "name": name[:180],
                "contract_type": self._contract_type(ws.cell(row, 4).value),
                "supplier": supplier,
                "owner": self._find_owner(ws.cell(row, 9).value, fallback_owner),
                "status": contract_status(start, end, notes),
                "start_date": start,
                "end_date": end,
                "amount": Decimal("0"),
                "amount_description": meaningful(ws.cell(row, 21).value),
                "cooperation_duration": meaningful(ws.cell(row, 6).value)[:120],
                "cooperation_type": meaningful(ws.cell(row, 7).value)[:120],
                "party_a": meaningful(ws.cell(row, 8).value)[:180],
                "party_a_contact": meaningful(ws.cell(row, 9).value)[:160],
                "party_a_address": meaningful(ws.cell(row, 10).value)[:500],
                "party_b_contact": meaningful(ws.cell(row, 12).value)[:160],
                "party_b_address": meaningful(ws.cell(row, 13).value)[:500],
                "payment_method": meaningful(ws.cell(row, 19).value)[:160],
                "payment_terms": meaningful(ws.cell(row, 20).value),
                "invoice_type": meaningful(ws.cell(row, 22).value)[:80],
                "invoice_tax_rate": meaningful(ws.cell(row, 23).value)[:80],
                "service_content": meaningful(ws.cell(row, 25).value),
                "notes": join_notes(
                    "来源：行政模块供应商合同档案",
                    f"供应商类别：{meaningful(ws.cell(row, 5).value)}" if meaningful(ws.cell(row, 5).value) else "",
                    f"费用说明：{meaningful(ws.cell(row, 24).value)}" if meaningful(ws.cell(row, 24).value) else "",
                    notes,
                ),
            })

        ws = workbook["02 其他模块行政供应商合同档案"]
        for row in range(4, 217):
            name = meaningful(ws.cell(row, 3).value)
            if not name:
                continue
            start = as_date(ws.cell(row, 14).value)
            end = as_date(ws.cell(row, 15).value)
            notes = meaningful(ws.cell(row, 21).value)
            external_id = f"admin-ledger:contract:other:{row}"
            self._upsert_contract(external_id, {
                "contract_no": f"LEG-OTH-{row - 3:04d}",
                "name": name[:180],
                "contract_type": self._contract_type(ws.cell(row, 4).value),
                "supplier": suppliers.get(normalize_name(ws.cell(row, 10).value)),
                "owner": self._find_owner(ws.cell(row, 8).value, fallback_owner),
                "status": contract_status(start, end, notes, legacy=True),
                "start_date": start,
                "end_date": end,
                "amount": Decimal("0"),
                "amount_description": meaningful(ws.cell(row, 19).value),
                "cooperation_duration": meaningful(ws.cell(row, 5).value)[:120],
                "cooperation_type": meaningful(ws.cell(row, 6).value)[:120],
                "party_a": meaningful(ws.cell(row, 7).value)[:180],
                "party_a_contact": meaningful(ws.cell(row, 8).value)[:160],
                "party_a_address": meaningful(ws.cell(row, 9).value)[:500],
                "party_b_contact": meaningful(ws.cell(row, 11).value)[:160],
                "party_b_address": meaningful(ws.cell(row, 12).value)[:500],
                "payment_method": meaningful(ws.cell(row, 18).value)[:160],
                "service_content": meaningful(ws.cell(row, 20).value),
                "notes": join_notes("来源：其他模块行政供应商合同档案", notes),
            })

    def _import_vehicles(self, workbook):
        ws = workbook["02 车辆保险到期管理"]
        for row in range(4, 11):
            plate = normalize_name(ws.cell(row, 3).value)
            if not plate:
                continue
            source_notes = meaningful(ws.cell(row, 21).value)
            retired = "已变卖" in source_notes or "已归属" in source_notes
            name = meaningful(ws.cell(row, 2).value) or plate
            defaults = {
                "name": name[:120],
                "brand": name[:80],
                "model_name": meaningful(ws.cell(row, 4).value)[:120],
                "vin": meaningful(ws.cell(row, 9).value)[:64],
                "engine_number": meaningful(ws.cell(row, 8).value)[:64],
                "energy_type": Vehicle.EnergyType.ELECTRIC if "纯电" in meaningful(ws.cell(row, 4).value) or "极氪" in name else Vehicle.EnergyType.GASOLINE,
                "seats": int(ws.cell(row, 7).value) if isinstance(ws.cell(row, 7).value, (int, float)) else 5,
                "status": Vehicle.Status.RETIRED if retired else Vehicle.Status.AVAILABLE,
                "purchase_cost": pure_decimal(ws.cell(row, 5).value),
                "registration_date": as_date(ws.cell(row, 6).value),
                "company": meaningful(ws.cell(row, 11).value)[:120],
                "use_scope": meaningful(ws.cell(row, 10).value)[:255],
                "insurance_started_at": as_date(ws.cell(row, 12).value),
                "insurance_expires_at": as_date(ws.cell(row, 13).value),
                "insurer_name": meaningful(ws.cell(row, 16).value)[:120],
                "asset_card_code": meaningful(ws.cell(row, 17).value)[:80],
                "asset_number": meaningful(ws.cell(row, 18).value)[:80],
                "handler_name": meaningful(ws.cell(row, 19).value)[:80],
                "supervisor_name": meaningful(ws.cell(row, 20).value)[:80],
                "notes": join_notes("来源：车辆保险到期管理", source_notes),
            }
            instance = Vehicle.objects.filter(plate_number=plate).first()
            if instance:
                self._record("vehicles", "updated" if self._set_fields(instance, defaults) else "unchanged")
            else:
                Vehicle.objects.create(plate_number=plate, **defaults)
                self._record("vehicles", "created")

    def _import_offices(self, workbook, owner):
        ws = workbook["各办事处登记表"]
        lease_type = self._contract_type("办事处租赁")
        for row in range(4, 31):
            sequence = int(ws.cell(row, 1).value)
            code = f"OFFICE-{sequence:03d}"
            name = meaningful(ws.cell(row, 4).value)
            region = meaningful(ws.cell(row, 3).value)
            notes = []
            if meaningful(ws.cell(row, 2).value) == "衡水" and region == "江苏":
                region = "河北"
                notes.append("源表区域为“江苏”，按城市“衡水”纠正为“河北”。")
                self.warnings.append(f"办事处表序号{sequence}区域已纠正")
            lease_start = as_date(ws.cell(row, 36).value)
            source_end = as_date(ws.cell(row, 37).value)
            period_start, period_end = period_dates(ws.cell(row, 35).value)
            lease_end = source_end
            expected_move_out = None
            renewal = meaningful(ws.cell(row, 33).value)
            if period_end and source_end and period_end != source_end and "退租" in renewal:
                lease_end = period_end
                expected_move_out = source_end
                notes.append(f"源表终止日期 {source_end} 结合退租反馈作为预计退租日；合同期到 {period_end}。")
                self.warnings.append(f"办事处表序号{sequence}已拆分合同到期日与预计退租日")
            elif period_end and source_end and period_end != source_end:
                notes.append(f"最近租赁期文本截止 {period_end}，终止日期列为 {source_end}，均保留待核对。")
            residents = meaningful(ws.cell(row, 41).value)
            resident_count = ws.cell(row, 40).value
            resident_count = int(resident_count) if isinstance(resident_count, (int, float)) else None
            resident_names = [item for item in re.split(r"[、,，;；\s]+", residents) if item]
            if resident_count is not None and residents and "流动" not in residents and len(resident_names) != resident_count:
                notes.append(f"源表居住人数为 {resident_count}，名单识别为 {len(resident_names)} 人，原值均保留。")
            latest_amount = pure_decimal(ws.cell(row, 29).value)
            if latest_amount is None and meaningful(ws.cell(row, 29).value):
                notes.append(f"近期付款金额原文：{meaningful(ws.cell(row, 29).value)}")
            notes.extend(filter(None, [
                f"开票明细：{meaningful(ws.cell(row, 42).value)}" if meaningful(ws.cell(row, 42).value) else "",
                f"房东税费：{meaningful(ws.cell(row, 43).value)}" if meaningful(ws.cell(row, 43).value) else "",
            ]))
            defaults = {
                "code": code,
                "name": name[:160],
                "status": Office.Status.ACTIVE if not lease_end or lease_end >= date.today() else Office.Status.CLOSED,
                "region": region[:80],
                "city": meaningful(ws.cell(row, 2).value)[:80],
                "address": meaningful(ws.cell(row, 8).value)[:500],
                "room_layout": meaningful(ws.cell(row, 9).value)[:120],
                "area_sqm": pure_decimal(ws.cell(row, 10).value),
                "sales_project": meaningful(ws.cell(row, 5).value)[:255],
                "cost_attribution": meaningful(ws.cell(row, 6).value)[:255],
                "landlord_name": meaningful(ws.cell(row, 11).value)[:160],
                "landlord_phone": meaningful(ws.cell(row, 12).value)[:80],
                "intermediary_name": meaningful(ws.cell(row, 14).value)[:160],
                "intermediary_phone": meaningful(ws.cell(row, 15).value)[:80],
                "intermediary_fee": pure_decimal(ws.cell(row, 16).value),
                "intermediary_invoice_status": meaningful(ws.cell(row, 17).value)[:120],
                "monthly_rent": pure_decimal(ws.cell(row, 13).value),
                "rent_description": meaningful(ws.cell(row, 13).value),
                "deposit": pure_decimal(ws.cell(row, 18).value),
                "deposit_status": meaningful(ws.cell(row, 19).value)[:160],
                "payment_frequency": meaningful(ws.cell(row, 20).value)[:80],
                "payment_method": meaningful(ws.cell(row, 21).value)[:120],
                "payment_terms": meaningful(ws.cell(row, 22).value),
                "latest_payment_period": meaningful(ws.cell(row, 23).value)[:160],
                "paid_period_start": as_date(ws.cell(row, 24).value),
                "paid_period_end": as_date(ws.cell(row, 25).value),
                "latest_payment_date": as_date(ws.cell(row, 26).value),
                "next_payment_date": as_date(ws.cell(row, 28).value),
                "latest_payment_amount": latest_amount,
                "responsible_name": meaningful(ws.cell(row, 30).value)[:80],
                "responsible_phone": meaningful(ws.cell(row, 31).value)[:80],
                "residents": residents,
                "resident_count": resident_count,
                "renewal_status": renewal,
                "lease_summary": meaningful(ws.cell(row, 34).value),
                "current_lease_period": meaningful(ws.cell(row, 35).value)[:160],
                "lease_start": lease_start or period_start,
                "lease_end": lease_end,
                "expected_move_out_date": expected_move_out,
                "feedback": meaningful(ws.cell(row, 32).value),
                "notes": join_notes(*notes),
                "external_id": f"admin-ledger:office:{sequence}",
            }
            office = Office.objects.filter(external_id=defaults["external_id"]).first() or Office.objects.filter(code=code).first()
            if office:
                self._record("offices", "updated" if self._set_fields(office, defaults) else "unchanged")
            else:
                office = Office.objects.create(**defaults)
                self._record("offices", "created")

            contract_external_id = f"admin-ledger:office-contract:{sequence}"
            contract_end = lease_end
            contract_notes = join_notes(
                "来源：办事处信息登记表",
                f"出租方：{defaults['landlord_name']} {defaults['landlord_phone']}".strip(),
                f"中介：{defaults['intermediary_name']} {defaults['intermediary_phone']}".strip(),
                f"押金：{meaningful(ws.cell(row, 18).value)}；状态：{defaults['deposit_status']}",
                f"续租情况：{renewal}" if renewal else "",
                defaults["notes"],
            )
            self._upsert_contract(contract_external_id, {
                "contract_no": re.sub(r"\s+", "", meaningful(ws.cell(row, 7).value)) or f"OFFICE-LEASE-{sequence:03d}",
                "name": f"{name}租赁合同"[:180],
                "contract_type": lease_type,
                "office": office,
                "owner": owner,
                "status": contract_status(lease_start, contract_end, contract_notes),
                "start_date": lease_start,
                "end_date": contract_end,
                "amount": Decimal("0"),
                "amount_description": meaningful(ws.cell(row, 13).value),
                "party_b_contact": join_notes(defaults["landlord_name"], defaults["landlord_phone"])[:160],
                "party_b_address": defaults["address"],
                "payment_method": defaults["payment_method"],
                "payment_terms": defaults["payment_terms"],
                "service_content": "办事处房屋租赁",
                "notes": contract_notes,
            })

    def _enable_modules(self):
        for code, label in [
            ("vehicles", "车辆管理"),
            ("suppliers", "供应商管理"),
            ("offices", "办事处管理"),
        ]:
            instance, created = ModuleToggle.objects.update_or_create(
                code=code, defaults={"label": label, "is_enabled": True}
            )
            self._record("modules", "created" if created else "updated")
        for role in AssetManagerRole.objects.all():
            scopes = list(role.scopes or [])
            before = list(scopes)
            if "procurement" in scopes and "suppliers" not in scopes:
                scopes.append("suppliers")
            if "contracts" in scopes and "offices" not in scopes:
                scopes.append("offices")
            if scopes != before:
                role.scopes = scopes
                role.save(update_fields=["scopes", "updated_at"])
