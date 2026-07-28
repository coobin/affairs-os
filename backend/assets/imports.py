import re
from datetime import date, datetime
from decimal import Decimal, InvalidOperation

from django.contrib.auth import get_user_model
from django.db import transaction
from django.db.models import Q
from openpyxl import load_workbook
import hashlib
import string

from .models import (
    Asset,
    AssetCategory,
    AssetEvent,
    AssetStatus,
    Department,
    InventoryItem,
    InventoryTransaction,
    Location,
)
from .services import generate_asset_tag

User = get_user_model()

MAX_IMPORT_ROWS = 2000
MAX_IMPORT_BYTES = 5 * 1024 * 1024

HEADER_ALIASES = {
    "asset_tag": {"编码", "资产编号", "资产编码"},
    "name": {"名称", "资产名称", "设备名称", "物品名称"},
    "system_code": {"系统编码", "系统编号"},
    "kingdee_code": {"金蝶编码", "金蝶资产编码"},
    "assignee": {"责任人", "使用人", "领用人", "保管人"},
    "status": {"状态", "资产状态"},
    "department": {"部门", "保管部门", "使用部门"},
    "class_type": {"资产大类", "资产归类"},
    "category": {"资产类型", "分类", "详细类型"},
    "location": {"资产位置", "位置", "当前地点"},
    "brand": {"品牌"},
    "model_name": {"型号", "设备型号"},
    "quantity": {"数量"},
    "serial_number": {"设备序列号", "序列号", "sn", "SN"},
    "specification": {"规格配置", "主要配置", "配置"},
    "cpu": {"CPU", "cpu", "处理器"},
    "storage": {"硬盘大小(G)", "硬盘大小（G）", "硬盘", "存储"},
    "memory": {"内存(g)", "内存(G)", "内存（G）", "内存"},
    "wired_mac": {"有线MAC地址", "有线 MAC 地址", "有线MAC"},
    "wireless_mac": {"无线MAC地址", "无线 MAC 地址", "无线MAC"},
    "purchase_date": {"购买日期", "采购日期"},
    "purchase_cost": {"采购金额", "购买金额", "金额"},
    "warranty_expires_at": {"保修到期", "保修截止日期"},
    "notes": {"备注", "说明"},
}

STATUS_ALIASES = {
    "在库": Asset.Status.AVAILABLE,
    "冻结": Asset.Status.FROZEN,
    "在库可用": Asset.Status.AVAILABLE,
    "库存": Asset.Status.AVAILABLE,
    "使用中": Asset.Status.ASSIGNED,
    "在用": Asset.Status.ASSIGNED,
    "借用中": Asset.Status.LOANED,
    "借用": Asset.Status.LOANED,
    "长期借用中": Asset.Status.LOANED,
    "报废": Asset.Status.DISPOSED,
    # 旧模板状态兼容导入并折叠到新的四种基础状态。
    "待验收": Asset.Status.AVAILABLE,
    "待检": Asset.Status.AVAILABLE,
    "调拨中": Asset.Status.AVAILABLE,
    "维修中": Asset.Status.AVAILABLE,
    "遗失": Asset.Status.DISPOSED,
    "已退役": Asset.Status.DISPOSED,
    "退役": Asset.Status.DISPOSED,
    "已处置": Asset.Status.DISPOSED,
}

CLASS_TYPE_ALIASES = {
    "it": AssetCategory.ClassType.IT,
    "it资产": AssetCategory.ClassType.IT,
    "信息资产": AssetCategory.ClassType.IT,
    "信息化资产": AssetCategory.ClassType.IT,
    "行政": AssetCategory.ClassType.ADMIN,
    "行政资产": AssetCategory.ClassType.ADMIN,
    "办公资产": AssetCategory.ClassType.ADMIN,
}


class AssetImportError(Exception):
    pass


class InventoryImportError(Exception):
    pass


def _text(value):
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _normalize_header(value):
    return re.sub(r"\s+", "", _text(value)).lower()


def _header_map(values):
    normalized_aliases = {
        field: {_normalize_header(alias) for alias in aliases}
        for field, aliases in HEADER_ALIASES.items()
    }
    result = {}
    normalized_values = [_normalize_header(value) for value in values]
    has_asset_type = "资产类型" in normalized_values
    for index, normalized in enumerate(normalized_values):
        # 新模板同时包含“资产分类”和“资产类型”；旧模板只有“资产分类”时，
        # 继续将其理解为笔记本电脑、显示器等详细类型。
        if normalized == "资产分类":
            result["class_type" if has_asset_type else "category"] = index
            continue
        for field, aliases in normalized_aliases.items():
            if normalized in aliases:
                result[field] = index
                break
    return result


def _resolve_class_type(value):
    text = _text(value)
    if not text:
        return AssetCategory.ClassType.IT, False
    resolved = CLASS_TYPE_ALIASES.get(_normalize_header(text))
    return resolved, True


def _find_header_row(sheet):
    last_row = min(sheet.max_row or 20, 20)
    for row_number, row in enumerate(
        sheet.iter_rows(min_row=1, max_row=last_row, values_only=True),
        start=1,
    ):
        mapping = _header_map(row)
        if len(mapping) >= 5 and any(
            field in mapping for field in ("name", "category", "model_name", "serial_number")
        ):
            return row_number, mapping
    raise AssetImportError("没有找到表头，请使用系统模板，或保留资产类型、型号等主要列。")


def _asset_display_name(brand, model_name, category_name):
    brand = _text(brand)
    model_name = _text(model_name)
    if brand and model_name:
        return (
            model_name
            if model_name.casefold().startswith(brand.casefold())
            else f"{brand} {model_name}"
        )
    return brand or model_name or category_name or "待完善资产"


def _parse_date(value):
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = _text(value).replace(".", "-").replace("/", "-")
    if re.fullmatch(r"\d{8}", text):
        text = f"{text[:4]}-{text[4:6]}-{text[6:]}"
    for fmt in ("%Y-%m-%d", "%Y-%m", "%Y"):
        try:
            parsed = datetime.strptime(text, fmt).date()
            return parsed.replace(month=1, day=1) if fmt == "%Y" else parsed
        except ValueError:
            pass
    raise ValueError("日期格式应为 2026-07-23")


def _parse_decimal(value):
    if value in (None, ""):
        return None
    try:
        result = Decimal(_text(value).replace(",", "").replace("￥", ""))
    except InvalidOperation as exc:
        raise ValueError("金额不是有效数字") from exc
    if result < 0:
        raise ValueError("金额不能小于 0")
    return result


def _generated_code(prefix, name):
    base = re.sub(r"[^A-Z0-9]", "", name.upper())[:8]
    candidate = f"{prefix}-{base}"[:12] if base else prefix
    index = 2
    model = Location
    while model.objects.filter(code=candidate).exists():
        suffix = str(index)
        candidate = (
            f"{prefix}-{base[: 11 - len(prefix) - len(suffix)]}{suffix}"[:12]
            if base
            else f"{prefix}{suffix}"
        )
        index += 1
    return candidate


def _generated_category_code(name):
    base = re.sub(r"[^A-Z]", "", name.upper())
    if len(base) >= 2:
        candidate = base[:2]
    elif len(base) == 1:
        candidate = base + "A"
    else:
        h = int(hashlib.md5(name.encode("utf-8")).hexdigest(), 16)
        candidate = string.ascii_uppercase[h % 26] + string.ascii_uppercase[(h // 26) % 26]

    index = 0
    original_candidate = candidate
    while AssetCategory.objects.filter(code=candidate).exists():
        index += 1
        if index < 26:
            c2 = string.ascii_uppercase[(string.ascii_uppercase.index(original_candidate[1]) + index) % 26]
            candidate = original_candidate[0] + c2
        else:
            candidate = f"{original_candidate[0]}{index}"
    return candidate


def _find_user(name):
    name = _text(name)
    if not name:
        return None
    queryset = User.objects.filter(is_active=True).filter(
        Q(username__iexact=name)
        | Q(first_name__iexact=name)
        | Q(last_name__iexact=name)
        | Q(employee_profile__employee_no__iexact=name)
    ).distinct()
    exact = [
        user
        for user in queryset
        if (user.get_full_name() or user.username).strip().lower() == name.lower()
        or user.username.lower() == name.lower()
        or getattr(getattr(user, "employee_profile", None), "employee_no", "").lower()
        == name.lower()
    ]
    if len(exact) == 1:
        return exact[0]
    if len(exact) > 1:
        raise ValueError(f"责任人“{name}”匹配到多人，请填写工号")
    raise ValueError(f"找不到责任人“{name}”，请先同步人员或填写正确姓名/工号")


def _resolve_asset_status(value):
    text = _text(value)
    if not text:
        return Asset.Status.AVAILABLE
    alias = STATUS_ALIASES.get(text)
    if alias:
        return alias
    configured = AssetStatus.objects.filter(is_active=True).filter(
        Q(code__iexact=text) | Q(name__iexact=text)
    ).values_list("code", flat=True).first()
    return configured


def _cell(row, mapping, field):
    index = mapping.get(field)
    return row[index] if index is not None and index < len(row) else None


def _find_existing_asset(serial_number, kingdee_code, system_code):
    matches = []
    identifiers = [
        ("序列号", serial_number, Asset.objects.filter(serial_number__iexact=serial_number)),
        ("金蝶编码", kingdee_code, Asset.objects.filter(kingdee_code__iexact=kingdee_code)),
        ("系统编码", system_code, Asset.objects.filter(custom_data__system_code=system_code)),
    ]
    for label, value, queryset in identifiers:
        if not value:
            continue
        candidates = list(queryset.order_by("id")[:2])
        if len(candidates) > 1:
            raise ValueError(f"{label}“{value}”匹配到多件资产，请先整理重复数据")
        if candidates:
            matches.append(candidates[0])
    matched_ids = {asset.pk for asset in matches}
    if len(matched_ids) > 1:
        raise ValueError("序列号、金蝶编码和系统编码指向不同资产")
    return matches[0] if matches else None


def parse_asset_workbook(file_obj):
    if getattr(file_obj, "size", 0) > MAX_IMPORT_BYTES:
        raise AssetImportError("文件不能超过 5MB。")
    try:
        workbook = load_workbook(file_obj, read_only=True, data_only=True)
    except Exception as exc:
        raise AssetImportError("无法读取 Excel，请上传有效的 .xlsx 文件。") from exc
    sheet = workbook.active
    header_row, mapping = _find_header_row(sheet)
    rows = []
    seen_existing_ids = set()
    seen_category_classes = {}
    existing_category_classes = {
        category.name.casefold(): category.class_type
        for category in AssetCategory.objects.only("name", "class_type")
    }
    for row_number, row in enumerate(
        sheet.iter_rows(min_row=header_row + 1, values_only=True),
        start=header_row + 1,
    ):
        if not any(value not in (None, "") for value in row):
            continue
        if len(rows) >= MAX_IMPORT_ROWS:
            raise AssetImportError(f"单次最多导入 {MAX_IMPORT_ROWS} 行。")
        raw = {field: _cell(row, mapping, field) for field in mapping}
        source_tag = _text(raw.get("asset_tag"))
        errors = []
        warnings = []
        if source_tag:
            warnings.append("表中原编码仅作参考，系统资产编号不会采用该值")

        source_name = _text(raw.get("name"))
        explicit_category_name = _text(raw.get("category"))
        category_name = explicit_category_name or source_name
        if not explicit_category_name and source_name:
            warnings.append(f"资产类型为空，已将名称“{source_name}”作为资产类型")
        elif not category_name:
            warnings.append("资产类型为空，导入后归入“待分类”")
        class_type, class_type_explicit = _resolve_class_type(raw.get("class_type"))
        if not class_type_explicit and category_name:
            class_type = existing_category_classes.get(category_name.casefold(), class_type)
        if class_type is None:
            errors.append(f"无法识别资产分类“{_text(raw.get('class_type'))}”，请填写 IT资产或行政资产")
        elif category_name and class_type_explicit:
            category_key = category_name.casefold()
            previous_class = seen_category_classes.get(category_key)
            if previous_class and previous_class != class_type:
                errors.append(f"同一资产类型“{category_name}”不能同时属于 IT资产和行政资产")
            else:
                seen_category_classes[category_key] = class_type
        quantity = _text(raw.get("quantity"))
        if quantity and quantity not in {"1", "1.0"}:
            errors.append("单件资产的数量必须为 1")

        status_text = _text(raw.get("status"))
        status = _resolve_asset_status(status_text)
        if status is None:
            errors.append(f"无法识别状态“{status_text}”")

        try:
            assignee = _find_user(raw.get("assignee"))
        except ValueError as exc:
            assignee = None
            warnings.append(str(exc))
        try:
            purchase_date = _parse_date(raw.get("purchase_date"))
        except ValueError as exc:
            purchase_date = None
            errors.append(str(exc))
        try:
            warranty = _parse_date(raw.get("warranty_expires_at"))
        except ValueError as exc:
            warranty = None
            errors.append(str(exc))
        try:
            purchase_cost = _parse_decimal(raw.get("purchase_cost"))
        except ValueError as exc:
            purchase_cost = None
            errors.append(str(exc))
        if purchase_date and warranty and warranty < purchase_date:
            errors.append("保修到期不能早于购买日期")

        serial = _text(raw.get("serial_number"))
        kingdee_code = _text(raw.get("kingdee_code"))
        system_code = _text(raw.get("system_code"))
        for field, label, limit in [
            ("kingdee_code", "金蝶编码", 64),
            ("brand", "品牌", 80),
            ("model_name", "型号", 120),
            ("serial_number", "设备序列号", 120),
            ("specification", "规格配置", 255),
            ("cpu", "CPU", 120),
            ("storage", "硬盘", 120),
            ("memory", "内存", 80),
            ("wired_mac", "有线 MAC 地址", 255),
            ("wireless_mac", "无线 MAC 地址", 255),
        ]:
            value = _text(raw.get(field))
            if len(value) > limit:
                errors.append(f"{label}不能超过 {limit} 个字符")
        try:
            existing = _find_existing_asset(serial, kingdee_code, system_code)
        except ValueError as exc:
            existing = None
            errors.append(str(exc))
        if existing and existing.pk in seen_existing_ids:
            errors.append(f"同一现有资产 {existing.asset_tag} 在文件中出现多次")
        if existing:
            seen_existing_ids.add(existing.pk)
        imported_status = status
        assignee_text = _text(raw.get("assignee"))
        if (
            not existing
            and assignee_text
            and assignee is None
            and status in {Asset.Status.ASSIGNED, Asset.Status.LOANED}
        ):
            imported_status = Asset.Status.AVAILABLE
            warnings.append("责任人未匹配，状态暂按“在库”导入")
        status_label = (
            AssetStatus.objects.filter(code=status).values_list("name", flat=True).first()
            if status
            else status_text
        )
        rows.append(
            {
                "row_number": row_number,
                "action": "update" if existing else "create",
                "errors": errors,
                "warnings": warnings,
                "data": {
                    "asset_tag": existing.asset_tag if existing else "",
                    "source_asset_tag": source_tag,
                    "existing_id": existing.pk if existing else None,
                    "source_name": source_name,
                    "category_name": category_name or "待分类",
                    "category_missing": not bool(category_name),
                    "class_type": class_type or AssetCategory.ClassType.IT,
                    "class_type_explicit": class_type_explicit,
                    "brand": _text(raw.get("brand")),
                    "model_name": _text(raw.get("model_name")),
                    "serial_number": serial,
                    "specification": _text(raw.get("specification")),
                    "cpu": _text(raw.get("cpu")),
                    "storage": _text(raw.get("storage")),
                    "memory": _text(raw.get("memory")),
                    "wired_mac": _text(raw.get("wired_mac")),
                    "wireless_mac": _text(raw.get("wireless_mac")),
                    "status": imported_status,
                    "original_status": status,
                    "assignee": assignee,
                    "assignee_text": assignee_text,
                    "department_name": _text(raw.get("department")),
                    "location_name": _text(raw.get("location")),
                    "purchase_date": purchase_date,
                    "purchase_cost": purchase_cost,
                    "warranty_expires_at": warranty,
                    "notes": _text(raw.get("notes")),
                    "system_code": system_code,
                    "kingdee_code": kingdee_code,
                },
                "preview": {
                    "asset_tag": existing.asset_tag if existing else "自动生成",
                    "name": _asset_display_name(
                        raw.get("brand"),
                        raw.get("model_name"),
                        category_name,
                    ),
                    "category": category_name,
                    "class_type": (
                        dict(AssetCategory.ClassType.choices).get(class_type, "")
                        if class_type
                        else _text(raw.get("class_type"))
                    ),
                    "brand_model": " ".join(
                        value
                        for value in [_text(raw.get("brand")), _text(raw.get("model_name"))]
                        if value
                    ),
                    "assignee": _text(raw.get("assignee")),
                    "status": status_label or "在库",
                },
            }
        )
    if not rows:
        raise AssetImportError("Excel 中没有可导入的数据行。")
    return rows


def summarize_import(rows):
    return {
        "total": len(rows),
        "create": sum(row["action"] == "create" for row in rows),
        "update": sum(row["action"] == "update" for row in rows),
        "invalid": sum(bool(row["errors"]) for row in rows),
        "warning": sum(bool(row["warnings"]) for row in rows),
        "rows": [
            {
                "row_number": row["row_number"],
                "action": row["action"],
                "errors": row["errors"],
                "warnings": row["warnings"],
                **row["preview"],
            }
            for row in rows[:100]
        ],
        "truncated": len(rows) > 100,
    }


@transaction.atomic
def apply_asset_import(rows, actor):
    invalid = [row for row in rows if row["errors"]]
    if invalid:
        raise AssetImportError("文件中仍有错误，请修正后重新上传。")
    created = updated = 0
    for row in rows:
        data = row["data"]
        category_missing = data.pop("category_missing")
        category_name_val = data.pop("category_name")
        class_type = data.pop("class_type")
        class_type_explicit = data.pop("class_type_explicit")
        category = AssetCategory.objects.filter(name__iexact=category_name_val).first()
        if category is None:
            category = AssetCategory.objects.create(
                name=category_name_val,
                code=(
                    "UC"
                    if category_missing
                    else _generated_category_code(category_name_val)
                ),
                class_type=class_type,
            )
        elif class_type_explicit and category.class_type != class_type:
            category.class_type = class_type
            category.save(update_fields=["class_type", "updated_at"])
        department_name = data.pop("department_name")
        department = None
        if department_name:
            department = Department.objects.filter(name__iexact=department_name).first()
        location_name = data.pop("location_name")
        location = None
        if location_name:
            location, _ = Location.objects.get_or_create(
                name=location_name,
                defaults={
                    "code": _generated_code("LOC", location_name),
                    "kind": Location.Kind.OFFICE,
                },
            )
        system_code = data.pop("system_code")
        kingdee_code = data.pop("kingdee_code")
        assignee = data.pop("assignee")
        assignee_text = data.pop("assignee_text")
        original_status = data.pop("original_status")
        data.pop("asset_tag")
        source_asset_tag = data.pop("source_asset_tag")
        source_name = data.pop("source_name")
        existing_id = data.pop("existing_id")
        defaults = {
            **data,
            "category": category,
            "assigned_to": assignee,
            "custodian_department": department,
            "current_location": location,
            "kingdee_code": kingdee_code,
        }
        asset = Asset.objects.filter(pk=existing_id).first() if existing_id else None
        tag = asset.asset_tag if asset else generate_asset_tag(category)
        before_status = asset.status if asset else ""
        if asset and category_missing:
            defaults["category"] = asset.category
        if asset and assignee_text and assignee is None:
            defaults["assigned_to"] = asset.assigned_to
            defaults["status"] = asset.status
        import_metadata = {
            "system_code": system_code,
            "import_warnings": row["warnings"],
            "import_original_assignee": assignee_text,
            "import_original_status": original_status,
            "import_original_asset_tag": source_asset_tag,
            "import_original_name": source_name,
        }
        if asset:
            for field, value in defaults.items():
                setattr(asset, field, value)
            asset.custom_data = {
                **asset.custom_data,
                **import_metadata,
            }
            asset.save()
            updated += 1
            action = AssetEvent.Action.UPDATED
            note = f"Excel 批量导入更新（第 {row['row_number']} 行）"
        else:
            asset = Asset.objects.create(
                asset_tag=tag,
                custom_data={
                    **import_metadata,
                },
                **defaults,
            )
            created += 1
            action = AssetEvent.Action.CREATED
            note = f"Excel 批量导入登记（第 {row['row_number']} 行）"
        AssetEvent.objects.create(
            asset=asset,
            action=action,
            from_status=before_status,
            to_status=asset.status,
            to_user=asset.assigned_to,
            to_location=asset.current_location,
            actor=actor,
            notes=note,
            metadata={"source": "excel_import", "row_number": row["row_number"]},
        )
    return {
        "created": created,
        "updated": updated,
        "total": created + updated,
        "warning": sum(bool(row["warnings"]) for row in rows),
    }


INVENTORY_HEADER_ALIASES = {
    "sku": {"物品编码", "库存编码", "SKU", "sku"},
    "name": {"物品名称", "名称", "库存名称"},
    "kind": {"物品分类", "库存分类", "类型", "分类"},
    "brand": {"品牌"},
    "model_name": {"型号", "规格型号"},
    "quantity": {"数量", "当前库存", "库存数量"},
    "unit": {"单位", "计量单位"},
    "unit_price": {"单价", "含税单价", "采购单价"},
    "minimum_quantity": {"保障数量", "最低库存", "预警数量", "库存下限"},
    "purchase_channel": {"采购途径", "采购渠道"},
    "location": {"存放地点", "存放位置", "位置", "地点"},
    "notes": {"备注", "说明"},
}

INVENTORY_KIND_ALIASES = {
    "配件": InventoryItem.Kind.ACCESSORY,
    "accessory": InventoryItem.Kind.ACCESSORY,
    "耗材": InventoryItem.Kind.CONSUMABLE,
    "consumable": InventoryItem.Kind.CONSUMABLE,
    "软件许可": InventoryItem.Kind.LICENSE,
    "软件": InventoryItem.Kind.LICENSE,
    "许可": InventoryItem.Kind.LICENSE,
    "license": InventoryItem.Kind.LICENSE,
    "其他": InventoryItem.Kind.OTHER,
    "other": InventoryItem.Kind.OTHER,
}

INVENTORY_PURCHASE_CHANNEL_ALIASES = {
    "合作供应商": InventoryItem.PurchaseChannel.SUPPLIER,
    "供应商": InventoryItem.PurchaseChannel.SUPPLIER,
    "supplier": InventoryItem.PurchaseChannel.SUPPLIER,
    "电商": InventoryItem.PurchaseChannel.ECOMMERCE,
    "电商平台": InventoryItem.PurchaseChannel.ECOMMERCE,
    "ecommerce": InventoryItem.PurchaseChannel.ECOMMERCE,
    "其他": InventoryItem.PurchaseChannel.OTHER,
    "other": InventoryItem.PurchaseChannel.OTHER,
}


def _inventory_header_map(values):
    normalized_aliases = {
        field: {_normalize_header(alias) for alias in aliases}
        for field, aliases in INVENTORY_HEADER_ALIASES.items()
    }
    result = {}
    for index, value in enumerate(values):
        normalized = _normalize_header(value)
        for field, aliases in normalized_aliases.items():
            if normalized in aliases:
                result[field] = index
                break
    return result


def _find_inventory_header_row(sheet):
    last_row = min(sheet.max_row or 20, 20)
    for row_number, row in enumerate(
        sheet.iter_rows(min_row=1, max_row=last_row, values_only=True),
        start=1,
    ):
        mapping = _inventory_header_map(row)
        if "name" in mapping and "quantity" in mapping:
            return row_number, mapping
    raise InventoryImportError(
        "没有找到库存表头，请保留“物品名称”和“数量”两列。"
    )


def _parse_inventory_integer(value, label, required=False):
    if value in (None, ""):
        if required:
            raise ValueError(f"{label}不能为空")
        return 0
    try:
        parsed = Decimal(_text(value).replace(",", ""))
    except InvalidOperation as exc:
        raise ValueError(f"{label}必须是整数") from exc
    if parsed < 0 or parsed != parsed.to_integral_value():
        raise ValueError(f"{label}必须是不小于 0 的整数")
    return int(parsed)


def _generate_inventory_sku():
    prefix = f"INV-{date.today().year}-"
    used = set(
        InventoryItem.objects.filter(sku__startswith=prefix).values_list("sku", flat=True)
    )
    sequence = 1
    while f"{prefix}{sequence:03d}" in used:
        sequence += 1
    return f"{prefix}{sequence:03d}"


def parse_inventory_workbook(file_obj):
    if getattr(file_obj, "size", 0) > MAX_IMPORT_BYTES:
        raise InventoryImportError("文件不能超过 5MB。")
    try:
        workbook = load_workbook(file_obj, read_only=True, data_only=True)
    except Exception as exc:
        raise InventoryImportError("无法读取 Excel，请上传有效的 .xlsx 文件。") from exc
    sheet = workbook.active
    header_row, mapping = _find_inventory_header_row(sheet)
    rows = []
    seen_skus = set()
    seen_existing_ids = set()
    for row_number, row in enumerate(
        sheet.iter_rows(min_row=header_row + 1, values_only=True),
        start=header_row + 1,
    ):
        if not any(value not in (None, "") for value in row):
            continue
        if len(rows) >= MAX_IMPORT_ROWS:
            raise InventoryImportError(f"单次最多导入 {MAX_IMPORT_ROWS} 行。")

        raw = {field: _cell(row, mapping, field) for field in mapping}
        errors = []
        warnings = []
        sku = _text(raw.get("sku"))
        name = _text(raw.get("name"))
        if not name:
            errors.append("物品名称不能为空")
        if sku:
            normalized_sku = sku.casefold()
            if normalized_sku in seen_skus:
                errors.append(f"物品编码“{sku}”在文件中重复")
            seen_skus.add(normalized_sku)
        else:
            warnings.append("物品编码为空，将由系统自动生成")

        kind_text = _text(raw.get("kind"))
        kind = INVENTORY_KIND_ALIASES.get(kind_text.casefold()) if kind_text else None
        if not kind:
            if kind_text:
                errors.append(f"无法识别物品分类“{kind_text}”")
            else:
                kind = InventoryItem.Kind.OTHER
                warnings.append("物品分类为空，将按“其他”导入")

        try:
            quantity = _parse_inventory_integer(raw.get("quantity"), "数量", required=True)
        except ValueError as exc:
            quantity = 0
            errors.append(str(exc))
        try:
            minimum_quantity = _parse_inventory_integer(
                raw.get("minimum_quantity"), "最低库存"
            )
        except ValueError as exc:
            minimum_quantity = 0
            errors.append(str(exc))
        try:
            unit_price = _parse_decimal(raw.get("unit_price"))
        except ValueError as exc:
            unit_price = None
            errors.append(str(exc).replace("金额", "单价"))

        purchase_channel_text = _text(raw.get("purchase_channel"))
        purchase_channel = (
            INVENTORY_PURCHASE_CHANNEL_ALIASES.get(purchase_channel_text.casefold())
            if purchase_channel_text
            else ""
        )
        if purchase_channel_text and not purchase_channel:
            errors.append(f"无法识别采购途径“{purchase_channel_text}”")

        existing = InventoryItem.objects.filter(sku__iexact=sku).first() if sku else None
        if existing and existing.pk in seen_existing_ids:
            errors.append(f"同一库存品 {existing.sku} 在文件中出现多次")
        if existing:
            seen_existing_ids.add(existing.pk)

        unit = _text(raw.get("unit")) or "个"
        location_name = _text(raw.get("location"))
        rows.append(
            {
                "row_number": row_number,
                "action": "update" if existing else "create",
                "errors": errors,
                "warnings": warnings,
                "data": {
                    "existing_id": existing.pk if existing else None,
                    "sku": sku,
                    "name": name,
                    "kind": kind,
                    "brand": _text(raw.get("brand")),
                    "model_name": _text(raw.get("model_name")),
                    "quantity": quantity,
                    "unit": unit,
                    "unit_price": unit_price,
                    "purchase_channel": purchase_channel,
                    "purchase_channel_present": "purchase_channel" in mapping,
                    "minimum_quantity": minimum_quantity,
                    "location_name": location_name,
                    "notes": _text(raw.get("notes")),
                },
                "preview": {
                    "sku": existing.sku if existing else (sku or "自动生成"),
                    "name": name,
                    "kind": dict(InventoryItem.Kind.choices).get(kind, kind_text or "其他"),
                    "brand_model": " ".join(
                        value
                        for value in [
                            _text(raw.get("brand")),
                            _text(raw.get("model_name")),
                        ]
                        if value
                    ),
                    "quantity": quantity,
                    "unit": unit,
                    "unit_price": str(unit_price) if unit_price is not None else "",
                    "purchase_channel": dict(InventoryItem.PurchaseChannel.choices).get(
                        purchase_channel,
                        "未设置",
                    ),
                    "location": location_name,
                },
            }
        )
    if not rows:
        raise InventoryImportError("Excel 中没有可导入的数据行。")
    return rows


def summarize_inventory_import(rows):
    return {
        "total": len(rows),
        "create": sum(row["action"] == "create" for row in rows),
        "update": sum(row["action"] == "update" for row in rows),
        "invalid": sum(bool(row["errors"]) for row in rows),
        "warning": sum(bool(row["warnings"]) for row in rows),
        "rows": [
            {
                "row_number": row["row_number"],
                "action": row["action"],
                "errors": row["errors"],
                "warnings": row["warnings"],
                **row["preview"],
            }
            for row in rows[:100]
        ],
        "truncated": len(rows) > 100,
    }


@transaction.atomic
def apply_inventory_import(rows, actor):
    invalid = [row for row in rows if row["errors"]]
    if invalid:
        raise InventoryImportError("文件中仍有错误，请修正后重新上传。")

    created = updated = adjusted = 0
    for row in rows:
        data = dict(row["data"])
        existing_id = data.pop("existing_id")
        desired_quantity = data.pop("quantity")
        location_name = data.pop("location_name")
        purchase_channel_present = data.pop("purchase_channel_present")
        location = None
        if location_name:
            location, _ = Location.objects.get_or_create(
                name=location_name,
                defaults={
                    "code": _generated_code("LOC", location_name),
                    "kind": Location.Kind.WAREHOUSE,
                },
            )
        data["location"] = location

        item = (
            InventoryItem.objects.select_for_update().filter(pk=existing_id).first()
            if existing_id
            else None
        )
        if item:
            before_quantity = item.quantity
            data.pop("sku", None)
            if not purchase_channel_present:
                data.pop("purchase_channel", None)
            for field, value in data.items():
                setattr(item, field, value)
            item.save()
            updated += 1
        else:
            before_quantity = 0
            data["sku"] = data["sku"] or _generate_inventory_sku()
            item = InventoryItem.objects.create(quantity=0, **data)
            created += 1

        delta = desired_quantity - before_quantity
        if delta:
            action = (
                InventoryTransaction.Action.INBOUND
                if delta > 0
                else InventoryTransaction.Action.WRITEOFF
            )
            item.quantity = desired_quantity
            item.save(update_fields=["quantity", "updated_at"])
            InventoryTransaction.objects.create(
                item=item,
                action=action,
                quantity=abs(delta),
                balance_after=desired_quantity,
                actor=actor,
                notes=f"Excel 导入校准库存（第 {row['row_number']} 行）",
            )
            adjusted += 1

    return {
        "created": created,
        "updated": updated,
        "adjusted": adjusted,
        "total": created + updated,
        "warning": sum(bool(row["warnings"]) for row in rows),
    }
