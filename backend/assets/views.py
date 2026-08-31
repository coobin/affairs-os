from datetime import date, timedelta
from pathlib import Path
import hashlib
import logging
import re
import secrets
import uuid
from urllib.parse import quote, urlencode

from django.conf import settings
from django.contrib.auth import get_user_model
from django.core.cache import cache
from django.http import FileResponse, HttpResponse, HttpResponseRedirect, StreamingHttpResponse
import openpyxl
from django.db import DatabaseError, transaction
from django.db.models.deletion import ProtectedError
from django.db.models import Count, F, Max, Q, Sum
from django.db.models.functions import ExtractMonth
from django.utils import timezone
from django.utils.http import content_disposition_header
from rest_framework import filters, viewsets
from rest_framework.authtoken.models import Token
from rest_framework.decorators import action, api_view, permission_classes
from rest_framework.exceptions import MethodNotAllowed
from rest_framework.permissions import AllowAny, IsAuthenticated
from rest_framework.response import Response
from rest_framework.parsers import MultiPartParser
from rest_framework.pagination import PageNumberPagination
from rest_framework.views import APIView

from .imports import (
    AssetImportError,
    InventoryImportError,
    apply_asset_import,
    apply_inventory_import,
    parse_asset_workbook,
    parse_inventory_workbook,
    summarize_import,
    summarize_inventory_import,
)
from .contract_reminders import expiry_reminder_dates
from .models import (
    Asset,
    AssetCategory,
    AssetEvent,
    AssetImage,
    AssetManagerRole,
    AssetRequest,
    AssetStatus,
    AdministrativeExpense,
    Contract,
    ContractAttachment,
    ContractChange,
    ContractType,
    Department,
    EmailNotification,
    ExpenseCategory,
    InventoryItem,
    InventoryTransaction,
    Location,
    ModuleToggle,
    Office,
    OperationLog,
    PurchaseOrder,
    PurchaseRequest,
    Supplier,
    SupplierAttachment,
    StocktakeRecord,
    StocktakeTask,
    Vehicle,
    VehicleDispatch,
    VehicleExpense,
)
from .oidc import oauth, sync_oidc_user
from .middleware import mark_operation
from .nextcloud import NextcloudStorageError, storage as nextcloud_storage
from .notifications import (
    notify_inventory_transaction,
    notify_request_cancelled,
    notify_request_processed,
    notify_request_submitted,
    notify_purchase_request_submitted,
    notify_vehicle_dispatch_submitted,
)
from .permissions import (
    HIDDEN_SYSTEM_USERNAME,
    MANAGEMENT_MODULES,
    IsModuleManager,
    IsSuperAdministrator,
    is_hidden_superuser,
    user_can_manage,
    user_can_manage_requests,
)
from .serializers import (
    AssetActionSerializer,
    AssetEventSerializer,
    AssetImageSerializer,
    AssetListSerializer,
    AssetRequestSerializer,
    AssetSerializer,
    AssetStatusSerializer,
    AdministrativeExpenseSerializer,
    CategorySerializer,
    ContractSerializer,
    ContractAttachmentSerializer,
    ContractChangeSerializer,
    ContractTypeSerializer,
    DepartmentSerializer,
    EmailNotificationSerializer,
    InventoryActionSerializer,
    InventoryItemSerializer,
    LocationSerializer,
    LoginSerializer,
    OperationLogSerializer,
    OfficeSerializer,
    StocktakeTaskSerializer,
    ExpenseCategorySerializer,
    PurchaseOrderSerializer,
    PurchaseRequestSerializer,
    SupplierSerializer,
    SupplierAttachmentSerializer,
    UserOptionSerializer,
    VehicleDispatchSerializer,
    VehicleExpenseSerializer,
    VehicleSerializer,
)
from .services import perform_asset_action

User = get_user_model()
logger = logging.getLogger(__name__)


def contracts_visible_to(user, queryset=None):
    queryset = queryset if queryset is not None else Contract.objects.all()
    if is_hidden_superuser(user):
        return queryset
    return queryset.filter(owner=user)

ASSET_IMAGE_EXTENSIONS = {".jpg", ".jpeg", ".png", ".webp"}
CONTRACT_FILE_EXTENSIONS = {
    ".pdf", ".doc", ".docx", ".xls", ".xlsx", ".ppt", ".pptx",
    ".jpg", ".jpeg", ".png", ".webp", ".zip", ".rar", ".7z", ".ofd", ".wps",
}
ASSET_IMAGE_MAX_BYTES = 10 * 1024 * 1024
CONTRACT_FILE_MAX_BYTES = 100 * 1024 * 1024
SUPPLIER_FILE_EXTENSIONS = {".pdf", ".jpg", ".jpeg", ".png", ".webp"}
SUPPLIER_FILE_MAX_BYTES = 20 * 1024 * 1024


def _safe_path_part(value):
    return re.sub(r"[^0-9A-Za-z._-]+", "-", str(value)).strip("-.") or "record"


def _validate_upload(upload, allowed_extensions, max_bytes, label):
    if not upload:
        return f"请选择要上传的{label}。"
    original_name = Path(upload.name).name
    extension = Path(original_name).suffix.lower()
    if extension not in allowed_extensions:
        return f"不支持这种{label}格式。"
    if upload.size <= 0:
        return f"{label}内容为空。"
    if upload.size > max_bytes:
        return f"{label}不能超过 {max_bytes // 1024 // 1024}MB。"
    return ""


def _file_sha256(upload):
    digest = hashlib.sha256()
    for chunk in upload.chunks():
        digest.update(chunk)
    upload.seek(0)
    return digest.hexdigest()


def _remote_file_path(section, year, identifier, upload):
    extension = Path(upload.name).suffix.lower()
    filename = f"{uuid.uuid4().hex}{extension}"
    return (
        f"{settings.NEXTCLOUD_ROOT}/{section}/{year}/"
        f"{_safe_path_part(identifier)}/{filename}"
    )


def _stream_remote_file(record, as_attachment):
    remote_response = nextcloud_storage.download(record.remote_path)

    def iterator():
        try:
            yield from remote_response.iter_content(chunk_size=64 * 1024)
        finally:
            remote_response.close()

    response = StreamingHttpResponse(
        iterator(),
        content_type=record.content_type or "application/octet-stream",
    )
    response["Content-Length"] = str(record.size_bytes)
    response["Content-Disposition"] = content_disposition_header(
        as_attachment,
        record.original_name,
    )
    response["X-Content-Type-Options"] = "nosniff"
    return response


def _next_supplement_contract_no(parent, index):
    candidate = f"{parent.contract_no}-S{index:02d}"
    while Contract.objects.filter(contract_no=candidate).exists():
        index += 1
        candidate = f"{parent.contract_no}-S{index:02d}"
    return candidate


@api_view(["GET"])
@permission_classes([AllowAny])
def health(request):
    return Response({"status": "ok", "service": "affairs-os"})


class OIDCLoginView(APIView):
    permission_classes = [AllowAny]

    def get(self, request):
        if not settings.OIDC_CLIENT_SECRET:
            return Response({"message": "OIDC 客户端尚未配置。"}, status=503)
        client = oauth.create_client("authelia")
        return client.authorize_redirect(request, settings.OIDC_REDIRECT_URI)


class OIDCCallbackView(APIView):
    permission_classes = [AllowAny]

    def get(self, request):
        try:
            client = oauth.create_client("authelia")
            token_data = client.authorize_access_token(request)
            userinfo = token_data.get("userinfo") or {}
            if not userinfo.get("preferred_username"):
                userinfo = client.userinfo(token=token_data)
            user = sync_oidc_user(userinfo)
            token, _ = Token.objects.get_or_create(user=user)
            one_time_code = secrets.token_urlsafe(32)
            cache.set(f"oidc-login:{one_time_code}", token.key, timeout=60)
            query = urlencode({"oidc_code": one_time_code})
        except Exception:
            logger.exception("OIDC callback failed")
            query = urlencode({"oidc_error": "公司账号登录未完成，请重试。"})
        return HttpResponseRedirect(f"{settings.FRONTEND_URL.rstrip('/')}/?{query}")


class OIDCCompleteView(APIView):
    permission_classes = [AllowAny]

    def post(self, request):
        code = str(request.data.get("code") or "").strip()
        cache_key = f"oidc-login:{code}"
        token_key = cache.get(cache_key) if code else None
        if not token_key:
            return Response({"message": "登录凭证已失效，请重新登录。"}, status=400)
        cache.delete(cache_key)
        token = Token.objects.select_related("user").get(key=token_key)
        mark_operation(
            request,
            user=token.user,
            action="login",
            target_label=token.user.get_full_name() or token.user.username,
        )
        return Response({"token": token.key, "user": UserOptionSerializer(token.user).data})


class LocalLoginView(APIView):
    permission_classes = [AllowAny]

    def post(self, request):
        username = str(request.data.get("username") or "").strip()
        if username.casefold() != settings.LOCAL_LOGIN_USERNAME.casefold():
            return Response({"message": "账号或密码不正确。"}, status=400)

        serializer = LoginSerializer(data=request.data, context={"request": request})
        if not serializer.is_valid():
            return Response({"message": "账号或密码不正确。"}, status=400)

        user = serializer.validated_data["user"]
        token, _ = Token.objects.get_or_create(user=user)
        mark_operation(
            request,
            user=user,
            action="login",
            target_label=user.get_full_name() or user.username,
        )
        return Response({"token": token.key, "user": UserOptionSerializer(user).data})


class LogoutView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        Token.objects.filter(user=request.user).delete()
        return Response(status=204)


class MeView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        return Response(UserOptionSerializer(request.user).data)


class AssetViewSet(viewsets.ModelViewSet):
    serializer_class = AssetSerializer
    permission_classes = [IsModuleManager]
    management_module = "assets"
    filter_backends = [filters.OrderingFilter]
    ordering_fields = [
        "asset_tag",
        "name",
        "status",
        "purchase_date",
        "warranty_expires_at",
        "created_at",
    ]
    ordering = [
        F("purchase_date").desc(nulls_last=True),
        F("created_at").desc(),
    ]

    def get_serializer_class(self):
        if self.action == "list":
            return AssetListSerializer
        return AssetSerializer

    def get_queryset(self):
        queryset = Asset.objects.select_related(
            "category",
            "current_location",
            "assigned_to",
            "custodian_department",
        )
        if self.action in {"retrieve", "perform_action"}:
            queryset = queryset.prefetch_related(
                "events__actor",
                "events__from_user",
                "events__to_user",
                "events__from_location",
                "events__to_location",
                "images__uploaded_by",
            )
        query = self.request.query_params.get("q", "").strip()
        if query:
            queryset = queryset.filter(
                Q(asset_tag__icontains=query)
                | Q(kingdee_code__icontains=query)
                | Q(custom_data__system_code__icontains=query)
                | Q(name__icontains=query)
                | Q(serial_number__icontains=query)
                | Q(brand__icontains=query)
                | Q(model_name__icontains=query)
                | Q(assigned_to__first_name__icontains=query)
                | Q(assigned_to__last_name__icontains=query)
                | Q(assigned_to__username__icontains=query)
            )

        for field, parameter in [
            ("status", "status"),
            ("category_id", "category"),
            ("current_location_id", "location"),
            ("custodian_department_id", "department"),
        ]:
            value = self.request.query_params.get(parameter)
            if value:
                queryset = queryset.filter(**{field: value})
        class_type = self.request.query_params.get("class_type", "").strip().upper()
        if class_type in dict(AssetCategory.ClassType.choices):
            queryset = queryset.filter(category__class_type=class_type)
        overdue = self.request.query_params.get("overdue", "").strip().lower()
        if overdue in {"1", "true", "yes"}:
            queryset = queryset.filter(
                status=Asset.Status.LOANED,
                expected_return_at__lt=timezone.localdate(),
            )
        return queryset

    def destroy(self, request, *args, **kwargs):
        if not is_hidden_superuser(request.user):
            raise MethodNotAllowed(
                "DELETE",
                detail="资产不能直接删除，请通过退役或处置保留完整历史。",
            )
        asset = self.get_object()
        try:
            for image in asset.images.all():
                try:
                    nextcloud_storage.delete(image.remote_path)
                except NextcloudStorageError as exc:
                    return Response({"message": str(exc)}, status=503)
            asset.events.all().delete()
            asset.delete()
        except ProtectedError:
            return Response(
                {"message": "该资产已被领用申请等业务记录引用，无法删除。"},
                status=400,
            )
        return Response(status=204)

    @action(detail=False, methods=["GET"])
    def export(self, request):
        queryset = self.get_queryset().order_by(*self.ordering)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "资产列表"

        headers = ["资产编号", "金蝶编码", "显示名称", "资产分类", "资产类型", "品牌", "型号", "序列号", "状态", "责任人", "部门", "所在地点", "入账时间", "保修到期"]
        ws.append(headers)

        for asset in queryset:
            class_type_label = asset.category.get_class_type_display() if asset.category else ""
            status_label = asset.get_status_display()
            assigned = f"{asset.assigned_to.first_name}{asset.assigned_to.last_name}" or asset.assigned_to.username if asset.assigned_to else ""
            department = asset.custodian_department.name if asset.custodian_department else ""
            location = asset.current_location.name if asset.current_location else ""
            category_name = asset.category.name if asset.category else ""

            ws.append([
                asset.asset_tag,
                asset.kingdee_code or "",
                asset.name,
                class_type_label,
                category_name,
                asset.brand or "",
                asset.model_name or "",
                asset.serial_number or "",
                status_label,
                assigned,
                department,
                location,
                asset.purchase_date.strftime("%Y-%m-%d") if asset.purchase_date else "",
                asset.warranty_expires_at.strftime("%Y-%m-%d") if asset.warranty_expires_at else "",
            ])

        response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response["Content-Disposition"] = f'attachment; filename="assets_export_{timezone.now().strftime("%Y%m%d%H%M%S")}.xlsx"'
        wb.save(response)
        return response


    @action(
        detail=False,
        methods=["post"],
        url_path="import",
        parser_classes=[MultiPartParser],
    )
    def import_excel(self, request):
        upload = request.FILES.get("file")
        if not upload:
            return Response({"message": "请选择 Excel 文件。"}, status=400)
        if not upload.name.lower().endswith(".xlsx"):
            return Response({"message": "只支持 .xlsx 文件。"}, status=400)
        try:
            rows = parse_asset_workbook(upload)
            preview = summarize_import(rows)
            if request.data.get("commit") in {"true", "1", True}:
                result = apply_asset_import(rows, request.user)
                return Response({"message": "资产导入完成。", **result})
            return Response(preview)
        except AssetImportError as exc:
            return Response({"message": str(exc)}, status=400)
        except DatabaseError:
            logger.exception("资产 Excel 写入数据库失败")
            return Response(
                {"message": "导入写入失败，请检查字段内容是否过长或存在重复数据。"},
                status=400,
            )

    @action(detail=False, methods=["get"], url_path="import-template")
    def import_template(self, request):
        template = Path(settings.BASE_DIR) / "assets" / "data" / "资产导入模板.xlsx"
        if not template.exists():
            return Response({"message": "导入模板暂不可用。"}, status=503)
        return FileResponse(
            template.open("rb"),
            as_attachment=True,
            filename="资产导入模板.xlsx",
        )

    @action(detail=True, methods=["post"], url_path="import-issues-resolved")
    def resolve_import_issues(self, request, pk=None):
        asset = self.get_object()
        custom_data = dict(asset.custom_data)
        warnings = custom_data.pop("import_warnings", [])
        custom_data.pop("import_original_assignee", None)
        custom_data.pop("import_original_status", None)
        asset.custom_data = custom_data
        asset.save(update_fields=["custom_data", "updated_at"])
        if warnings:
            AssetEvent.objects.create(
                asset=asset,
                action=AssetEvent.Action.UPDATED,
                from_status=asset.status,
                to_status=asset.status,
                from_user=asset.assigned_to,
                to_user=asset.assigned_to,
                from_location=asset.current_location,
                to_location=asset.current_location,
                actor=request.user,
                notes="确认 Excel 导入待完善资料已处理",
                metadata={"resolved_import_warnings": warnings},
            )
        return Response(AssetSerializer(asset, context={"request": request}).data)

    @action(
        detail=True,
        methods=["get", "post"],
        url_path="images",
        parser_classes=[MultiPartParser],
    )
    def images(self, request, pk=None):
        asset = self.get_object()
        if request.method == "GET":
            return Response(
                AssetImageSerializer(asset.images.select_related("uploaded_by"), many=True).data
            )

        if asset.images.count() >= 10:
            return Response({"message": "每件资产最多上传 10 张图片。"}, status=400)
        upload = request.FILES.get("file")
        error = _validate_upload(
            upload,
            ASSET_IMAGE_EXTENSIONS,
            ASSET_IMAGE_MAX_BYTES,
            "图片",
        )
        if error:
            return Response({"message": error}, status=400)
        if not (upload.content_type or "").startswith("image/"):
            return Response({"message": "文件内容不是可识别的图片。"}, status=400)

        remote_path = _remote_file_path(
            "assets",
            asset.created_at.year,
            asset.asset_tag,
            upload,
        )
        try:
            nextcloud_storage.upload(upload, remote_path)
            try:
                current_max = asset.images.aggregate(value=Max("sort_order"))["value"]
                image = AssetImage.objects.create(
                    asset=asset,
                    remote_path=remote_path,
                    original_name=Path(upload.name).name[:255],
                    content_type=(upload.content_type or "")[:120],
                    size_bytes=upload.size,
                    sha256=_file_sha256(upload),
                    uploaded_by=request.user,
                    is_cover=not asset.images.exists(),
                    sort_order=(current_max if current_max is not None else -1) + 1,
                )
            except Exception:
                nextcloud_storage.delete(remote_path)
                raise
        except NextcloudStorageError as exc:
            return Response({"message": str(exc)}, status=503)

        return Response(AssetImageSerializer(image).data, status=201)

    @action(
        detail=True,
        methods=["get", "delete", "patch"],
        url_path=r"images/(?P<image_id>\d+)",
    )
    def image_file(self, request, pk=None, image_id=None):
        asset = self.get_object()
        image = asset.images.filter(pk=image_id).first()
        if not image:
            return Response({"message": "没有找到这张资产图片。"}, status=404)

        if request.method == "GET":
            try:
                return _stream_remote_file(image, as_attachment=False)
            except NextcloudStorageError as exc:
                return Response({"message": str(exc)}, status=503)

        if request.method == "PATCH":
            asset.images.exclude(pk=image.pk).update(is_cover=False)
            image.is_cover = True
            image.save(update_fields=["is_cover", "updated_at"])
            return Response(AssetImageSerializer(image).data)

        try:
            nextcloud_storage.delete(image.remote_path)
        except NextcloudStorageError as exc:
            return Response({"message": str(exc)}, status=503)
        was_cover = image.is_cover
        image.delete()
        if was_cover:
            replacement = asset.images.order_by("sort_order", "created_at").first()
            if replacement:
                replacement.is_cover = True
                replacement.save(update_fields=["is_cover", "updated_at"])
        return Response(status=204)

    @action(detail=True, methods=["post"], url_path="actions")
    def perform_action(self, request, pk=None):
        asset = self.get_object()
        serializer = AssetActionSerializer(
            data=request.data,
            context={"request": request, "asset": asset},
        )
        serializer.is_valid(raise_exception=True)
        result = serializer.save()
        return Response(
            {
                "asset": AssetSerializer(result.asset, context={"request": request}).data,
                "event": AssetEventSerializer(result.event).data,
            }
        )


class DashboardView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        queryset = Asset.objects.all() if user_can_manage(request.user, "assets") else Asset.objects.none()
        status_counts = {
            row["status"]: row["total"]
            for row in queryset.values("status").annotate(total=Count("id"))
        }
        today = timezone.localdate()
        overdue = queryset.filter(
            status=Asset.Status.LOANED,
            expected_return_at__lt=today,
        ).count()
        recent_events = AssetEvent.objects.select_related("asset", "actor") if user_can_manage(request.user, "assets") else AssetEvent.objects.none()
        status_options = AssetStatus.objects.filter(
            Q(is_active=True) | Q(code__in=status_counts.keys())
        ).distinct()
        expiring_contracts = Contract.objects.none()
        if user_can_manage(request.user, "contracts"):
            expiring_contracts = contracts_visible_to(
                request.user,
                Contract.objects.filter(
                    status__in=[Contract.Status.ACTIVE, Contract.Status.EXPIRED],
                    renewal_contracts__isnull=True,
                    supplement_of__isnull=True,
                ).filter(
                    Q(end_date__lte=today)
                    | Q(end_date__in=expiry_reminder_dates(today))
                ),
            ).distinct()
        vehicle_insurance_due = Vehicle.objects.none()
        if user_can_manage(request.user, "vehicles"):
            reminder_dates = expiry_reminder_dates(today)
            vehicle_insurance_due = Vehicle.objects.filter(
                Q(insurance_expires_at__lte=today)
                | Q(inspection_expires_at__lte=today)
                | Q(insurance_expires_at__in=reminder_dates)
                | Q(inspection_expires_at__in=reminder_dates)
            ).exclude(status=Vehicle.Status.RETIRED)
        return Response(
            {
                "summary": {
                    "total": queryset.count(),
                    "available": status_counts.get(Asset.Status.AVAILABLE, 0),
                    "assigned": status_counts.get(Asset.Status.ASSIGNED, 0)
                    + status_counts.get(Asset.Status.LOANED, 0),
                },
                "tasks": {
                    "overdue_loans": overdue,
                },
                "admin_tasks": {
                    "pending_vehicle_dispatches": VehicleDispatch.objects.filter(status=VehicleDispatch.Status.PENDING).count() if user_can_manage(request.user, "vehicles") else 0,
                    "vehicle_insurance_due": vehicle_insurance_due.count(),
                    "pending_purchase_requests": PurchaseRequest.objects.filter(status=PurchaseRequest.Status.PENDING).count() if user_can_manage(request.user, "procurement") else 0,
                    "expiring_contracts": expiring_contracts.count(),
                },
                "status_distribution": [
                    {
                        "status": value,
                        "label": label,
                        "total": status_counts.get(value, 0),
                    }
                    for value, label in status_options.values_list("code", "name")
                    if status_counts.get(value, 0)
                ],
                "recent_events": AssetEventSerializer(recent_events[:8], many=True).data,
                "generated_at": timezone.now(),
            }
        )


class LookupView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        enabled_modules = list(
            ModuleToggle.objects.filter(is_enabled=True)
            .values_list("code", flat=True)
        )
        users = User.objects.filter(is_active=True).exclude(
            username__iexact=HIDDEN_SYSTEM_USERNAME
        ).select_related(
            "employee_profile__department"
        )
        if not any(user_can_manage(request.user, scope) for scope in ("assets", "inventory", "vehicles", "procurement", "contracts", "offices", "expenses", "settings")):
            users = users.filter(pk=request.user.pk)
        return Response(
            {
                "enabled_modules": enabled_modules,
                "users": UserOptionSerializer(users, many=True).data,
                "departments": DepartmentSerializer(
                    Department.objects.filter(is_active=True), many=True
                ).data,
                "locations": LocationSerializer(
                    Location.objects.filter(is_active=True), many=True
                ).data,
                "categories": CategorySerializer(
                    AssetCategory.objects.filter(is_active=True), many=True
                ).data,
                "statuses": [
                    {"value": value, "label": label}
                    for value, label in AssetStatus.objects.filter(is_active=True).values_list(
                        "code", "name"
                    )
                ],
            }
        )


class ManagerSettingsView(APIView):
    permission_classes = [IsSuperAdministrator]

    def get(self, request):
        users = User.objects.filter(is_active=True).exclude(
            username__iexact=HIDDEN_SYSTEM_USERNAME
        ).select_related("employee_profile__department", "asset_manager_role")
        enabled = set(
            ModuleToggle.objects.filter(is_enabled=True)
            .values_list("code", flat=True)
        )
        return Response(
            {
                "modules": [
                    {"value": value, "label": label}
                    for value, label in MANAGEMENT_MODULES
                    if value in enabled
                ],
                "users": UserOptionSerializer(users, many=True).data,
            }
        )

    @transaction.atomic
    def patch(self, request):
        user_id = request.data.get("user_id")
        scopes = request.data.get("scopes", [])
        allowed = {value for value, _ in MANAGEMENT_MODULES}
        if not isinstance(scopes, list) or any(scope not in allowed for scope in scopes):
            return Response({"message": "管理板块设置不正确。"}, status=400)
        user = User.objects.filter(
            pk=user_id,
            is_active=True,
            is_superuser=False,
        ).exclude(username__iexact=HIDDEN_SYSTEM_USERNAME).first()
        if not user:
            return Response({"message": "找不到这个用户。"}, status=404)
        if scopes:
            AssetManagerRole.objects.update_or_create(
                user=user,
                defaults={"scopes": list(dict.fromkeys(scopes))},
            )
            user.is_staff = True
        else:
            AssetManagerRole.objects.filter(user=user).delete()
            user.is_staff = False
        user.save(update_fields=["is_staff"])
        user = User.objects.select_related(
            "employee_profile__department", "asset_manager_role"
        ).get(pk=user.pk)
        return Response(UserOptionSerializer(user).data)


class ModuleSettingsView(APIView):
    permission_classes = [IsSuperAdministrator]

    def get(self, request):
        toggles = {
            item.code: item
            for item in ModuleToggle.objects.all()
        }
        return Response(
            [
                {
                    "code": code,
                    "label": label,
                    "enabled": toggles[code].is_enabled if code in toggles else True,
                }
                for code, label in MANAGEMENT_MODULES
                if code != "settings"
            ]
        )

    def patch(self, request):
        code = str(request.data.get("code") or "").strip()
        enabled = request.data.get("enabled")
        allowed = {value for value, _ in MANAGEMENT_MODULES} - {"settings"}
        if code not in allowed or not isinstance(enabled, bool):
            return Response({"message": "模块设置不正确。"}, status=400)
        ModuleToggle.objects.update_or_create(
            code=code,
            defaults={"is_enabled": enabled},
        )
        return Response(
            {
                "code": code,
                "label": dict(MANAGEMENT_MODULES)[code],
                "enabled": enabled,
            }
        )


class OperationLogPagination(PageNumberPagination):
    page_size = 50
    page_size_query_param = "page_size"
    max_page_size = 200


class OperationLogView(APIView):
    permission_classes = [IsSuperAdministrator]

    def get(self, request):
        queryset = OperationLog.objects.select_related("user")
        username = request.query_params.get("username", "").strip()
        module = request.query_params.get("module", "").strip()
        action = request.query_params.get("action", "").strip()
        result = request.query_params.get("result", "").strip()
        date_from = request.query_params.get("date_from", "").strip()
        date_to = request.query_params.get("date_to", "").strip()
        query = request.query_params.get("q", "").strip()

        if username:
            queryset = queryset.filter(username=username)
        if module:
            queryset = queryset.filter(module=module)
        if action:
            queryset = queryset.filter(action=action)
        if result == "success":
            queryset = queryset.filter(succeeded=True)
        elif result == "failed":
            queryset = queryset.filter(succeeded=False)
        try:
            if date_from:
                queryset = queryset.filter(occurred_at__date__gte=date.fromisoformat(date_from))
            if date_to:
                queryset = queryset.filter(occurred_at__date__lte=date.fromisoformat(date_to))
        except ValueError:
            return Response({"message": "日志日期格式不正确。"}, status=400)
        if query:
            queryset = queryset.filter(
                Q(username__icontains=query)
                | Q(display_name__icontains=query)
                | Q(target_label__icontains=query)
                | Q(path__icontains=query)
                | Q(action_label__icontains=query)
            )

        paginator = OperationLogPagination()
        page = paginator.paginate_queryset(queryset, request, view=self)
        response = paginator.get_paginated_response(
            OperationLogSerializer(page, many=True).data
        )
        response.data["filters"] = {
            "users": [
                {"username": row["username"], "display_name": row["display_name"]}
                for row in OperationLog.objects.values("username", "display_name")
                .order_by("display_name", "username")
                .distinct()
            ],
            "modules": [
                {"value": row["module"], "label": row["module_label"]}
                for row in OperationLog.objects.values("module", "module_label")
                .order_by("module_label", "module")
                .distinct()
            ],
            "actions": [
                {"value": row["action"], "label": row["action_label"]}
                for row in OperationLog.objects.values("action", "action_label")
                .order_by("action_label", "action")
                .distinct()
            ],
        }
        return response


class EmailNotificationPagination(PageNumberPagination):
    page_size = 50
    page_size_query_param = "page_size"
    max_page_size = 200


class EmailNotificationView(APIView):
    permission_classes = [IsSuperAdministrator]

    def get(self, request):
        queryset = EmailNotification.objects.select_related("recipient_user")
        status_value = request.query_params.get("status", "").strip()
        event_type = request.query_params.get("event_type", "").strip()
        date_from = request.query_params.get("date_from", "").strip()
        date_to = request.query_params.get("date_to", "").strip()
        query = request.query_params.get("q", "").strip()

        if status_value:
            queryset = queryset.filter(status=status_value)
        if event_type:
            queryset = queryset.filter(event_type=event_type)
        try:
            if date_from:
                queryset = queryset.filter(created_at__date__gte=date.fromisoformat(date_from))
            if date_to:
                queryset = queryset.filter(created_at__date__lte=date.fromisoformat(date_to))
        except ValueError:
            return Response({"message": "邮件记录日期格式不正确。"}, status=400)
        if query:
            queryset = queryset.filter(
                Q(recipient_email__icontains=query)
                | Q(recipient_user__username__icontains=query)
                | Q(recipient_user__first_name__icontains=query)
                | Q(recipient_user__last_name__icontains=query)
                | Q(subject__icontains=query)
                | Q(body__icontains=query)
            )

        paginator = EmailNotificationPagination()
        page = paginator.paginate_queryset(queryset, request, view=self)
        response = paginator.get_paginated_response(
            EmailNotificationSerializer(page, many=True).data
        )
        response.data["filters"] = {
            "statuses": [
                {"value": value, "label": label}
                for value, label in EmailNotification.Status.choices
            ],
            "event_types": [
                {
                    "value": value,
                    "label": EmailNotificationSerializer().get_event_type_label(
                        EmailNotification(event_type=value)
                    ),
                }
                for value in queryset.order_by()
                .values_list("event_type", flat=True)
                .distinct()
                .order_by("event_type")
            ],
        }
        return response


class AssetRequestViewSet(viewsets.ModelViewSet):
    serializer_class = AssetRequestSerializer
    permission_classes = [IsAuthenticated]
    pagination_class = None
    http_method_names = ["get", "post", "head", "options"]

    def get_queryset(self):
        queryset = AssetRequest.objects.select_related(
            "requester",
            "requester__employee_profile__department",
            "assigned_asset",
            "inventory_item",
            "issued_inventory_transaction",
            "handled_by",
        )
        mine_only = self.request.query_params.get("mine", "").lower() in {"1", "true", "yes"}
        if mine_only or not user_can_manage_requests(self.request.user):
            queryset = queryset.filter(requester=self.request.user)
        status_value = self.request.query_params.get("status")
        if status_value:
            queryset = queryset.filter(status=status_value)
        return queryset

    def perform_create(self, serializer):
        asset_request = serializer.save(requester=self.request.user)
        notify_request_submitted(asset_request)

    @action(detail=False, methods=["get"], url_path="my-loaned-assets")
    def my_loaned_assets(self, request):
        assets = Asset.objects.filter(
            status=Asset.Status.LOANED,
            assigned_to=request.user,
        ).select_related(
            "category",
            "current_location",
            "assigned_to",
            "custodian_department",
        ).order_by("expected_return_at", "asset_tag")
        return Response(AssetListSerializer(assets, many=True).data)

    @action(detail=False, methods=["get"], url_path="device-options")
    def device_options(self, request):
        asset_rows = (
            Asset.objects.filter(
                status=Asset.Status.AVAILABLE,
                is_requestable=True,
            )
            .values("category_id", "category__name")
            .annotate(available_count=Count("id"))
            .order_by("category__name")
        )
        options = [
            {
                "key": f"asset:{row['category_id']}",
                "item_type": AssetRequest.ItemType.ASSET,
                "item_id": row["category_id"],
                "name": row["category__name"],
                "description": "资产",
                "available_count": row["available_count"],
                "unit": "件",
            }
            for row in asset_rows
        ]
        options.extend(
            {
                "key": f"inventory:{item.pk}",
                "item_type": AssetRequest.ItemType.INVENTORY,
                "item_id": item.pk,
                "name": item.name,
                "description": " · ".join(part for part in ["库存", item.brand, item.model_name, item.sku] if part),
                "available_count": item.quantity,
                "unit": item.unit,
            }
            for item in InventoryItem.objects.filter(is_active=True, quantity__gt=0).order_by("name", "sku")
        )
        return Response(options)

    @action(detail=True, methods=["get"])
    def candidates(self, request, pk=None):
        if not user_can_manage_requests(request.user):
            return Response({"message": "只有资产或库存管理员可以分配具体设备。"}, status=403)
        asset_request = self.get_object()
        if asset_request.requested_item_type == AssetRequest.ItemType.INVENTORY:
            return Response({"message": "库存领用申请无需选择具体资产。"}, status=400)
        assets = Asset.objects.filter(
            category__name=asset_request.requested_name,
            status=Asset.Status.AVAILABLE,
            is_requestable=True,
        ).select_related("category", "current_location", "assigned_to", "custodian_department")
        return Response(AssetListSerializer(assets, many=True).data)

    @action(detail=True, methods=["post"])
    @transaction.atomic
    def fulfill(self, request, pk=None):
        if not user_can_manage_requests(request.user):
            return Response({"message": "只有资产或库存管理员可以分配具体设备。"}, status=403)
        asset_request = AssetRequest.objects.select_for_update().get(pk=self.get_object().pk)
        if asset_request.status != AssetRequest.Status.PENDING:
            return Response({"message": "这条申请已经处理。"}, status=400)

        if asset_request.requested_item_type == AssetRequest.ItemType.INVENTORY:
            item = InventoryItem.objects.select_for_update().filter(
                pk=asset_request.inventory_item_id,
                is_active=True,
            ).first()
            if not item or item.quantity < asset_request.requested_quantity:
                return Response({"message": "库存数量不足，暂时不能完成发放。"}, status=400)
            item.quantity -= asset_request.requested_quantity
            item.save(update_fields=["quantity", "updated_at"])
            inventory_transaction = InventoryTransaction.objects.create(
                item=item,
                action=InventoryTransaction.Action.ISSUE,
                quantity=asset_request.requested_quantity,
                balance_after=item.quantity,
                recipient=asset_request.requester,
                actor=request.user,
                notes=str(request.data.get("manager_notes") or "申请领用发放").strip(),
            )
            asset_request.status = AssetRequest.Status.FULFILLED
            asset_request.issued_inventory_transaction = inventory_transaction
            asset_request.handled_by = request.user
            asset_request.handled_at = timezone.now()
            asset_request.manager_notes = str(request.data.get("manager_notes") or "").strip()
            asset_request.save()
            notify_request_processed(asset_request)
            return Response(self.get_serializer(asset_request).data)

        asset = Asset.objects.select_for_update().filter(pk=request.data.get("asset_id")).first()
        if not asset or asset.status != Asset.Status.AVAILABLE or not asset.is_requestable:
            return Response({"message": "选择的设备已不可分配。"}, status=400)
        if asset.category.name != asset_request.requested_name:
            return Response({"message": "请分配与申请类型一致的设备。"}, status=400)
        action_notes = str(request.data.get("manager_notes") or "").strip()
        if asset_request.request_type == AssetRequest.RequestType.LOAN:
            action_notes = action_notes or str(asset_request.reason or "").strip()
        action_result = perform_asset_action(
            asset=asset,
            action=asset_request.request_type,
            actor=request.user,
            target_user=asset_request.requester,
            expected_return_at=asset_request.expected_return_at,
            notes=action_notes,
            send_notification=False,
        )
        asset_request.status = AssetRequest.Status.FULFILLED
        asset_request.assigned_asset = action_result.asset
        asset_request.handled_by = request.user
        asset_request.handled_at = timezone.now()
        asset_request.manager_notes = str(request.data.get("manager_notes") or "").strip()
        asset_request.save()
        notify_request_processed(asset_request)
        return Response(self.get_serializer(asset_request).data)

    @action(detail=True, methods=["post"])
    @transaction.atomic
    def reject(self, request, pk=None):
        if not user_can_manage_requests(request.user):
            return Response({"message": "只有资产或库存管理员可以驳回申请。"}, status=403)
        asset_request = AssetRequest.objects.select_for_update().get(pk=self.get_object().pk)
        if asset_request.status != AssetRequest.Status.PENDING:
            return Response({"message": "这条申请已经处理。"}, status=400)
        asset_request.status = AssetRequest.Status.REJECTED
        asset_request.handled_by = request.user
        asset_request.handled_at = timezone.now()
        asset_request.manager_notes = str(request.data.get("manager_notes") or "").strip()
        asset_request.save()
        notify_request_processed(asset_request)
        return Response(self.get_serializer(asset_request).data)

    @action(detail=True, methods=["post"])
    @transaction.atomic
    def cancel(self, request, pk=None):
        asset_request = AssetRequest.objects.select_for_update().get(pk=self.get_object().pk)
        if asset_request.requester_id != request.user.id:
            return Response({"message": "只能取消自己的申请。"}, status=403)
        if asset_request.status != AssetRequest.Status.PENDING:
            return Response({"message": "已处理的申请不能取消。"}, status=400)
        asset_request.status = AssetRequest.Status.CANCELLED
        asset_request.handled_at = timezone.now()
        asset_request.save(update_fields=["status", "handled_at", "updated_at"])
        notify_request_cancelled(asset_request)
        return Response(self.get_serializer(asset_request).data)


class NoDeleteViewSet(viewsets.ModelViewSet):
    permission_classes = [IsModuleManager]
    pagination_class = None

    def destroy(self, request, *args, **kwargs):
        if not is_hidden_superuser(request.user):
            raise MethodNotAllowed("DELETE", detail="基础资料不能直接删除，请改为停用。")
        obj = self.get_object()
        try:
            obj.delete()
        except ProtectedError:
            return Response(
                {"message": "该基础资料仍被业务记录引用，无法删除，请先停用。"},
                status=400,
            )
        return Response(status=204)


class DepartmentViewSet(NoDeleteViewSet):
    management_module = "settings"
    queryset = Department.objects.all()
    serializer_class = DepartmentSerializer


class LocationViewSet(NoDeleteViewSet):
    management_module = "settings"
    queryset = Location.objects.all()
    serializer_class = LocationSerializer


class CategoryViewSet(NoDeleteViewSet):
    management_module = "settings"
    queryset = AssetCategory.objects.all()
    serializer_class = CategorySerializer


class AssetStatusViewSet(NoDeleteViewSet):
    management_module = "settings"
    queryset = AssetStatus.objects.all()
    serializer_class = AssetStatusSerializer


class InventoryItemViewSet(NoDeleteViewSet):
    management_module = "inventory"
    serializer_class = InventoryItemSerializer

    def get_queryset(self):
        queryset = InventoryItem.objects.select_related("location").prefetch_related(
            "transactions__recipient",
            "transactions__actor",
        )
        query = self.request.query_params.get("q", "").strip()
        if query:
            queryset = queryset.filter(
                Q(sku__icontains=query)
                | Q(name__icontains=query)
                | Q(brand__icontains=query)
                | Q(model_name__icontains=query)
            )
        kind = self.request.query_params.get("kind")
        if kind:
            queryset = queryset.filter(kind=kind)
        return queryset

    def destroy(self, request, *args, **kwargs):
        if not is_hidden_superuser(request.user):
            raise MethodNotAllowed(
                "DELETE",
                detail="库存物品不能直接删除，请通过出入库记录处理。",
            )
        item = self.get_object()
        try:
            item.delete()
        except ProtectedError:
            return Response(
                {"message": "该库存物品已被出入库或申请记录引用，无法删除。"},
                status=400,
            )
        return Response(status=204)

    @action(detail=False, methods=["get"])
    def export(self, request):
        queryset = self.get_queryset().order_by("kind", "sku")
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "库存列表"
        headers = [
            "物品编码",
            "物品名称",
            "物品分类",
            "品牌",
            "型号",
            "当前数量",
            "单位",
            "单价",
            "库存金额",
            "采购途径",
            "保障数量",
            "存放地点",
            "状态",
            "备注",
        ]
        sheet.append(headers)
        for item in queryset:
            inventory_value = (
                item.unit_price * item.quantity if item.unit_price is not None else None
            )
            sheet.append(
                [
                    item.sku,
                    item.name,
                    item.get_kind_display(),
                    item.brand,
                    item.model_name,
                    item.quantity,
                    item.unit,
                    item.unit_price,
                    inventory_value,
                    item.get_purchase_channel_display() if item.purchase_channel else "",
                    item.minimum_quantity,
                    item.location.name if item.location else "",
                    "启用" if item.is_active else "停用",
                    item.notes,
                ]
            )
        header = sheet[1]
        for cell in header:
            cell.font = openpyxl.styles.Font(bold=True, color="FFFFFF")
            cell.fill = openpyxl.styles.PatternFill("solid", fgColor="167D91")
            cell.alignment = openpyxl.styles.Alignment(horizontal="center")
        widths = [18, 22, 12, 14, 18, 12, 9, 13, 15, 16, 13, 18, 10, 28]
        for index, width in enumerate(widths, start=1):
            sheet.column_dimensions[openpyxl.utils.get_column_letter(index)].width = width
        for row in range(2, sheet.max_row + 1):
            sheet.cell(row, 6).number_format = "0"
            sheet.cell(row, 8).number_format = "¥#,##0.00"
            sheet.cell(row, 9).number_format = "¥#,##0.00"
            sheet.cell(row, 11).number_format = "0"
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = (
            f'attachment; filename="inventory_export_{timezone.now().strftime("%Y%m%d%H%M%S")}.xlsx"'
        )
        workbook.save(response)
        return response

    @action(detail=False, methods=["get"], url_path="purchase-export")
    def purchase_export(self, request):
        queryset = self.get_queryset().filter(is_active=True)
        include_sufficient = request.query_params.get("include_sufficient") in {
            "true",
            "1",
        }
        if not include_sufficient:
            queryset = queryset.filter(quantity__lt=F("minimum_quantity"))
        queryset = queryset.order_by("kind", "sku")

        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "采购清单"
        headers = [
            "物品编码",
            "物品名称",
            "物品分类",
            "品牌",
            "型号",
            "当前库存",
            "保障数量",
            "建议采购数量",
            "单位",
            "单价",
            "预计金额",
            "采购途径",
            "存放地点",
            "备注",
        ]
        sheet.append(headers)
        for item in queryset:
            suggested_quantity = max(item.minimum_quantity - item.quantity, 0)
            estimated_amount = (
                item.unit_price * suggested_quantity
                if item.unit_price is not None
                else None
            )
            sheet.append(
                [
                    item.sku,
                    item.name,
                    item.get_kind_display(),
                    item.brand,
                    item.model_name,
                    item.quantity,
                    item.minimum_quantity,
                    suggested_quantity,
                    item.unit,
                    item.unit_price,
                    estimated_amount,
                    item.get_purchase_channel_display()
                    if item.purchase_channel
                    else "未设置",
                    item.location.name if item.location else "",
                    item.notes,
                ]
            )
        for cell in sheet[1]:
            cell.font = openpyxl.styles.Font(bold=True, color="FFFFFF")
            cell.fill = openpyxl.styles.PatternFill("solid", fgColor="B76532")
            cell.alignment = openpyxl.styles.Alignment(horizontal="center")
        widths = [18, 22, 12, 14, 18, 12, 13, 16, 9, 13, 15, 16, 18, 28]
        for index, width in enumerate(widths, start=1):
            sheet.column_dimensions[openpyxl.utils.get_column_letter(index)].width = width
        for row in range(2, sheet.max_row + 1):
            for column in (6, 7, 8):
                sheet.cell(row, column).number_format = "0"
            for column in (10, 11):
                sheet.cell(row, column).number_format = "¥#,##0.00"
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        filename = f"采购清单_{timezone.now().strftime('%Y%m%d%H%M%S')}.xlsx"
        response["Content-Disposition"] = (
            f"attachment; filename*=UTF-8''{quote(filename)}"
        )
        workbook.save(response)
        return response

    @action(
        detail=False,
        methods=["post"],
        url_path="import",
        parser_classes=[MultiPartParser],
    )
    def import_excel(self, request):
        upload = request.FILES.get("file")
        if not upload:
            return Response({"message": "请选择 Excel 文件。"}, status=400)
        if not upload.name.lower().endswith(".xlsx"):
            return Response({"message": "只支持 .xlsx 文件。"}, status=400)
        try:
            rows = parse_inventory_workbook(upload)
            preview = summarize_inventory_import(rows)
            if request.data.get("commit") in {"true", "1", True}:
                result = apply_inventory_import(rows, request.user)
                return Response({"message": "库存导入完成。", **result})
            return Response(preview)
        except InventoryImportError as exc:
            return Response({"message": str(exc)}, status=400)

    @action(detail=False, methods=["get"], url_path="import-template")
    def import_template(self, request):
        template = Path(settings.BASE_DIR) / "assets" / "data" / "库存导入模板.xlsx"
        if not template.exists():
            return Response({"message": "导入模板暂不可用。"}, status=503)
        return FileResponse(
            template.open("rb"),
            as_attachment=True,
            filename="库存导入模板.xlsx",
        )

    @action(detail=True, methods=["post"], url_path="transactions")
    @transaction.atomic
    def transact(self, request, pk=None):
        serializer = InventoryActionSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        item = InventoryItem.objects.select_for_update().get(pk=self.get_object().pk)
        action_value = serializer.validated_data["action"]
        quantity = serializer.validated_data["quantity"]
        if action_value in {
            InventoryTransaction.Action.ISSUE,
            InventoryTransaction.Action.WRITEOFF,
        }:
            if quantity > item.quantity:
                return Response({"message": "库存不足，不能完成这次操作。"}, status=400)
            item.quantity -= quantity
        else:
            item.quantity += quantity
        item.save(update_fields=["quantity", "updated_at"])
        inventory_transaction = InventoryTransaction.objects.create(
            item=item,
            action=action_value,
            quantity=quantity,
            balance_after=item.quantity,
            recipient=serializer.validated_data.get("recipient"),
            actor=request.user,
            notes=serializer.validated_data.get("notes", ""),
        )
        notify_inventory_transaction(inventory_transaction)
        return Response(InventoryItemSerializer(item, context={"request": request}).data)


class StocktakeTaskViewSet(NoDeleteViewSet):
    management_module = "stocktake"
    serializer_class = StocktakeTaskSerializer

    def get_queryset(self):
        queryset = StocktakeTask.objects.select_related(
            "scope_location",
            "created_by",
        ).annotate(
            scanned_count=Count(
                "records",
                filter=~Q(records__result=StocktakeRecord.Result.PENDING),
            ),
            missing_count=Count(
                "records",
                filter=Q(records__result=StocktakeRecord.Result.MISSING),
            ),
        )
        if self.action == "retrieve":
            queryset = queryset.prefetch_related(
                "records__asset",
                "records__expected_location",
                "records__expected_user",
            )
        return queryset

    @transaction.atomic
    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        task = serializer.save(created_by=request.user)
        assets = Asset.objects.exclude(status=Asset.Status.DISPOSED).select_related(
            "current_location",
            "assigned_to",
        )
        if task.scope_location_id:
            assets = assets.filter(current_location=task.scope_location)
        records = [
            StocktakeRecord(
                task=task,
                asset=asset,
                expected_location=asset.current_location,
                expected_user=asset.assigned_to,
            )
            for asset in assets
        ]
        StocktakeRecord.objects.bulk_create(records)
        task.snapshot_count = len(records)
        task.save(update_fields=["snapshot_count", "updated_at"])
        task = self.get_queryset().get(pk=task.pk)
        return Response(self.get_serializer(task).data, status=201)

    @action(detail=True, methods=["post"], url_path="scan")
    @transaction.atomic
    def scan(self, request, pk=None):
        task = StocktakeTask.objects.select_for_update().get(pk=self.get_object().pk)
        if task.status != StocktakeTask.Status.IN_PROGRESS:
            return Response({"message": "这项盘点已经完成。"}, status=400)
        asset_tag = str(request.data.get("asset_tag", "")).strip()
        if not asset_tag:
            return Response({"message": "请输入或扫描资产编号。"}, status=400)
        record = StocktakeRecord.objects.select_related("asset").filter(
            task=task,
            asset__asset_tag__iexact=asset_tag,
        ).first()
        if not record:
            return Response({"message": "这件资产不在本次应盘清单中。"}, status=400)
        actual_location_id = request.data.get("actual_location_id")
        record.result = (
            StocktakeRecord.Result.LOCATION_MISMATCH
            if actual_location_id
            and str(actual_location_id) != str(record.expected_location_id or "")
            else StocktakeRecord.Result.MATCHED
        )
        record.scanned_at = timezone.now()
        record.scanned_by = request.user
        record.save(update_fields=["result", "scanned_at", "scanned_by"])
        record.asset.last_audited_at = timezone.now()
        record.asset.save(update_fields=["last_audited_at", "updated_at"])
        task = self.get_queryset().get(pk=task.pk)
        return Response(self.get_serializer(task).data)

    @action(detail=True, methods=["post"], url_path="complete")
    @transaction.atomic
    def complete(self, request, pk=None):
        task = StocktakeTask.objects.select_for_update().get(pk=self.get_object().pk)
        if task.status == StocktakeTask.Status.COMPLETED:
            return Response(self.get_serializer(self.get_queryset().get(pk=task.pk)).data)
        task.records.filter(result=StocktakeRecord.Result.PENDING).update(
            result=StocktakeRecord.Result.MISSING
        )
        task.status = StocktakeTask.Status.COMPLETED
        task.completed_at = timezone.now()
        task.save(update_fields=["status", "completed_at", "updated_at"])
        task = self.get_queryset().get(pk=task.pk)
        return Response(self.get_serializer(task).data)


class ReportsView(APIView):
    permission_classes = [IsModuleManager]
    management_module = "reports"

    def get(self, request):
        assets = Asset.objects.all()
        total_cost = assets.aggregate(total=Sum("purchase_cost", default=0))["total"]
        used_statuses = set(assets.values_list("status", flat=True))
        status_options = AssetStatus.objects.filter(
            Q(is_active=True) | Q(code__in=used_statuses)
        ).distinct()
        return Response(
            {
                "summary": {
                    "assets": assets.count(),
                    "purchase_cost": total_cost,
                    "in_use": assets.filter(
                        status__in=[Asset.Status.ASSIGNED, Asset.Status.LOANED]
                    ).count(),
                    "available": assets.filter(status=Asset.Status.AVAILABLE).count(),
                },
                "by_category": list(
                    assets.values("category__name")
                    .annotate(total=Count("id"))
                    .order_by("-total", "category__name")
                ),
                "by_department": list(
                    assets.values("custodian_department_id", "custodian_department__name")
                    .annotate(total=Count("id"))
                    .order_by("-total", "custodian_department__name")[:15]
                ),
                "by_status": [
                    {
                        "status": value,
                        "label": label,
                        "total": assets.filter(status=value).count(),
                    }
                    for value, label in status_options.values_list("code", "name")
                ],
                "quality": {
                    "missing_category": assets.filter(category__code="UC").count(),
                    "missing_location": assets.filter(current_location__isnull=True).count(),
                    "missing_serial": assets.filter(serial_number="").count(),
                    "import_warnings": assets.filter(
                        custom_data__has_key="import_warnings"
                    ).count(),
                },
                "low_stock": InventoryItemSerializer(
                    InventoryItem.objects.filter(
                        quantity__lt=F("minimum_quantity"),
                        is_active=True,
                    ).select_related("location"),
                    many=True,
                ).data,
            }
        )


class ReportAssetDetailView(APIView):
    permission_classes = [IsModuleManager]
    management_module = "reports"

    DETAIL_TITLES = {
        "import_warnings": "导入待完善",
        "missing_category": "待分类资产",
        "missing_location": "缺少地点",
        "missing_serial": "缺少序列号",
        "department": "部门资产",
    }

    def _queryset(self, kind, department_id=None):
        assets = Asset.objects.select_related(
            "category", "current_location", "assigned_to", "custodian_department"
        ).order_by("asset_tag")
        if kind == "import_warnings":
            return assets.filter(custom_data__has_key="import_warnings")
        if kind == "missing_category":
            return assets.filter(category__code="UC")
        if kind == "missing_location":
            return assets.filter(current_location__isnull=True)
        if kind == "missing_serial":
            return assets.filter(serial_number="")
        if kind == "department":
            return assets.filter(custodian_department_id=department_id) if department_id else assets.filter(custodian_department__isnull=True)
        return None

    def get(self, request):
        kind = request.query_params.get("kind", "").strip()
        department_id = request.query_params.get("department_id")
        queryset = self._queryset(kind, department_id)
        if queryset is None:
            return Response({"message": "未知的报表明细类型。"}, status=400)
        rows = AssetListSerializer(queryset, many=True, context={"request": request}).data
        if kind == "import_warnings":
            warning_map = {
                asset.pk: asset.custom_data.get("import_warnings", [])
                for asset in queryset
            }
            for row in rows:
                row["import_warnings"] = warning_map.get(row["id"], [])
        title = self.DETAIL_TITLES[kind]
        if kind == "department":
            first_asset = queryset.first()
            title = first_asset.custodian_department.name if first_asset and first_asset.custodian_department else "未分配部门"
        return Response({"kind": kind, "title": title, "count": len(rows), "results": rows})

    @transaction.atomic
    def post(self, request):
        if not user_can_manage(request.user, "assets"):
            return Response({"message": "只有资产管理员可以批量修改资产。"}, status=403)
        kind = str(request.data.get("kind") or "").strip()
        if kind not in self.DETAIL_TITLES:
            return Response({"message": "这个明细不支持批量补齐。"}, status=400)
        if kind == "department":
            return Response(
                {"message": "归属部门由责任人自动确定，请到资产详情中选择责任人。"},
                status=400,
            )
        raw_ids = request.data.get("asset_ids") or []
        try:
            asset_ids = list(dict.fromkeys(int(value) for value in raw_ids))
        except (TypeError, ValueError):
            return Response({"message": "资产选择有误。"}, status=400)
        if not asset_ids:
            return Response({"message": "请至少选择一件资产。"}, status=400)
        queryset = self._queryset(kind).select_for_update(of=("self",)).filter(pk__in=asset_ids)
        assets = list(queryset)
        if len(assets) != len(asset_ids):
            return Response({"message": "部分资产已完成补齐，请刷新后重试。"}, status=409)

        category = location = None
        serial_numbers = request.data.get("serial_numbers") or {}
        if kind == "missing_category":
            category = AssetCategory.objects.filter(pk=request.data.get("category_id"), is_active=True).first()
            if not category:
                return Response({"message": "请选择有效的资产类型。"}, status=400)
        elif kind == "missing_location":
            location = Location.objects.filter(pk=request.data.get("location_id"), is_active=True).first()
            if not location:
                return Response({"message": "请选择有效的地点。"}, status=400)
        elif kind == "missing_serial":
            cleaned = {str(key): str(value).strip() for key, value in serial_numbers.items() if str(value).strip()}
            if any(not cleaned.get(str(asset.pk)) for asset in assets):
                return Response({"message": "请填写所有已勾选资产的序列号。"}, status=400)
            values = list(cleaned.values())
            if len({value.casefold() for value in values}) != len(values):
                return Response({"message": "本次填写的序列号不能重复。"}, status=400)
            duplicate_query = Q()
            for value in values:
                duplicate_query |= Q(serial_number__iexact=value)
            duplicates = Asset.objects.exclude(pk__in=asset_ids).filter(duplicate_query)
            if duplicates.exists():
                return Response({"message": "填写的序列号已被其他资产使用。"}, status=400)

        for asset in assets:
            before_location = asset.current_location
            if kind == "missing_category":
                asset.category = category
                asset.save(update_fields=["category", "updated_at"])
                change_note = f"批量补齐资产类型：{category.name}"
            elif kind == "missing_location":
                asset.current_location = location
                asset.save(update_fields=["current_location", "updated_at"])
                change_note = f"批量补齐地点：{location.name}"
            elif kind == "missing_serial":
                asset.serial_number = str(serial_numbers[str(asset.pk)]).strip()
                asset.save(update_fields=["serial_number", "updated_at"])
                change_note = "批量补齐序列号"
            else:
                custom_data = dict(asset.custom_data)
                resolved = custom_data.pop("import_warnings", [])
                custom_data.pop("import_original_assignee", None)
                custom_data.pop("import_original_status", None)
                asset.custom_data = custom_data
                asset.save(update_fields=["custom_data", "updated_at"])
                change_note = "批量确认导入待完善资料已处理"
            AssetEvent.objects.create(
                asset=asset,
                action=AssetEvent.Action.UPDATED,
                from_status=asset.status,
                to_status=asset.status,
                from_user=asset.assigned_to,
                to_user=asset.assigned_to,
                from_location=before_location,
                to_location=asset.current_location,
                actor=request.user,
                notes=change_note,
                metadata={"report_batch_completion": kind},
            )
        return Response({"message": f"已补齐 {len(assets)} 件资产。", "updated": len(assets)})


class ExpenseCategoryViewSet(NoDeleteViewSet):
    management_module = "settings"
    queryset = ExpenseCategory.objects.all()
    serializer_class = ExpenseCategorySerializer

    def get_permissions(self):
        if self.action in {"list", "retrieve"}:
            return [IsAuthenticated()]
        return super().get_permissions()


class ContractTypeViewSet(NoDeleteViewSet):
    management_module = "settings"
    queryset = ContractType.objects.all()
    serializer_class = ContractTypeSerializer

    def get_permissions(self):
        if self.action in {"list", "retrieve"}:
            return [IsAuthenticated()]
        return super().get_permissions()


class SupplierViewSet(NoDeleteViewSet):
    management_module = "suppliers"
    serializer_class = SupplierSerializer

    def get_permissions(self):
        if self.action in {"list", "retrieve"} or (
            self.action in {"files", "file_content"} and self.request.method == "GET"
        ):
            return [IsAuthenticated()]
        return super().get_permissions()

    def get_queryset(self):
        queryset = Supplier.objects.prefetch_related("attachments__uploaded_by")
        if not any(user_can_manage(self.request.user, scope) for scope in ("suppliers", "procurement", "vehicles", "expenses", "contracts", "offices")):
            return queryset.none()
        query = self.request.query_params.get("q", "").strip()
        active = self.request.query_params.get("active", "").strip().lower()
        license_status = self.request.query_params.get("business_license_status", "").strip()
        if query:
            queryset = queryset.filter(
                Q(code__icontains=query)
                | Q(name__icontains=query)
                | Q(category__icontains=query)
                | Q(business_scope__icontains=query)
                | Q(contact_name__icontains=query)
            )
        if active in {"1", "true", "yes"}:
            queryset = queryset.filter(is_active=True)
        elif active in {"0", "false", "no"}:
            queryset = queryset.filter(is_active=False)
        if license_status in dict(Supplier._meta.get_field("business_license_status").choices):
            queryset = queryset.filter(business_license_status=license_status)
        return queryset

    def destroy(self, request, *args, **kwargs):
        if not is_hidden_superuser(request.user):
            raise MethodNotAllowed("DELETE", detail="供应商不能直接删除，请改为停用。")
        supplier = self.get_object()
        try:
            for attachment in supplier.attachments.all():
                nextcloud_storage.delete(attachment.remote_path)
        except NextcloudStorageError as exc:
            return Response({"message": str(exc)}, status=503)
        try:
            supplier.delete()
        except ProtectedError:
            return Response({"message": "该供应商仍被合同或业务记录引用，无法删除，请改为停用。"}, status=400)
        return Response(status=204)

    @action(detail=True, methods=["get", "post"], url_path="files", parser_classes=[MultiPartParser])
    def files(self, request, pk=None):
        supplier = self.get_object()
        if request.method == "GET":
            return Response(
                SupplierAttachmentSerializer(
                    supplier.attachments.select_related("uploaded_by"), many=True
                ).data
            )
        if supplier.attachments.count() >= 20:
            return Response({"message": "每家供应商最多保存 20 个证照文件。"}, status=400)
        upload = request.FILES.get("file")
        error = _validate_upload(
            upload, SUPPLIER_FILE_EXTENSIONS, SUPPLIER_FILE_MAX_BYTES, "供应商证照"
        )
        if error:
            return Response({"message": error}, status=400)
        document_type = request.data.get(
            "document_type", SupplierAttachment.DocumentType.BUSINESS_LICENSE
        )
        if document_type not in dict(SupplierAttachment.DocumentType.choices):
            return Response({"message": "请选择正确的供应商文件类别。"}, status=400)
        remote_path = _remote_file_path(
            "suppliers", timezone.localdate().year, supplier.code, upload
        )
        try:
            nextcloud_storage.upload(upload, remote_path)
            try:
                attachment = SupplierAttachment.objects.create(
                    supplier=supplier,
                    document_type=document_type,
                    remote_path=remote_path,
                    original_name=Path(upload.name).name[:255],
                    content_type=(upload.content_type or "")[:120],
                    size_bytes=upload.size,
                    sha256=_file_sha256(upload),
                    uploaded_by=request.user,
                )
                if document_type == SupplierAttachment.DocumentType.BUSINESS_LICENSE:
                    supplier.business_license_status = "registered"
                    supplier.save(update_fields=["business_license_status", "updated_at"])
            except Exception:
                nextcloud_storage.delete(remote_path)
                raise
        except NextcloudStorageError as exc:
            return Response({"message": str(exc)}, status=503)
        return Response(SupplierAttachmentSerializer(attachment).data, status=201)

    @action(detail=True, methods=["get", "delete"], url_path=r"files/(?P<file_id>\d+)")
    def file_content(self, request, pk=None, file_id=None):
        supplier = self.get_object()
        attachment = supplier.attachments.filter(pk=file_id).first()
        if not attachment:
            return Response({"message": "没有找到这个供应商文件。"}, status=404)
        if request.method == "GET":
            try:
                return _stream_remote_file(attachment, as_attachment=True)
            except NextcloudStorageError as exc:
                return Response({"message": str(exc)}, status=503)
        try:
            nextcloud_storage.delete(attachment.remote_path)
        except NextcloudStorageError as exc:
            return Response({"message": str(exc)}, status=503)
        attachment.delete()
        return Response(status=204)


class OfficeViewSet(NoDeleteViewSet):
    management_module = "offices"
    serializer_class = OfficeSerializer

    def get_permissions(self):
        if self.action in {"list", "retrieve"}:
            return [IsAuthenticated()]
        return super().get_permissions()

    def get_queryset(self):
        queryset = Office.objects.prefetch_related(
            "contracts__owner",
            "contracts__contract_type",
            "resident_users__employee_profile__department",
        )
        if not any(user_can_manage(self.request.user, scope) for scope in ("offices", "contracts")):
            return queryset.none()
        query = self.request.query_params.get("q", "").strip()
        status_value = self.request.query_params.get("status", "").strip()
        city = self.request.query_params.get("city", "").strip()
        if query:
            queryset = queryset.filter(
                Q(code__icontains=query)
                | Q(name__icontains=query)
                | Q(city__icontains=query)
                | Q(address__icontains=query)
                | Q(responsible_name__icontains=query)
            )
        if status_value in dict(Office.Status.choices):
            queryset = queryset.filter(status=status_value)
        if city:
            queryset = queryset.filter(city=city)
        return queryset.distinct()


class ContractViewSet(viewsets.ModelViewSet):
    management_module = "contracts"
    permission_classes = [IsModuleManager]
    serializer_class = ContractSerializer
    pagination_class = None

    def get_permissions(self):
        if self.action in {"list", "retrieve", "history"}:
            return [IsAuthenticated()]
        return super().get_permissions()

    def get_queryset(self):
        queryset = Contract.objects.select_related(
            "contract_type", "supplier", "office", "category", "department", "owner", "previous_contract", "supplement_of"
        ).prefetch_related(
            "attachments__uploaded_by", "attachments__change", "changes__created_by", "renewal_contracts", "supplement_contracts"
        )
        if not any(user_can_manage(self.request.user, scope) for scope in ("contracts", "expenses", "procurement")):
            return queryset.none()
        queryset = contracts_visible_to(self.request.user, queryset)
        if self.action == "list":
            queryset = queryset.filter(renewal_contracts__isnull=True, supplement_of__isnull=True)
        query = self.request.query_params.get("q", "").strip()
        status_value = self.request.query_params.get("status", "").strip()
        contract_type = self.request.query_params.get("contract_type", "").strip()
        office = self.request.query_params.get("office", "").strip()
        due = self.request.query_params.get("due", "").strip().lower()
        if query:
            queryset = queryset.filter(
                Q(contract_no__icontains=query)
                | Q(name__icontains=query)
                | Q(supplier__name__icontains=query)
                | Q(owner__first_name__icontains=query)
                | Q(owner__username__icontains=query)
            )
        if status_value:
            queryset = queryset.filter(status=status_value)
        if contract_type:
            queryset = queryset.filter(contract_type_id=contract_type)
        if office:
            queryset = queryset.filter(office_id=office)
        if due in {"1", "true", "yes"}:
            today = timezone.localdate()
            queryset = queryset.filter(
                status__in=[Contract.Status.ACTIVE, Contract.Status.EXPIRED],
            ).filter(
                Q(end_date__lte=today)
                | Q(end_date__in=expiry_reminder_dates(today))
            )
        return queryset.distinct()

    def perform_create(self, serializer):
        if is_hidden_superuser(self.request.user):
            serializer.save()
        else:
            serializer.save(owner=self.request.user)

    def perform_update(self, serializer):
        if is_hidden_superuser(self.request.user):
            serializer.save()
        else:
            serializer.save(owner=self.request.user)

    @action(detail=True, methods=["get"], url_path="history")
    def history(self, request, pk=None):
        contract = self.get_object()
        chain = []
        visited = set()
        current = contract
        while current and current.pk not in visited:
            if not is_hidden_superuser(request.user) and current.owner_id != request.user.id:
                break
            visited.add(current.pk)
            chain.append(current)
            current = current.previous_contract
        chain.reverse()
        return Response(self.get_serializer(chain, many=True).data)

    @action(detail=True, methods=["post"], url_path="renew")
    def renew(self, request, pk=None):
        previous = self.get_object()
        if previous.renewal_contracts.exists():
            return Response({"message": "这份合同已经生成续签合同，请进入续签合同继续处理。"}, status=400)
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        with transaction.atomic():
            save_kwargs = {"previous_contract": previous}
            if not is_hidden_superuser(request.user):
                save_kwargs["owner"] = request.user
            renewed = serializer.save(**save_kwargs)
            if previous.status not in {Contract.Status.COMPLETED, Contract.Status.TERMINATED}:
                previous.status = Contract.Status.COMPLETED
                previous.save(update_fields=["status", "updated_at"])
        return Response(self.get_serializer(renewed).data, status=201)

    @action(detail=True, methods=["post"], url_path="changes")
    def register_change(self, request, pk=None):
        contract = self.get_object()
        serializer = ContractChangeSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        data = serializer.validated_data
        if data["change_type"] == ContractChange.ChangeType.SUPPLEMENT:
            return self._register_supplement_change(contract, data, request.user)
        effective_start = data.get("new_start_date") or contract.start_date
        effective_end = data.get("new_end_date") or contract.end_date
        if effective_start and effective_end and effective_end < effective_start:
            return Response({"errors": {"new_end_date": ["变更后的结束日期不能早于开始日期。"]}}, status=400)
        with transaction.atomic():
            change = ContractChange.objects.create(
                contract=contract,
                old_start_date=contract.start_date,
                old_end_date=contract.end_date,
                old_amount=contract.amount,
                created_by=request.user,
                **data,
            )
            if data.get("new_start_date"):
                contract.start_date = data["new_start_date"]
            if data.get("new_end_date"):
                contract.end_date = data["new_end_date"]
            if data.get("new_amount") is not None:
                contract.amount = data["new_amount"]
            if data["change_type"] == ContractChange.ChangeType.TERMINATION:
                contract.status = Contract.Status.TERMINATED
            elif contract.status == Contract.Status.EXPIRED and (
                not contract.end_date or contract.end_date >= date.today()
            ):
                contract.status = Contract.Status.ACTIVE
            contract.save()
        return Response(ContractChangeSerializer(change).data, status=201)

    def _register_supplement_change(self, parent, data, user):
        """把补充协议登记为一份附属合同，母合同金额与补充金额在列表界面自动合计。"""
        if parent.supplement_of_id:
            return Response(
                {"errors": {"change_type": ["补充协议请登记在母合同上，不能在补充协议上再次登记补充协议。"]}},
                status=400,
            )
        with transaction.atomic():
            index = parent.supplement_contracts.count() + 1
            contract_no = _next_supplement_contract_no(parent, index)
            supplement = Contract.objects.create(
                contract_no=contract_no,
                name=f"{parent.name}（补充协议 {index}）",
                contract_type=parent.contract_type,
                supplier=parent.supplier,
                office=parent.office,
                category=parent.category,
                department=parent.department,
                owner=parent.owner,
                status=Contract.Status.ACTIVE,
                start_date=data["new_start_date"],
                end_date=data["new_end_date"],
                amount=data["new_amount"],
                renewal_notice_days=parent.renewal_notice_days,
                auto_renew=False,
                supplement_of=parent,
            )
            change = ContractChange.objects.create(
                contract=parent,
                change_type=ContractChange.ChangeType.SUPPLEMENT,
                changed_on=data["changed_on"],
                old_start_date=parent.start_date,
                old_end_date=parent.end_date,
                old_amount=parent.amount,
                new_start_date=supplement.start_date,
                new_end_date=supplement.end_date,
                new_amount=supplement.amount,
                notes=data["notes"],
                created_by=user,
            )
        payload = ContractChangeSerializer(change).data
        payload["supplement"] = {
            "id": supplement.id,
            "contract_no": supplement.contract_no,
            "name": supplement.name,
            "amount": str(supplement.amount),
            "start_date": supplement.start_date,
            "end_date": supplement.end_date,
        }
        return Response(payload, status=201)

    def destroy(self, request, *args, **kwargs):
        contract = self.get_object()
        if not is_hidden_superuser(request.user):
            raise MethodNotAllowed("DELETE", detail="只有超级管理员可以删除合同。")
        targets = [contract, *contract.supplement_contracts.all()]
        try:
            for item in targets:
                for attachment in item.attachments.all():
                    nextcloud_storage.delete(attachment.remote_path)
        except NextcloudStorageError as exc:
            return Response({"message": str(exc)}, status=503)
        contract.delete()
        return Response(status=204)

    @action(
        detail=True,
        methods=["get", "post"],
        url_path="files",
        parser_classes=[MultiPartParser],
    )
    def files(self, request, pk=None):
        contract = self.get_object()
        if request.method == "GET":
            return Response(
                ContractAttachmentSerializer(
                    contract.attachments.select_related("uploaded_by"),
                    many=True,
                ).data
            )

        if contract.attachments.count() >= 30:
            return Response({"message": "每份合同最多保存 30 个文件。"}, status=400)
        upload = request.FILES.get("file")
        error = _validate_upload(
            upload,
            CONTRACT_FILE_EXTENSIONS,
            CONTRACT_FILE_MAX_BYTES,
            "合同文件",
        )
        if error:
            return Response({"message": error}, status=400)
        document_type = request.data.get(
            "document_type",
            ContractAttachment.DocumentType.SIGNED,
        )
        if document_type not in dict(ContractAttachment.DocumentType.choices):
            return Response({"message": "请选择正确的合同文件类别。"}, status=400)
        change = None
        change_id = request.data.get("change_id")
        if change_id:
            change = contract.changes.filter(pk=change_id).first()
            if not change:
                return Response({"message": "没有找到对应的合同变更记录。"}, status=400)

        contract_year = contract.start_date.year if contract.start_date else contract.created_at.year
        remote_path = _remote_file_path(
            "contracts",
            contract_year,
            contract.contract_no,
            upload,
        )
        try:
            nextcloud_storage.upload(upload, remote_path)
            try:
                attachment = ContractAttachment.objects.create(
                    contract=contract,
                    change=change,
                    document_type=document_type,
                    remote_path=remote_path,
                    original_name=Path(upload.name).name[:255],
                    content_type=(upload.content_type or "")[:120],
                    size_bytes=upload.size,
                    sha256=_file_sha256(upload),
                    uploaded_by=request.user,
                )
            except Exception:
                nextcloud_storage.delete(remote_path)
                raise
        except NextcloudStorageError as exc:
            return Response({"message": str(exc)}, status=503)

        return Response(ContractAttachmentSerializer(attachment).data, status=201)

    @action(
        detail=True,
        methods=["get", "delete"],
        url_path=r"files/(?P<file_id>\d+)",
    )
    def file_content(self, request, pk=None, file_id=None):
        contract = self.get_object()
        attachment = contract.attachments.filter(pk=file_id).first()
        if not attachment:
            return Response({"message": "没有找到这个合同文件。"}, status=404)

        if request.method == "GET":
            try:
                return _stream_remote_file(attachment, as_attachment=True)
            except NextcloudStorageError as exc:
                return Response({"message": str(exc)}, status=503)

        try:
            nextcloud_storage.delete(attachment.remote_path)
        except NextcloudStorageError as exc:
            return Response({"message": str(exc)}, status=503)
        attachment.delete()
        return Response(status=204)


class VehicleViewSet(viewsets.ModelViewSet):
    management_module = "vehicles"
    permission_classes = [IsModuleManager]
    serializer_class = VehicleSerializer
    pagination_class = None

    def get_queryset(self):
        queryset = Vehicle.objects.select_related("department", "custodian")
        query = self.request.query_params.get("q", "").strip()
        status_value = self.request.query_params.get("status", "").strip()
        insurance_due = self.request.query_params.get("insurance_due", "").strip().lower()
        if query:
            queryset = queryset.filter(Q(plate_number__icontains=query) | Q(name__icontains=query) | Q(brand__icontains=query) | Q(model_name__icontains=query) | Q(insurer_name__icontains=query) | Q(company__icontains=query))
        if status_value:
            queryset = queryset.filter(status=status_value)
        if insurance_due in {"1", "true", "yes"}:
            today = timezone.localdate()
            reminder_dates = expiry_reminder_dates(today)
            queryset = queryset.filter(
                Q(insurance_expires_at__lte=today)
                | Q(inspection_expires_at__lte=today)
                | Q(insurance_expires_at__in=reminder_dates)
                | Q(inspection_expires_at__in=reminder_dates)
            ).exclude(status=Vehicle.Status.RETIRED)
        return queryset

    def destroy(self, request, *args, **kwargs):
        if not is_hidden_superuser(request.user):
            raise MethodNotAllowed("DELETE", detail="只有超级管理员可以删除车辆。")
        vehicle = self.get_object()
        try:
            vehicle.delete()
        except ProtectedError:
            return Response(
                {"message": "该车辆已有派车或费用记录，无法删除，请先处置为“已处置”。"},
                status=400,
            )
        return Response(status=204)


class VehicleDispatchViewSet(viewsets.ModelViewSet):
    permission_classes = [IsAuthenticated]
    serializer_class = VehicleDispatchSerializer
    pagination_class = None

    def perform_create(self, serializer):
        instance = serializer.save()
        notify_vehicle_dispatch_submitted(instance)

    def get_queryset(self):
        queryset = VehicleDispatch.objects.select_related("requester", "department", "vehicle", "driver", "handled_by")
        if not user_can_manage(self.request.user, "vehicles"):
            queryset = queryset.filter(requester=self.request.user)
        status_value = self.request.query_params.get("status", "").strip()
        if status_value:
            queryset = queryset.filter(status=status_value)
        return queryset

    def update(self, request, *args, **kwargs):
        dispatch = self.get_object()
        if dispatch.requester_id != request.user.id or dispatch.status != VehicleDispatch.Status.PENDING:
            return Response({"message": "只能修改自己尚未处理的派车申请。"}, status=403)
        return super().update(request, *args, **kwargs)

    def destroy(self, request, *args, **kwargs):
        return Response({"message": "派车申请不能删除，可取消保留记录。"}, status=405)

    def _require_manager(self, request):
        if not user_can_manage(request.user, "vehicles"):
            return Response({"message": "只有车辆管理员可以处理派车申请。"}, status=403)
        return None

    @action(detail=True, methods=["post"])
    @transaction.atomic
    def approve(self, request, pk=None):
        denied = self._require_manager(request)
        if denied:
            return denied
        dispatch = VehicleDispatch.objects.select_for_update().get(pk=self.get_object().pk)
        if dispatch.status != VehicleDispatch.Status.PENDING:
            return Response({"message": "这条申请已经处理。"}, status=400)
        dispatch.status = VehicleDispatch.Status.APPROVED
        dispatch.handled_by = request.user
        dispatch.notes = str(request.data.get("notes") or "").strip()
        dispatch.save()
        return Response(self.get_serializer(dispatch).data)

    @action(detail=True, methods=["post"], url_path="dispatch")
    @transaction.atomic
    def dispatch_vehicle(self, request, pk=None):
        denied = self._require_manager(request)
        if denied:
            return denied
        dispatch = VehicleDispatch.objects.select_for_update().get(pk=self.get_object().pk)
        if dispatch.status not in {VehicleDispatch.Status.PENDING, VehicleDispatch.Status.APPROVED}:
            return Response({"message": "当前状态不能派车。"}, status=400)
        vehicle = Vehicle.objects.select_for_update().filter(pk=request.data.get("vehicle_id")).first()
        if not vehicle or vehicle.status != Vehicle.Status.AVAILABLE:
            return Response({"message": "请选择一辆当前可用的车辆。"}, status=400)
        if dispatch.passenger_count > vehicle.seats:
            return Response({"message": "乘车人数超过车辆座位数。"}, status=400)
        conflict = VehicleDispatch.objects.exclude(pk=dispatch.pk).filter(
            vehicle=vehicle,
            status__in=[VehicleDispatch.Status.DISPATCHED, VehicleDispatch.Status.IN_PROGRESS],
            planned_departure_at__lt=dispatch.planned_return_at,
            planned_return_at__gt=dispatch.planned_departure_at,
        ).exists()
        if conflict:
            return Response({"message": "该车辆在计划时段已有任务。"}, status=400)
        driver = User.objects.filter(pk=request.data.get("driver_id"), is_active=True).first()
        driver_name = str(request.data.get("driver_name") or "").strip()
        if not driver and not driver_name:
            return Response({"message": "请选择驾驶员或填写外部驾驶员。"}, status=400)
        dispatch.vehicle = vehicle
        dispatch.driver = driver
        dispatch.driver_name = "" if driver else driver_name
        dispatch.status = VehicleDispatch.Status.DISPATCHED
        dispatch.handled_by = request.user
        dispatch.notes = str(request.data.get("notes") or "").strip()
        dispatch.save()
        return Response(self.get_serializer(dispatch).data)

    @action(detail=True, methods=["post"])
    @transaction.atomic
    def depart(self, request, pk=None):
        denied = self._require_manager(request)
        if denied:
            return denied
        dispatch = VehicleDispatch.objects.select_for_update().get(pk=self.get_object().pk)
        if dispatch.status != VehicleDispatch.Status.DISPATCHED or not dispatch.vehicle:
            return Response({"message": "只有已派车的任务可以出车。"}, status=400)
        vehicle = Vehicle.objects.select_for_update().get(pk=dispatch.vehicle_id)
        mileage = int(request.data.get("mileage") or vehicle.current_mileage)
        if mileage < vehicle.current_mileage:
            return Response({"message": "出车里程不能小于车辆当前里程。"}, status=400)
        dispatch.start_mileage = mileage
        dispatch.actual_departure_at = timezone.now()
        dispatch.status = VehicleDispatch.Status.IN_PROGRESS
        dispatch.save()
        vehicle.status = Vehicle.Status.IN_USE
        vehicle.current_mileage = mileage
        vehicle.save(update_fields=["status", "current_mileage", "updated_at"])
        return Response(self.get_serializer(dispatch).data)

    @action(detail=True, methods=["post"])
    @transaction.atomic
    def complete(self, request, pk=None):
        denied = self._require_manager(request)
        if denied:
            return denied
        dispatch = VehicleDispatch.objects.select_for_update().get(pk=self.get_object().pk)
        if dispatch.status != VehicleDispatch.Status.IN_PROGRESS or not dispatch.vehicle:
            return Response({"message": "只有出车中的任务可以办理返回。"}, status=400)
        vehicle = Vehicle.objects.select_for_update().get(pk=dispatch.vehicle_id)
        mileage = int(request.data.get("mileage") or 0)
        if mileage < (dispatch.start_mileage or 0):
            return Response({"message": "返回里程不能小于出车里程。"}, status=400)
        dispatch.end_mileage = mileage
        dispatch.actual_return_at = timezone.now()
        dispatch.status = VehicleDispatch.Status.COMPLETED
        dispatch.notes = str(request.data.get("notes") or dispatch.notes).strip()
        dispatch.save()
        vehicle.status = Vehicle.Status.AVAILABLE
        vehicle.current_mileage = mileage
        vehicle.save(update_fields=["status", "current_mileage", "updated_at"])
        return Response(self.get_serializer(dispatch).data)

    @action(detail=True, methods=["post"])
    def reject(self, request, pk=None):
        denied = self._require_manager(request)
        if denied:
            return denied
        dispatch = self.get_object()
        if dispatch.status != VehicleDispatch.Status.PENDING:
            return Response({"message": "这条申请已经处理。"}, status=400)
        dispatch.status = VehicleDispatch.Status.REJECTED
        dispatch.handled_by = request.user
        dispatch.notes = str(request.data.get("notes") or "").strip()
        dispatch.save()
        return Response(self.get_serializer(dispatch).data)

    @action(detail=True, methods=["post"])
    def cancel(self, request, pk=None):
        dispatch = self.get_object()
        if dispatch.requester_id != request.user.id or dispatch.status not in {VehicleDispatch.Status.PENDING, VehicleDispatch.Status.APPROVED}:
            return Response({"message": "当前申请不能取消。"}, status=403)
        dispatch.status = VehicleDispatch.Status.CANCELLED
        dispatch.save(update_fields=["status", "updated_at"])
        return Response(self.get_serializer(dispatch).data)


class VehicleExpenseViewSet(viewsets.ModelViewSet):
    management_module = "vehicles"
    permission_classes = [IsModuleManager]
    serializer_class = VehicleExpenseSerializer
    pagination_class = None

    def get_queryset(self):
        queryset = VehicleExpense.objects.select_related("vehicle", "supplier", "expense", "created_by")
        vehicle_id = self.request.query_params.get("vehicle")
        year = self.request.query_params.get("year")
        if vehicle_id:
            queryset = queryset.filter(vehicle_id=vehicle_id)
        if year:
            queryset = queryset.filter(occurred_on__year=year)
        return queryset

    def destroy(self, request, *args, **kwargs):
        if not is_hidden_superuser(request.user):
            raise MethodNotAllowed(
                "DELETE",
                detail="只有超级管理员可以删除车辆事项与费用。",
            )
        record = self.get_object()
        if record.expense:
            record.expense.delete()
        record.delete()
        return Response(status=204)


class AdministrativeExpenseViewSet(viewsets.ModelViewSet):
    management_module = "expenses"
    permission_classes = [IsModuleManager]
    serializer_class = AdministrativeExpenseSerializer

    def get_queryset(self):
        queryset = AdministrativeExpense.objects.select_related("category", "department", "supplier", "contract", "created_by")
        year = self.request.query_params.get("year")
        category = self.request.query_params.get("category")
        query = self.request.query_params.get("q", "").strip()
        if year:
            queryset = queryset.filter(fiscal_year=year)
        if category:
            queryset = queryset.filter(category_id=category)
        if query:
            queryset = queryset.filter(Q(title__icontains=query) | Q(source_no__icontains=query) | Q(object_label__icontains=query) | Q(supplier__name__icontains=query))
        return queryset

    def destroy(self, request, *args, **kwargs):
        return Response({"message": "费用记录不能删除，请新增冲销记录。"}, status=405)

    @action(detail=False, methods=["get"])
    def summary(self, request):
        year = int(request.query_params.get("year") or date.today().year)
        rows = AdministrativeExpense.objects.filter(fiscal_year=year)
        totals = {row["amount_type"]: row["total"] or 0 for row in rows.values("amount_type").annotate(total=Sum("amount"))}
        actual = totals.get(AdministrativeExpense.AmountType.ACTUAL, 0)
        reversal = totals.get(AdministrativeExpense.AmountType.REVERSAL, 0)
        by_category = list(rows.values("category__name").annotate(total=Sum("amount")).order_by("-total"))
        by_month = list(rows.filter(amount_type=AdministrativeExpense.AmountType.ACTUAL).annotate(month=ExtractMonth("occurred_on")).values("month").annotate(total=Sum("amount")).order_by("month"))
        return Response({
            "year": year,
            "totals": {"estimated": totals.get("estimated", 0), "approved": totals.get("approved", 0), "committed": totals.get("committed", 0), "actual": actual, "reversal": reversal, "net_actual": actual - reversal},
            "by_category": by_category,
            "by_month": by_month,
        })

    @action(detail=False, methods=["get"])
    def export(self, request):
        rows = self.get_queryset().order_by("occurred_on", "id")
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "行政费用台账"
        sheet.append(["发生日期", "年度", "费用类别", "归属部门", "费用事项", "金额类型", "金额", "供应商", "合同", "来源", "来源单号", "费用对象", "发票状态", "发票号码", "金蝶编码", "预算系统标识", "同步状态", "备注"])
        for item in rows:
            sheet.append([item.occurred_on, item.fiscal_year, item.category.name, item.department.name if item.department else "", item.title, item.get_amount_type_display(), item.amount, item.supplier.name if item.supplier else "", item.contract.name if item.contract else "", item.source_type, item.source_no, item.object_label, item.get_invoice_status_display(), item.invoice_number, item.kingdee_code, item.external_id, item.sync_status, item.notes])
        for cell in sheet[1]:
            cell.font = openpyxl.styles.Font(bold=True, color="FFFFFF")
            cell.fill = openpyxl.styles.PatternFill("solid", fgColor="146B80")
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for column in range(1, sheet.max_column + 1):
            sheet.column_dimensions[openpyxl.utils.get_column_letter(column)].width = 16 if column != 5 else 28
        for row in range(2, sheet.max_row + 1):
            sheet.cell(row, 7).number_format = "¥#,##0.00"
        response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        filename = f"行政费用台账_{request.query_params.get('year') or '全部'}_{timezone.now():%Y%m%d%H%M%S}.xlsx"
        response["Content-Disposition"] = f"attachment; filename*=UTF-8''{quote(filename)}"
        workbook.save(response)
        return response


class PurchaseRequestViewSet(viewsets.ModelViewSet):
    permission_classes = [IsAuthenticated]
    serializer_class = PurchaseRequestSerializer
    pagination_class = None

    def perform_create(self, serializer):
        instance = serializer.save()
        notify_purchase_request_submitted(instance)

    def get_queryset(self):
        queryset = PurchaseRequest.objects.select_related("requester", "department", "category", "handled_by").prefetch_related("items")
        if not user_can_manage(self.request.user, "procurement"):
            queryset = queryset.filter(requester=self.request.user)
        return queryset

    def update(self, request, *args, **kwargs):
        instance = self.get_object()
        if instance.requester_id != request.user.id or instance.status not in {PurchaseRequest.Status.DRAFT, PurchaseRequest.Status.PENDING}:
            return Response({"message": "只能修改自己尚未审批的申请。"}, status=403)
        return super().update(request, *args, **kwargs)

    def destroy(self, request, *args, **kwargs):
        return Response({"message": "采购申请不能删除，可取消保留记录。"}, status=405)

    @action(detail=True, methods=["post"])
    def approve(self, request, pk=None):
        if not user_can_manage(request.user, "procurement"):
            return Response({"message": "只有采购管理员可以审批。"}, status=403)
        instance = self.get_object()
        if instance.status != PurchaseRequest.Status.PENDING:
            return Response({"message": "这条申请已经处理。"}, status=400)
        instance.status = PurchaseRequest.Status.APPROVED
        instance.handled_by = request.user
        instance.handled_at = timezone.now()
        instance.manager_notes = str(request.data.get("manager_notes") or "").strip()
        instance.save()
        category = instance.category or ExpenseCategory.objects.get_or_create(code="PURCHASE", defaults={"name": "行政采购"})[0]
        AdministrativeExpense.objects.update_or_create(
            source_type="purchase_request", source_id=instance.pk,
            defaults={
                "occurred_on": instance.needed_on or date.today(), "fiscal_year": (instance.needed_on or date.today()).year,
                "category": category, "department": instance.department,
                "amount_type": AdministrativeExpense.AmountType.APPROVED, "amount": instance.estimated_amount,
                "title": f"采购申请 · {instance.request_no}", "source_no": instance.request_no,
                "object_label": "；".join(item.name for item in instance.items.all())[:160],
                "created_by": request.user, "notes": instance.manager_notes,
            },
        )
        return Response(self.get_serializer(instance).data)

    @action(detail=True, methods=["post"])
    def reject(self, request, pk=None):
        if not user_can_manage(request.user, "procurement"):
            return Response({"message": "只有采购管理员可以审批。"}, status=403)
        instance = self.get_object()
        if instance.status != PurchaseRequest.Status.PENDING:
            return Response({"message": "这条申请已经处理。"}, status=400)
        instance.status = PurchaseRequest.Status.REJECTED
        instance.handled_by = request.user
        instance.handled_at = timezone.now()
        instance.manager_notes = str(request.data.get("manager_notes") or "").strip()
        instance.save()
        return Response(self.get_serializer(instance).data)

    @action(detail=True, methods=["post"])
    def cancel(self, request, pk=None):
        instance = self.get_object()
        if instance.requester_id != request.user.id or instance.status not in {PurchaseRequest.Status.DRAFT, PurchaseRequest.Status.PENDING}:
            return Response({"message": "当前申请不能取消。"}, status=403)
        instance.status = PurchaseRequest.Status.CANCELLED
        instance.save(update_fields=["status", "updated_at"])
        return Response(self.get_serializer(instance).data)


class PurchaseOrderViewSet(viewsets.ModelViewSet):
    management_module = "procurement"
    permission_classes = [IsModuleManager]
    serializer_class = PurchaseOrderSerializer
    pagination_class = None

    def get_queryset(self):
        return PurchaseOrder.objects.select_related("request", "supplier", "contract", "created_by").prefetch_related("items")

    def destroy(self, request, *args, **kwargs):
        return Response({"message": "采购订单不能删除，请改为已取消。"}, status=405)
