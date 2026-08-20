import io
import json
from datetime import date, timedelta
from decimal import Decimal
from tempfile import TemporaryDirectory
from unittest.mock import patch

from django.contrib.auth import get_user_model
from django.core import mail
from django.core.management import call_command
from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import TestCase, override_settings
from django.utils import timezone
from openpyxl import Workbook, load_workbook
from rest_framework.authtoken.models import Token
from rest_framework.test import APIClient
from rest_framework.exceptions import ValidationError

from .models import (
    Asset,
    AssetCategory,
    AssetEvent,
    AssetImage,
    AssetManagerRole,
    AssetNumberSequence,
    AssetRequest,
    AssetStatus,
    AdministrativeExpense,
    Contract,
    ContractAttachment,
    ContractChange,
    ContractType,
    Department,
    EmployeeProfile,
    EmailNotification,
    ExpenseCategory,
    InventoryItem,
    InventoryTransaction,
    Location,
    Office,
    OperationLog,
    PurchaseOrder,
    PurchaseRequest,
    Supplier,
    SupplierAttachment,
    StocktakeRecord,
    StocktakeTask,
    Vehicle,
    VehicleDispatch,
    VehicleExpense,
)
from .services import perform_asset_action
from .oidc import sync_oidc_user
from .tasks import send_daily_operational_notifications, send_email_notification

User = get_user_model()


class OperationLogTests(TestCase):
    def setUp(self):
        self.superuser = User.objects.create_user(
            "operation-super",
            password="pass",
            first_name="超级管理员",
            is_superuser=True,
            is_staff=True,
        )
        self.manager = User.objects.create_user(
            "operation-manager",
            password="pass",
            first_name="具体使用人",
        )
        AssetManagerRole.objects.create(user=self.manager, scopes=["settings", "inventory"])
        self.manager_token = Token.objects.create(user=self.manager)
        self.super_token = Token.objects.create(user=self.superuser)
        self.manager_client = APIClient()
        self.manager_client.credentials(
            HTTP_AUTHORIZATION=f"Token {self.manager_token.key}",
            HTTP_X_FORWARDED_FOR="10.1.22.88, 10.1.6.15",
            HTTP_USER_AGENT="AffairsOS test client",
        )
        self.super_client = APIClient()
        self.super_client.credentials(HTTP_AUTHORIZATION=f"Token {self.super_token.key}")

    def test_mutation_is_logged_with_user_action_target_and_ip(self):
        created = self.manager_client.post(
            "/api/v1/categories/",
            {
                "name": "日志测试资产类型",
                "code": "AUDIT-LOG",
                "class_type": "IT",
                "is_active": True,
            },
            format="json",
        )

        self.assertEqual(created.status_code, 201)
        log = OperationLog.objects.get(username=self.manager.username)
        self.assertEqual(log.display_name, "具体使用人")
        self.assertEqual(log.module, "settings")
        self.assertEqual(log.action, "create")
        self.assertEqual(log.action_label, "新增")
        self.assertEqual(log.target_id, str(created.data["id"]))
        self.assertIn("日志测试资产类型", log.target_label)
        self.assertEqual(log.status_code, 201)
        self.assertTrue(log.succeeded)
        self.assertEqual(log.ip_address, "10.1.22.88")
        self.assertNotIn("password", str(log.details).lower())

    def test_failed_mutation_is_logged_and_get_is_not_logged(self):
        AssetCategory.objects.create(name="已有类型", code="AUDIT-DUP")
        listed = self.manager_client.get("/api/v1/categories/")
        failed = self.manager_client.post(
            "/api/v1/categories/",
            {"name": "重复类型", "code": "AUDIT-DUP", "class_type": "IT"},
            format="json",
        )

        self.assertEqual(listed.status_code, 200)
        self.assertEqual(failed.status_code, 400)
        self.assertEqual(OperationLog.objects.count(), 1)
        log = OperationLog.objects.get()
        self.assertFalse(log.succeeded)
        self.assertEqual(log.status_code, 400)

    def test_inventory_action_uses_specific_business_action_label(self):
        item = InventoryItem.objects.create(
            sku="AUDIT-STOCK-001",
            name="日志测试库存",
            kind=InventoryItem.Kind.CONSUMABLE,
            quantity=1,
        )

        response = self.manager_client.post(
            f"/api/v1/inventory/{item.pk}/transactions/",
            {"action": "inbound", "quantity": 2},
            format="json",
        )

        self.assertEqual(response.status_code, 200)
        log = OperationLog.objects.get()
        self.assertEqual(log.action, "inventory_inbound")
        self.assertEqual(log.action_label, "库存入库")
        self.assertEqual(log.target_id, str(item.pk))

    def test_only_superuser_can_view_and_filter_operation_logs(self):
        OperationLog.objects.create(
            user=self.manager,
            username=self.manager.username,
            display_name="具体使用人",
            module="assets",
            module_label="资产管理",
            action="partial_update",
            action_label="编辑",
            target_type="asset",
            target_id="9",
            target_label="IT-NB-009 · 测试电脑",
            method="PATCH",
            path="/api/v1/assets/9/",
            status_code=200,
            succeeded=True,
            ip_address="10.1.22.88",
        )

        denied = self.manager_client.get("/api/v1/settings/operation-logs/")
        visible = self.super_client.get(
            "/api/v1/settings/operation-logs/",
            {"username": self.manager.username, "module": "assets", "result": "success", "q": "测试电脑"},
        )

        self.assertEqual(denied.status_code, 403)
        self.assertEqual(visible.status_code, 200)
        self.assertEqual(visible.data["count"], 1)
        self.assertEqual(visible.data["results"][0]["target_label"], "IT-NB-009 · 测试电脑")
        self.assertEqual(visible.data["filters"]["users"][0]["username"], self.manager.username)

    @override_settings(LOCAL_LOGIN_USERNAME="operation-super")
    def test_successful_login_and_logout_are_logged_without_credentials(self):
        self.super_token.delete()
        login_client = APIClient()
        logged_in = login_client.post(
            "/api/v1/auth/local/login/",
            {"username": "operation-super", "password": "pass"},
            format="json",
        )
        login_client.credentials(HTTP_AUTHORIZATION=f"Token {logged_in.data['token']}")
        logged_out = login_client.post("/api/v1/auth/logout/", {}, format="json")

        self.assertEqual(logged_in.status_code, 200)
        self.assertEqual(logged_out.status_code, 204)
        logs = OperationLog.objects.filter(username="operation-super").order_by("occurred_at")
        self.assertEqual(list(logs.values_list("action", flat=True)), ["login", "logout"])
        self.assertTrue(all("pass" not in str(log.details).lower() for log in logs))


class AssetActionServiceTests(TestCase):
    def setUp(self):
        self.admin = User.objects.create_user("admin", password="pass", is_staff=True, is_superuser=True)
        self.employee = User.objects.create_user("employee", password="pass")
        self.department = Department.objects.create(name="研发部", code="RND")
        EmployeeProfile.objects.create(user=self.employee, employee_no="E001", department=self.department)
        self.category = AssetCategory.objects.create(name="笔记本电脑", code="LT")
        self.location = Location.objects.create(name="IT 库房", code="WH", kind="warehouse")
        self.asset = Asset.objects.create(
            asset_tag="IT-LT-TEST-001",
            name="测试电脑",
            category=self.category,
            status=Asset.Status.AVAILABLE,
            current_location=self.location,
        )

    def test_assign_and_return_create_history(self):
        assigned = perform_asset_action(
            asset=self.asset,
            action="assign",
            actor=self.admin,
            target_user=self.employee,
        )
        self.assertEqual(assigned.asset.status, Asset.Status.ASSIGNED)
        self.assertEqual(assigned.asset.assigned_to, self.employee)
        self.assertEqual(assigned.asset.custodian_department, self.department)

        returned = perform_asset_action(
            asset=assigned.asset,
            action="return",
            actor=self.admin,
            requires_inspection=True,
        )
        self.assertEqual(returned.asset.status, Asset.Status.INSPECTION)
        self.assertIsNone(returned.asset.assigned_to)
        self.assertIsNone(returned.asset.custodian_department)
        self.assertEqual(self.asset.events.count(), 2)

    def test_loan_requires_expected_return_date(self):
        with self.assertRaisesMessage(ValidationError, "预计归还日期"):
            perform_asset_action(
                asset=self.asset,
                action="loan",
                actor=self.admin,
                target_user=self.employee,
            )

    def test_non_requestable_asset_can_still_be_managed_directly(self):
        self.asset.is_requestable = False
        self.asset.save(update_fields=["is_requestable", "updated_at"])
        result = perform_asset_action(
            asset=self.asset,
            action="assign",
            actor=self.admin,
            target_user=self.employee,
        )
        self.assertEqual(result.asset.status, Asset.Status.ASSIGNED)

    def test_cannot_assign_disposed_asset(self):
        self.asset.status = Asset.Status.DISPOSED
        self.asset.save()
        with self.assertRaisesMessage(ValidationError, "不能执行"):
            perform_asset_action(
                asset=self.asset,
                action="assign",
                actor=self.admin,
                target_user=self.employee,
            )


class AssetApiTests(TestCase):
    def setUp(self):
        self.admin = User.objects.create_user("admin", password="pass", is_staff=True, is_superuser=True)
        self.employee = User.objects.create_user("employee", password="pass")
        self.category = AssetCategory.objects.create(name="显示器", code="MN")
        self.location = Location.objects.create(name="库房", code="WH", kind="warehouse")
        self.asset = Asset.objects.create(
            asset_tag="IT-MN-TEST-001",
            name="测试显示器",
            category=self.category,
            status=Asset.Status.AVAILABLE,
            current_location=self.location,
            warranty_expires_at=date.today() + timedelta(days=30),
        )
        self.client = APIClient()
        self.client.force_authenticate(self.admin)

    def test_dashboard_returns_actionable_counts(self):
        response = self.client.get("/api/v1/dashboard/")
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data["summary"]["total"], 1)
        self.assertNotIn("warranty_due", response.data["tasks"])
        self.assertNotIn("attention", response.data["tasks"])
        self.assertNotIn("attention", response.data["summary"])

    def test_overdue_dashboard_count_matches_overdue_asset_list(self):
        overdue = Asset.objects.create(
            asset_tag="IT-MN-OVERDUE-001",
            category=self.category,
            status=Asset.Status.LOANED,
            assigned_to=self.employee,
            expected_return_at=timezone.localdate() - timedelta(days=1),
        )
        Asset.objects.create(
            asset_tag="IT-MN-FUTURE-001",
            category=self.category,
            status=Asset.Status.LOANED,
            assigned_to=self.employee,
            expected_return_at=timezone.localdate() + timedelta(days=1),
        )

        dashboard = self.client.get("/api/v1/dashboard/")
        asset_list = self.client.get("/api/v1/assets/?status=loaned&overdue=1")

        self.assertEqual(dashboard.status_code, 200)
        self.assertEqual(dashboard.data["tasks"]["overdue_loans"], 1)
        self.assertEqual(asset_list.status_code, 200)
        self.assertEqual(asset_list.data["count"], 1)
        self.assertEqual(asset_list.data["results"][0]["id"], overdue.id)

    def test_action_endpoint_assigns_asset(self):
        department = Department.objects.create(name="研发部", code="RND")
        EmployeeProfile.objects.create(user=self.employee, employee_no="E001", department=department)
        response = self.client.post(
            f"/api/v1/assets/{self.asset.pk}/actions/",
            {"action": "assign", "target_user_id": self.employee.pk},
            format="json",
        )
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data["asset"]["status"], Asset.Status.ASSIGNED)
        self.assertEqual(response.data["asset"]["custodian_department"], department.id)
        self.assertEqual(AssetEvent.objects.filter(asset=self.asset).count(), 1)

    def test_direct_assignee_change_automatically_uses_employee_department(self):
        department = Department.objects.create(name="财务部", code="FIN")
        EmployeeProfile.objects.create(user=self.employee, employee_no="E002", department=department)
        response = self.client.patch(
            f"/api/v1/assets/{self.asset.pk}/",
            {"status": Asset.Status.ASSIGNED, "assigned_to": self.employee.pk},
            format="json",
        )
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data["custodian_department"], department.id)
        self.assertEqual(response.data["department_name"], "财务部")
        lookups = self.client.get("/api/v1/lookups/")
        employee = next(item for item in lookups.data["users"] if item["id"] == self.employee.id)
        self.assertEqual(employee["department"], department.id)

    def test_clearing_assignee_also_clears_department_and_returns_asset_to_stock(self):
        department = Department.objects.create(name="行政部", code="ADM")
        EmployeeProfile.objects.create(user=self.employee, employee_no="E003", department=department)
        self.asset.status = Asset.Status.ASSIGNED
        self.asset.assigned_to = self.employee
        self.asset.save()

        response = self.client.patch(
            f"/api/v1/assets/{self.asset.pk}/",
            {"assigned_to": None},
            format="json",
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data["status"], Asset.Status.AVAILABLE)
        self.assertIsNone(response.data["assigned_to"])
        self.assertIsNone(response.data["custodian_department"])

    def test_department_cannot_be_set_independently(self):
        department = Department.objects.create(name="法务部", code="LEGAL")
        response = self.client.patch(
            f"/api/v1/assets/{self.asset.pk}/",
            {"custodian_department": department.pk},
            format="json",
        )
        self.assertEqual(response.status_code, 400)
        self.assertIn("不允许单独设置", str(response.data))

    def test_employee_department_change_updates_assigned_assets(self):
        old_department = Department.objects.create(name="旧部门", code="OLD-API")
        new_department = Department.objects.create(name="新部门", code="NEW-API")
        profile = EmployeeProfile.objects.create(
            user=self.employee,
            employee_no="E004",
            department=old_department,
        )
        self.asset.status = Asset.Status.ASSIGNED
        self.asset.assigned_to = self.employee
        self.asset.save()

        profile.department = new_department
        profile.save()

        self.asset.refresh_from_db()
        self.assertEqual(self.asset.custodian_department, new_department)

    def test_employee_cannot_open_asset_register(self):
        self.asset.assigned_to = self.employee
        self.asset.status = Asset.Status.ASSIGNED
        self.asset.save()
        self.client.force_authenticate(self.employee)
        response = self.client.get("/api/v1/assets/")
        self.assertEqual(response.status_code, 403)

    def test_asset_can_be_deleted_by_superuser(self):
        response = self.client.delete(f"/api/v1/assets/{self.asset.pk}/")
        self.assertEqual(response.status_code, 204)
        self.assertFalse(Asset.objects.filter(pk=self.asset.pk).exists())

    def test_loan_can_be_extended(self):
        self.client.force_authenticate(self.admin)
        loan_due = date.today() + timedelta(days=5)
        extended_due = date.today() + timedelta(days=30)
        loaned = self.client.post(
            f"/api/v1/assets/{self.asset.pk}/actions/",
            {
                "action": "loan",
                "target_user_id": self.employee.pk,
                "expected_return_at": loan_due.isoformat(),
            },
            format="json",
        )
        self.assertEqual(loaned.status_code, 200)
        extended = self.client.post(
            f"/api/v1/assets/{self.asset.pk}/actions/",
            {
                "action": "extend",
                "expected_return_at": extended_due.isoformat(),
                "notes": "项目延期",
            },
            format="json",
        )
        self.assertEqual(extended.status_code, 200)
        self.asset.refresh_from_db()
        self.assertEqual(self.asset.status, Asset.Status.LOANED)
        self.assertEqual(self.asset.expected_return_at, extended_due)
        event = AssetEvent.objects.filter(
            asset=self.asset,
            action=AssetEvent.Action.EXTENDED,
        ).latest("id")
        self.assertEqual(event.to_status, Asset.Status.LOANED)

    def test_extend_requires_loaned_asset(self):
        self.client.force_authenticate(self.admin)
        blocked = self.client.post(
            f"/api/v1/assets/{self.asset.pk}/actions/",
            {"action": "extend", "expected_return_at": "2026-09-10"},
            format="json",
        )
        self.assertEqual(blocked.status_code, 400)
        self.assertIn("只有借用中的资产", json.dumps(blocked.data, ensure_ascii=False))

    def test_extend_requires_return_date(self):
        self.client.force_authenticate(self.admin)
        self.asset.status = Asset.Status.LOANED
        self.asset.assigned_to = self.employee
        self.asset.save()
        blocked = self.client.post(
            f"/api/v1/assets/{self.asset.pk}/actions/",
            {"action": "extend"},
            format="json",
        )
        self.assertEqual(blocked.status_code, 400)
        self.assertIn("延期必须填写新的预计归还日期", json.dumps(blocked.data, ensure_ascii=False))

    def test_asset_configuration_can_be_saved(self):
        response = self.client.patch(
            f"/api/v1/assets/{self.asset.pk}/",
            {
                "specification": "27 英寸 4K",
                "cpu": "i5-13500H",
                "memory": "16G",
                "storage": "1T SSD",
                "wired_mac": "F4A8-0DE1-67F1",
                "wireless_mac": "C815-4ED5-2235",
            },
            format="json",
        )
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data["memory"], "16G")
        self.asset.refresh_from_db()
        self.assertEqual(self.asset.storage, "1T SSD")

    def test_asset_tag_is_generated_and_kingdee_code_is_saved(self):
        response = self.client.post(
            "/api/v1/assets/",
            {
                "category": self.category.pk,
                "asset_tag": "MANUAL-CODE-001",
                "kingdee_code": "KD-000123",
                "purchase_date": "2021-06-18",
            },
            format="json",
        )
        self.assertEqual(response.status_code, 201)
        self.assertEqual(response.data["asset_tag"], "IT-MN-2021-001")
        self.assertNotEqual(response.data["asset_tag"], "MANUAL-CODE-001")
        self.assertEqual(response.data["kingdee_code"], "KD-000123")
        self.assertEqual(response.data["name"], "显示器")

        updated = self.client.patch(
            f"/api/v1/assets/{response.data['id']}/",
            {"purchase_date": "2022-06-18"},
            format="json",
        )
        self.assertEqual(updated.status_code, 200)
        self.assertEqual(updated.data["asset_tag"], "IT-MN-2022-001")

        second = self.client.post(
            "/api/v1/assets/",
            {"category": self.category.pk, "purchase_date": "2024-01-01"},
            format="json",
        )
        self.assertEqual(second.status_code, 201)
        self.assertEqual(second.data["asset_tag"], "IT-MN-2024-002")

    def test_asset_page_size_is_capped_at_100(self):
        Asset.objects.bulk_create(
            [
                Asset(
                    asset_tag=f"IT-MN-PAGE-{index:03d}",
                    name=f"分页资产 {index}",
                    category=self.category,
                )
                for index in range(104)
            ]
        )
        response = self.client.get("/api/v1/assets/?page_size=500")
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data["count"], 105)
        self.assertEqual(len(response.data["results"]), 100)

    def test_asset_list_defaults_to_newest_purchase_date_with_missing_dates_last(self):
        older_asset = Asset.objects.create(
            asset_tag="IT-MN-SORT-001",
            name="较早采购资产",
            category=self.category,
            purchase_date=date(2024, 1, 1),
        )
        newer_asset = Asset.objects.create(
            asset_tag="IT-MN-SORT-002",
            name="较新采购资产",
            category=self.category,
            purchase_date=date(2025, 1, 1),
        )

        response = self.client.get("/api/v1/assets/")

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            [item["id"] for item in response.data["results"]],
            [newer_asset.id, older_asset.id, self.asset.id],
        )

    def test_asset_status_can_be_added_in_settings_and_used(self):
        created = self.client.post(
            "/api/v1/asset-statuses/",
            {"name": "维修中", "code": "repairing", "is_active": True},
            format="json",
        )
        self.assertEqual(created.status_code, 201)
        lookups = self.client.get("/api/v1/lookups/")
        self.assertIn(
            {"value": "repairing", "label": "维修中"},
            lookups.data["statuses"],
        )
        updated = self.client.patch(
            f"/api/v1/assets/{self.asset.pk}/",
            {"status": "repairing"},
            format="json",
        )
        self.assertEqual(updated.status_code, 200)
        self.assertEqual(updated.data["status_label"], "维修中")

    def test_responsible_person_is_removed_from_asset_data(self):
        response = self.client.patch(
            f"/api/v1/assets/{self.asset.pk}/",
            {"custom_data": {"system_code": "OLD-001", "responsible_person": "旧责任人"}},
            format="json",
        )
        self.assertEqual(response.status_code, 200)
        self.assertNotIn("responsible_person", response.data["custom_data"])
        self.asset.refresh_from_db()
        self.assertNotIn("responsible_person", self.asset.custom_data)

    @patch("assets.views.nextcloud_storage.delete")
    @patch("assets.views.nextcloud_storage.upload")
    def test_asset_image_upload_list_and_delete(self, storage_upload, storage_delete):
        upload = SimpleUploadedFile(
            "设备正面.png",
            b"\x89PNG\r\n\x1a\nasset-image",
            content_type="image/png",
        )
        created = self.client.post(
            f"/api/v1/assets/{self.asset.pk}/images/",
            {"file": upload},
            format="multipart",
        )
        self.assertEqual(created.status_code, 201)
        self.assertTrue(created.data["is_cover"])
        self.assertEqual(created.data["original_name"], "设备正面.png")
        storage_upload.assert_called_once()
        image = AssetImage.objects.get(pk=created.data["id"])
        self.assertTrue(image.remote_path.startswith("/AffairsOS/assets/"))

        detail = self.client.get(f"/api/v1/assets/{self.asset.pk}/")
        self.assertEqual(detail.data["images"][0]["id"], image.id)

        deleted = self.client.delete(
            f"/api/v1/assets/{self.asset.pk}/images/{image.pk}/"
        )
        self.assertEqual(deleted.status_code, 204)
        storage_delete.assert_called_once_with(image.remote_path)
        self.assertFalse(AssetImage.objects.filter(pk=image.pk).exists())

    @patch("assets.views.nextcloud_storage.upload")
    def test_asset_image_rejects_non_image_content_type(self, storage_upload):
        upload = SimpleUploadedFile("伪装图片.png", b"not-an-image", content_type="text/plain")
        response = self.client.post(
            f"/api/v1/assets/{self.asset.pk}/images/",
            {"file": upload},
            format="multipart",
        )
        self.assertEqual(response.status_code, 400)
        storage_upload.assert_not_called()


class AssetExcelImportTests(TestCase):
    def setUp(self):
        self.admin = User.objects.create_user("admin", password="pass", is_staff=True, is_superuser=True)
        self.employee = User.objects.create_user(
            "zhangsan",
            password="pass",
            first_name="张三",
        )
        self.department = Department.objects.create(name="信息技术部", code="IT")
        EmployeeProfile.objects.create(
            user=self.employee,
            employee_no="E001",
            department=self.department,
        )
        self.client = APIClient()
        self.client.force_authenticate(self.admin)

    def workbook_upload(self, rows, headers=None):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "IT资产统计表"
        sheet.append(
            headers
            or [
                "编码",
                "系统编码",
                "金蝶编码",
                "责任人",
                "使用人",
                "状态",
                "部门",
                "资产类型",
                "资产位置",
                "品牌",
                "型号",
                "数量",
                "设备序列号",
                "规格配置",
                "CPU",
                "硬盘大小(G)",
                "内存(G)",
                "有线MAC地址",
                "无线MAC地址",
                "购买日期",
                "备注",
            ]
        )
        for row in rows:
            sheet.append(row)
        output = io.BytesIO()
        workbook.save(output)
        return SimpleUploadedFile(
            "assets.xlsx",
            output.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    def test_legacy_name_column_falls_back_to_category(self):
        headers = ["编码", "名称", "品牌", "型号", "设备序列号"]
        row = ["", "笔记本电脑", "Lenovo", "ThinkBook 14", "SN-NAME-001"]
        preview = self.client.post(
            "/api/v1/assets/import/",
            {"file": self.workbook_upload([row], headers=headers)},
            format="multipart",
        )
        self.assertEqual(preview.status_code, 200)
        self.assertEqual(preview.data["rows"][0]["name"], "Lenovo ThinkBook 14")
        self.assertEqual(preview.data["rows"][0]["category"], "笔记本电脑")
        self.assertIn("已将名称", preview.data["rows"][0]["warnings"][0])

        result = self.client.post(
            "/api/v1/assets/import/",
            {"file": self.workbook_upload([row], headers=headers), "commit": "true"},
            format="multipart",
        )
        self.assertEqual(result.status_code, 200)
        asset = Asset.objects.get(serial_number="SN-NAME-001")
        self.assertEqual(asset.name, "Lenovo ThinkBook 14")
        self.assertEqual(asset.category.name, "笔记本电脑")

    def test_import_ignores_source_asset_tag_and_generates_one(self):
        headers = ["编码", "资产名称", "资产类型", "型号", "设备序列号"]
        row = ["LEGACY-001", "导入电脑", "笔记本电脑", "ThinkBook 14", "SN-AUTO-001"]
        preview = self.client.post(
            "/api/v1/assets/import/",
            {"file": self.workbook_upload([row], headers=headers)},
            format="multipart",
        )
        self.assertEqual(preview.status_code, 200)
        self.assertEqual(preview.data["rows"][0]["asset_tag"], "自动生成")
        self.assertIn("原编码仅作参考", preview.data["rows"][0]["warnings"][0])

        result = self.client.post(
            "/api/v1/assets/import/",
            {"file": self.workbook_upload([row], headers=headers), "commit": "true"},
            format="multipart",
        )
        self.assertEqual(result.status_code, 200)
        asset = Asset.objects.get(serial_number="SN-AUTO-001")
        self.assertRegex(asset.asset_tag, r"^IT-[A-Z]{2}-\d{4}-001$")
        self.assertNotEqual(asset.asset_tag, "LEGACY-001")
        self.assertEqual(asset.custom_data["import_original_asset_tag"], "LEGACY-001")

    def test_preview_then_import_with_automatic_tag(self):
        row = [
            "",
            "QC-DZ-26-001",
            "KD-IT-001",
            "贺凯旋",
            "E001",
            "在用",
            "信息技术部",
            "笔记本电脑",
            "个人工位",
            "Lenovo",
            "ThinkBook 14",
            1,
            "SN-001",
            "i5/16G/1T",
            "i5-13500H",
            "1T",
            "16G",
            "AA-BB-CC",
            "DD-EE-FF",
            date(2026, 7, 23),
            "测试导入",
        ]
        preview = self.client.post(
            "/api/v1/assets/import/",
            {"file": self.workbook_upload([row])},
            format="multipart",
        )
        self.assertEqual(preview.status_code, 200)
        self.assertEqual(preview.data["invalid"], 0)
        self.assertEqual(preview.data["rows"][0]["asset_tag"], "自动生成")

        result = self.client.post(
            "/api/v1/assets/import/",
            {"file": self.workbook_upload([row]), "commit": "true"},
            format="multipart",
        )
        self.assertEqual(result.status_code, 200)
        self.assertEqual(result.data["created"], 1)
        asset = Asset.objects.get(serial_number="SN-001")
        self.assertRegex(asset.asset_tag, r"^IT-[A-Z]{2}-\d{4}-001$")
        self.assertEqual(asset.kingdee_code, "KD-IT-001")
        self.assertEqual(asset.assigned_to, self.employee)
        self.assertEqual(asset.custodian_department, self.department)
        self.assertEqual(asset.custom_data["system_code"], "QC-DZ-26-001")
        self.assertEqual(asset.events.count(), 1)

    def test_invalid_row_does_not_write_any_assets(self):
        row = [
            "",
            "",
            "",
            "",
            "不存在的人",
            "在用",
            "",
            "笔记本电脑",
            "",
            "",
            "",
            2,
        ]
        response = self.client.post(
            "/api/v1/assets/import/",
            {"file": self.workbook_upload([row]), "commit": "true"},
            format="multipart",
        )
        self.assertEqual(response.status_code, 400)
        self.assertEqual(Asset.objects.count(), 0)

    def test_missing_category_and_user_can_import_as_pending_remediation(self):
        row = [
            "",
            "",
            "",
            "",
            "历史员工",
            "借用中",
            "",
            "",
            "",
            "Lenovo",
            "E480",
            1,
            "SN-WARNING-001",
        ]
        preview = self.client.post(
            "/api/v1/assets/import/",
            {"file": self.workbook_upload([row])},
            format="multipart",
        )
        self.assertEqual(preview.status_code, 200)
        self.assertEqual(preview.data["invalid"], 0)
        self.assertEqual(preview.data["warning"], 1)

        result = self.client.post(
            "/api/v1/assets/import/",
            {"file": self.workbook_upload([row]), "commit": "true"},
            format="multipart",
        )
        self.assertEqual(result.status_code, 200)
        asset = Asset.objects.get(serial_number="SN-WARNING-001")
        self.assertEqual(asset.category.code, "UC")
        self.assertEqual(asset.status, Asset.Status.PENDING)
        self.assertIsNone(asset.assigned_to)
        self.assertTrue(asset.custom_data["import_warnings"])

    def test_asset_template_keeps_responsible_person_and_kingdee_code_only(self):
        response = self.client.get("/api/v1/assets/import-template/")
        self.assertEqual(response.status_code, 200)
        workbook = load_workbook(io.BytesIO(b"".join(response.streaming_content)), data_only=True)
        sheet = workbook["资产导入"]
        headers = [cell.value for cell in sheet[1] if cell.value is not None]
        self.assertIn("责任人", headers)
        self.assertIn("金蝶编码", headers)
        self.assertIn("资产分类", headers)
        self.assertIn("资产类型", headers)
        self.assertIn("状态", headers)
        self.assertNotIn("部门", headers)
        self.assertNotIn("使用人", headers)
        self.assertNotIn("系统编码", headers)
        self.assertEqual(len(headers), 18)

    def test_import_separates_asset_classification_and_type(self):
        headers = ["资产分类", "资产类型", "品牌", "型号", "设备序列号", "数量"]
        row = ["行政资产", "办公家具", "震旦", "办公桌", "ADMIN-001", 1]
        preview = self.client.post(
            "/api/v1/assets/import/",
            {"file": self.workbook_upload([row], headers=headers)},
            format="multipart",
        )
        self.assertEqual(preview.status_code, 200)
        self.assertEqual(preview.data["invalid"], 0)
        self.assertEqual(preview.data["rows"][0]["class_type"], "行政资产")
        self.assertEqual(preview.data["rows"][0]["category"], "办公家具")

        result = self.client.post(
            "/api/v1/assets/import/",
            {"file": self.workbook_upload([row], headers=headers), "commit": "true"},
            format="multipart",
        )
        self.assertEqual(result.status_code, 200)
        asset = Asset.objects.get(serial_number="ADMIN-001")
        self.assertEqual(asset.category.class_type, AssetCategory.ClassType.ADMIN)
        self.assertEqual(asset.category.name, "办公家具")
        self.assertRegex(asset.asset_tag, r"^AD-[A-Z]{2}-\d{4}-001$")

    def test_legacy_asset_category_header_still_means_asset_type(self):
        headers = ["资产分类", "品牌", "型号", "设备序列号", "数量"]
        row = ["笔记本电脑", "Lenovo", "ThinkBook 14", "LEGACY-CATEGORY-001", 1]
        preview = self.client.post(
            "/api/v1/assets/import/",
            {"file": self.workbook_upload([row], headers=headers)},
            format="multipart",
        )
        self.assertEqual(preview.status_code, 200)
        self.assertEqual(preview.data["rows"][0]["category"], "笔记本电脑")
        self.assertEqual(preview.data["rows"][0]["class_type"], "IT资产")

    def test_import_accepts_multiple_wired_mac_addresses(self):
        headers = ["资产类型", "金蝶编码", "责任人", "状态", "有线MAC地址", "设备序列号"]
        wired_mac = "LAN1:EC-D6-8A-4F-AB-12\nLAN2:C0-8A-CD-D2-A8-79"
        row = ["网络设备", "KD-NET-001", "E001", "使用中", wired_mac, "SN-NET-001"]
        preview = self.client.post(
            "/api/v1/assets/import/",
            {"file": self.workbook_upload([row], headers=headers)},
            format="multipart",
        )
        self.assertEqual(preview.status_code, 200)
        self.assertEqual(preview.data["invalid"], 0)
        result = self.client.post(
            "/api/v1/assets/import/",
            {"file": self.workbook_upload([row], headers=headers), "commit": "true"},
            format="multipart",
        )
        self.assertEqual(result.status_code, 200)
        self.assertEqual(Asset.objects.get(serial_number="SN-NET-001").wired_mac, wired_mac)


class AssetResequenceCommandTests(TestCase):
    def test_resequence_updates_tags_history_and_sequence(self):
        it_category = AssetCategory.objects.create(name="笔记本", code="LT")
        admin_category = AssetCategory.objects.create(
            name="办公家具",
            code="FU",
            class_type=AssetCategory.ClassType.ADMIN,
        )
        first = Asset.objects.create(
            asset_tag="OLD-001",
            name="电脑一",
            category=it_category,
            purchase_date=date(2024, 5, 1),
        )
        second = Asset.objects.create(
            asset_tag="OLD-002",
            name="电脑二",
            category=it_category,
            purchase_date=date(2022, 1, 1),
        )
        furniture = Asset.objects.create(
            asset_tag="OLD-003",
            name="办公桌",
            category=admin_category,
            purchase_date=date(2020, 8, 8),
        )

        with TemporaryDirectory() as output_dir:
            output_path = f"{output_dir}/asset_tag_mapping.csv"
            call_command("resequence_asset_tags", output=output_path)
            with open(output_path, encoding="utf-8-sig") as mapping_file:
                self.assertEqual(len(mapping_file.readlines()), 4)

        first.refresh_from_db()
        second.refresh_from_db()
        furniture.refresh_from_db()
        self.assertEqual(first.asset_tag, "IT-LT-2024-002")
        self.assertEqual(second.asset_tag, "IT-LT-2022-001")
        self.assertEqual(furniture.asset_tag, "AD-FU-2020-001")
        self.assertIn("OLD-001", first.custom_data["previous_asset_tags"])
        self.assertEqual(
            AssetNumberSequence.objects.get(category=it_category).current_value,
            2,
        )
        self.assertTrue(
            first.events.filter(metadata__old_asset_tag="OLD-001").exists()
        )

    def test_normalize_asset_categories_merges_and_resequences_assets(self):
        categories = {
            name: AssetCategory.objects.create(name=name, code=code)
            for name, code in [
                ("笔记本电脑", "LI"),
                ("台式机", "WL"),
                ("显示器", "FE"),
                ("显示屏", "NW"),
                ("会议设备", "XK"),
                ("录像机", "ZO"),
                ("服务器", "XF"),
                ("交换机", "CT"),
                ("AP", "AP"),
            ]
        }
        assets = [
            Asset.objects.create(
                asset_tag=old_tag,
                name=name,
                category=categories[category_name],
                purchase_date=purchase_date,
            )
            for old_tag, name, category_name, purchase_date in [
                ("OLD-AP", "无线接入点", "AP", date(2025, 3, 1)),
                ("OLD-SWITCH", "交换机", "交换机", date(2022, 2, 1)),
                ("OLD-MONITOR", "显示器", "显示器", date(2021, 1, 1)),
                ("OLD-SCREEN", "显示屏", "显示屏", date(2022, 1, 1)),
                ("OLD-RECORDER", "录像机", "录像机", date(2021, 1, 1)),
                ("OLD-SERVER", "服务器", "服务器", date(2022, 1, 1)),
            ]
        ]
        assets[0].custom_data = {
            "previous_asset_tags": ["LEGACY-AP"],
            "import_original_asset_tag": "IMPORT-AP",
        }
        assets[0].save(update_fields=["custom_data", "updated_at"])
        AssetEvent.objects.create(
            asset=assets[0],
            action=AssetEvent.Action.UPDATED,
            metadata={"old_asset_tag": "LEGACY-AP", "new_asset_tag": "OLD-AP"},
        )

        call_command("normalize_asset_categories")

        self.assertEqual(
            dict(
                AssetCategory.objects.filter(is_active=True).values_list("name", "code")
            ),
            {
                "笔记本电脑": "LT",
                "台式机": "DT",
                "显示屏": "MN",
                "会议设备": "AV",
                "网络设备": "NW",
                "服务器": "SV",
            },
        )
        self.assertFalse(
            AssetCategory.objects.filter(name__in=["AP", "显示器", "录像机"]).exists()
        )
        for asset in assets:
            asset.refresh_from_db()
            self.assertIn(asset.asset_tag.split("-")[1], {"MN", "NW", "SV"})
            self.assertNotIn("previous_asset_tags", asset.custom_data)
            self.assertNotIn("import_original_asset_tag", asset.custom_data)
        self.assertFalse(AssetEvent.objects.filter(metadata__has_key="old_asset_tag").exists())

        network_tags = list(
            Asset.objects.filter(category__name="网络设备")
            .order_by("purchase_date")
            .values_list("asset_tag", flat=True)
        )
        self.assertEqual(network_tags, ["IT-NW-2022-001", "IT-NW-2025-002"])

        call_command("normalize_asset_categories")
        self.assertEqual(
            Asset.objects.filter(category__name="网络设备").count(),
            2,
        )


class PeopleImportTests(TestCase):
    def test_import_people_creates_department_and_non_login_employee(self):
        payload = {
            "departments": [
                {"source_id": "10", "name": "总部", "parent_source_id": None},
                {"source_id": "11", "name": "信息技术部", "parent_source_id": "10"},
            ],
            "users": [
                {
                    "username": "zhangsan",
                    "email": "zhangsan@example.com",
                    "full_name": "张三",
                    "department_source_ids": ["11"],
                    "mobile": "13800000000",
                    "active": True,
                }
            ],
        }
        with patch("sys.stdin", io.StringIO(json.dumps(payload, ensure_ascii=False))):
            call_command("import_people", input="-")

        user = User.objects.get(username="zhangsan")
        profile = EmployeeProfile.objects.get(user=user)
        self.assertFalse(user.has_usable_password())
        self.assertEqual(profile.department.name, "信息技术部")
        self.assertEqual(Department.objects.get(code="263-11").parent.code, "263-10")

    def test_import_people_syncs_assigned_asset_department(self):
        old_department = Department.objects.create(name="旧部门", code="OLD")
        category = AssetCategory.objects.create(name="笔记本电脑", code="LT")
        user = User.objects.create_user("zhangsan")
        EmployeeProfile.objects.create(
            user=user,
            employee_no="zhangsan",
            department=old_department,
        )
        asset = Asset.objects.create(
            asset_tag="IT-LT-2026-001",
            category=category,
            assigned_to=user,
            custodian_department=old_department,
        )
        payload = {
            "departments": [
                {"source_id": "11", "name": "信息技术部", "parent_source_id": None},
            ],
            "users": [
                {
                    "username": "zhangsan",
                    "email": "zhangsan@example.com",
                    "full_name": "张三",
                    "department_source_ids": ["11"],
                    "active": True,
                }
            ],
        }

        with patch("sys.stdin", io.StringIO(json.dumps(payload, ensure_ascii=False))):
            call_command("import_people", input="-")

        asset.refresh_from_db()
        self.assertEqual(asset.custodian_department.name, "信息技术部")


class InventoryWorkflowTests(TestCase):
    def setUp(self):
        self.admin = User.objects.create_user("admin", password="pass", is_staff=True, is_superuser=True)
        self.employee = User.objects.create_user("employee", password="pass")
        self.location = Location.objects.create(name="IT 库房", code="IT-WH", kind="warehouse")
        self.client = APIClient()
        self.client.force_authenticate(self.admin)

    def test_create_issue_and_prevent_negative_stock(self):
        created = self.client.post(
            "/api/v1/inventory/",
            {
                "sku": "KB-001",
                "name": "无线键盘",
                "kind": "accessory",
                "unit": "个",
                "initial_quantity": 10,
                "minimum_quantity": 2,
                "location": self.location.pk,
            },
            format="json",
        )
        self.assertEqual(created.status_code, 201)
        self.assertEqual(created.data["quantity"], 10)
        item_id = created.data["id"]

        issued = self.client.post(
            f"/api/v1/inventory/{item_id}/transactions/",
            {
                "action": "issue",
                "quantity": 3,
                "recipient_id": self.employee.pk,
                "notes": "新员工领用",
            },
            format="json",
        )
        self.assertEqual(issued.status_code, 200)
        self.assertEqual(issued.data["quantity"], 7)
        self.assertEqual(InventoryTransaction.objects.count(), 2)

        rejected = self.client.post(
            f"/api/v1/inventory/{item_id}/transactions/",
            {"action": "issue", "quantity": 8},
            format="json",
        )
        self.assertEqual(rejected.status_code, 400)
        self.assertEqual(InventoryItem.objects.get(pk=item_id).quantity, 7)

    def test_low_stock_only_when_quantity_is_below_minimum(self):
        equal_item = InventoryItem.objects.create(
            sku="EQUAL-001",
            name="库存刚好充足",
            quantity=5,
            minimum_quantity=5,
        )
        low_item = InventoryItem.objects.create(
            sku="LOW-001",
            name="库存不足",
            quantity=4,
            minimum_quantity=5,
        )

        response = self.client.get("/api/v1/inventory/")
        self.assertEqual(response.status_code, 200)
        rows = {row["id"]: row for row in response.data}
        self.assertFalse(rows[equal_item.id]["low_stock"])
        self.assertTrue(rows[low_item.id]["low_stock"])


class InventoryExcelImportExportTests(TestCase):
    def setUp(self):
        self.admin = User.objects.create_user(
            "admin",
            password="pass",
            is_staff=True,
            is_superuser=True,
        )
        self.client = APIClient()
        self.client.force_authenticate(self.admin)

    def workbook_upload(self, rows):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "库存导入"
        sheet.append(
            [
                "物品编码",
                "物品名称",
                "物品分类",
                "品牌",
                "型号",
                "数量",
                "单位",
                "单价",
                "采购途径",
                "保障数量",
                "存放地点",
                "备注",
            ]
        )
        for row in rows:
            sheet.append(row)
        output = io.BytesIO()
        workbook.save(output)
        return SimpleUploadedFile(
            "inventory.xlsx",
            output.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    def test_import_creates_item_auto_sku_and_opening_transaction(self):
        row = [
            "",
            "无线鼠标",
            "配件",
            "Logitech",
            "M720",
            20,
            "个",
            189.5,
            "合作供应商",
            3,
            "IT 库房",
            "新员工备用",
        ]
        preview = self.client.post(
            "/api/v1/inventory/import/",
            {"file": self.workbook_upload([row])},
            format="multipart",
        )
        self.assertEqual(preview.status_code, 200)
        self.assertEqual(preview.data["create"], 1)
        self.assertEqual(preview.data["invalid"], 0)
        self.assertEqual(preview.data["rows"][0]["sku"], "自动生成")

        result = self.client.post(
            "/api/v1/inventory/import/",
            {"file": self.workbook_upload([row]), "commit": "true"},
            format="multipart",
        )
        self.assertEqual(result.status_code, 200)
        item = InventoryItem.objects.get(name="无线鼠标")
        self.assertRegex(item.sku, r"^INV-\d{4}-\d{3}$")
        self.assertEqual(item.brand, "Logitech")
        self.assertEqual(item.model_name, "M720")
        self.assertEqual(str(item.unit_price), "189.50")
        self.assertEqual(item.purchase_channel, InventoryItem.PurchaseChannel.SUPPLIER)
        self.assertEqual(item.quantity, 20)
        transaction_record = item.transactions.get()
        self.assertEqual(transaction_record.action, InventoryTransaction.Action.INBOUND)
        self.assertEqual(transaction_record.quantity, 20)
        self.assertEqual(transaction_record.balance_after, 20)

    def test_import_existing_sku_reconciles_quantity_with_writeoff(self):
        item = InventoryItem.objects.create(
            sku="KB-001",
            name="键盘",
            kind=InventoryItem.Kind.ACCESSORY,
            quantity=10,
            unit="个",
        )
        row = ["KB-001", "无线键盘", "配件", "Logitech", "K380", 7, "个", 169, "电商", 2, "", ""]
        result = self.client.post(
            "/api/v1/inventory/import/",
            {"file": self.workbook_upload([row]), "commit": "true"},
            format="multipart",
        )
        self.assertEqual(result.status_code, 200)
        self.assertEqual(result.data["updated"], 1)
        item.refresh_from_db()
        self.assertEqual(item.name, "无线键盘")
        self.assertEqual(item.purchase_channel, InventoryItem.PurchaseChannel.ECOMMERCE)
        self.assertEqual(item.quantity, 7)
        movement = item.transactions.get()
        self.assertEqual(movement.action, InventoryTransaction.Action.WRITEOFF)
        self.assertEqual(movement.quantity, 3)
        self.assertEqual(movement.balance_after, 7)

    def test_import_rejects_invalid_quantity_without_writing(self):
        row = ["", "A4 纸", "耗材", "", "", "3.5", "包", 25, "其他", 2, "", ""]
        preview = self.client.post(
            "/api/v1/inventory/import/",
            {"file": self.workbook_upload([row])},
            format="multipart",
        )
        self.assertEqual(preview.status_code, 200)
        self.assertEqual(preview.data["invalid"], 1)
        result = self.client.post(
            "/api/v1/inventory/import/",
            {"file": self.workbook_upload([row]), "commit": "true"},
            format="multipart",
        )
        self.assertEqual(result.status_code, 400)
        self.assertFalse(InventoryItem.objects.exists())

    def test_template_and_export_include_inventory_columns(self):
        InventoryItem.objects.create(
            sku="PAPER-001",
            name="A4 纸",
            kind=InventoryItem.Kind.CONSUMABLE,
            brand="Double A",
            model_name="80g",
            unit="包",
            unit_price="25.00",
            purchase_channel=InventoryItem.PurchaseChannel.SUPPLIER,
            quantity=6,
            minimum_quantity=2,
        )
        template_response = self.client.get("/api/v1/inventory/import-template/")
        self.assertEqual(template_response.status_code, 200)
        template_bytes = b"".join(template_response.streaming_content)
        self.assertGreater(len(template_bytes), 1000)
        template_sheet = load_workbook(io.BytesIO(template_bytes), data_only=True).active
        self.assertEqual(
            [cell.value for cell in template_sheet[1]],
            ["物品编码", "物品名称", "物品分类", "品牌", "型号", "数量", "单位", "单价", "采购途径", "保障数量", "存放地点", "备注"],
        )

        export_response = self.client.get("/api/v1/inventory/export/")
        self.assertEqual(export_response.status_code, 200)
        exported = io.BytesIO(export_response.content)
        sheet = load_workbook(exported, data_only=True).active
        self.assertEqual(
            [cell.value for cell in sheet[1]][:11],
            ["物品编码", "物品名称", "物品分类", "品牌", "型号", "当前数量", "单位", "单价", "库存金额", "采购途径", "保障数量"],
        )
        self.assertEqual(sheet["A2"].value, "PAPER-001")
        self.assertEqual(sheet["I2"].value, 150)
        self.assertEqual(sheet["J2"].value, "合作供应商")

    def test_purchase_export_only_contains_shortages(self):
        InventoryItem.objects.create(
            sku="LOW-001",
            name="无线鼠标",
            kind=InventoryItem.Kind.ACCESSORY,
            unit="个",
            unit_price="80.00",
            purchase_channel=InventoryItem.PurchaseChannel.ECOMMERCE,
            quantity=2,
            minimum_quantity=5,
        )
        InventoryItem.objects.create(
            sku="OK-001",
            name="键盘",
            kind=InventoryItem.Kind.ACCESSORY,
            unit="个",
            unit_price="100.00",
            purchase_channel=InventoryItem.PurchaseChannel.SUPPLIER,
            quantity=5,
            minimum_quantity=5,
        )

        response = self.client.get("/api/v1/inventory/purchase-export/")
        self.assertEqual(response.status_code, 200)
        self.assertIn("filename*=UTF-8''", response["Content-Disposition"])
        sheet = load_workbook(io.BytesIO(response.content), data_only=True).active
        self.assertEqual(sheet.max_row, 2)
        self.assertEqual(sheet["A2"].value, "LOW-001")
        self.assertEqual(sheet["H2"].value, 3)
        self.assertEqual(sheet["K2"].value, 240)
        self.assertEqual(sheet["L2"].value, "电商")


class StocktakeWorkflowTests(TestCase):
    def setUp(self):
        self.admin = User.objects.create_user("admin", password="pass", is_staff=True, is_superuser=True)
        self.category = AssetCategory.objects.create(name="笔记本", code="LT")
        self.location = Location.objects.create(name="办公室", code="OFFICE")
        self.first = Asset.objects.create(
            asset_tag="IT-LT-001",
            name="电脑 1",
            category=self.category,
            current_location=self.location,
        )
        self.second = Asset.objects.create(
            asset_tag="IT-LT-002",
            name="电脑 2",
            category=self.category,
            current_location=self.location,
        )
        self.client = APIClient()
        self.client.force_authenticate(self.admin)

    def test_snapshot_scan_and_complete_marks_missing(self):
        created = self.client.post(
            "/api/v1/stocktakes/",
            {"name": "办公室盘点", "scope_location": self.location.pk},
            format="json",
        )
        self.assertEqual(created.status_code, 201)
        self.assertEqual(created.data["snapshot_count"], 2)
        task_id = created.data["id"]

        scanned = self.client.post(
            f"/api/v1/stocktakes/{task_id}/scan/",
            {"asset_tag": self.first.asset_tag},
            format="json",
        )
        self.assertEqual(scanned.status_code, 200)
        self.assertEqual(scanned.data["scanned_count"], 1)

        completed = self.client.post(
            f"/api/v1/stocktakes/{task_id}/complete/",
            {},
            format="json",
        )
        self.assertEqual(completed.status_code, 200)
        self.assertEqual(completed.data["missing_count"], 1)
        self.assertEqual(
            StocktakeRecord.objects.get(task_id=task_id, asset=self.second).result,
            StocktakeRecord.Result.MISSING,
        )
        self.assertEqual(StocktakeTask.objects.get(pk=task_id).status, "completed")


class SettingsAndReportsTests(TestCase):
    def setUp(self):
        self.admin = User.objects.create_user("admin", password="pass", is_staff=True, is_superuser=True)
        self.client = APIClient()
        self.client.force_authenticate(self.admin)

    def test_settings_create_toggle_and_reports(self):
        created = self.client.post(
            "/api/v1/categories/",
            {"name": "平板电脑", "code": "TB", "icon": "box", "custom_fields": []},
            format="json",
        )
        self.assertEqual(created.status_code, 201)
        toggled = self.client.patch(
            f"/api/v1/categories/{created.data['id']}/",
            {"is_active": False},
            format="json",
        )
        self.assertEqual(toggled.status_code, 200)
        self.assertFalse(toggled.data["is_active"])
        reports = self.client.get("/api/v1/reports/")
        self.assertEqual(reports.status_code, 200)
        self.assertIn("quality", reports.data)

    def test_report_excludes_inventory_at_guarantee_quantity(self):
        InventoryItem.objects.create(
            sku="EQUAL-001",
            name="库存刚好充足",
            quantity=5,
            minimum_quantity=5,
        )
        low_item = InventoryItem.objects.create(
            sku="LOW-001",
            name="库存不足",
            quantity=4,
            minimum_quantity=5,
        )

        reports = self.client.get("/api/v1/reports/")
        self.assertEqual(reports.status_code, 200)
        self.assertEqual(
            [row["id"] for row in reports.data["low_stock"]],
            [low_item.id],
        )


class ScopedPermissionTests(TestCase):
    def setUp(self):
        self.user = User.objects.create_user("inventory_manager", password="pass")
        AssetManagerRole.objects.create(user=self.user, scopes=["inventory"])
        self.client = APIClient()
        self.client.force_authenticate(self.user)

    def test_manager_only_opens_assigned_modules(self):
        self.assertEqual(self.client.get("/api/v1/inventory/").status_code, 200)
        self.assertEqual(self.client.get("/api/v1/assets/").status_code, 403)
        self.assertEqual(self.client.get("/api/v1/reports/").status_code, 403)
        response = self.client.get("/api/v1/auth/me/")
        self.assertEqual(response.data["management_scopes"], ["inventory"])

    def test_inventory_manager_can_process_asset_requests(self):
        requester = User.objects.create_user("requester", password="pass")
        category = AssetCategory.objects.create(name="显示器", code="MN")
        asset = Asset.objects.create(
            asset_tag="IT-MN-2026-201",
            category=category,
            brand="Dell",
            model_name="U2720Q",
            status=Asset.Status.AVAILABLE,
        )
        asset_request = AssetRequest.objects.create(
            requester=requester,
            request_type=AssetRequest.RequestType.LOAN,
            requested_name="显示器",
            expected_return_at=date.today() + timedelta(days=3),
        )

        listed = self.client.get("/api/v1/requests/")
        self.assertEqual([row["id"] for row in listed.data], [asset_request.id])
        candidates = self.client.get(f"/api/v1/requests/{asset_request.id}/candidates/")
        self.assertEqual(candidates.status_code, 200)
        fulfilled = self.client.post(
            f"/api/v1/requests/{asset_request.id}/fulfill/",
            {"asset_id": asset.id},
            format="json",
        )
        self.assertEqual(fulfilled.status_code, 200)
        self.assertEqual(fulfilled.data["status"], AssetRequest.Status.FULFILLED)
        asset.refresh_from_db()
        self.assertEqual(asset.status, Asset.Status.LOANED)
        self.assertEqual(asset.assigned_to, requester)

    def test_manager_can_switch_to_own_request_list(self):
        other = User.objects.create_user("other_requester", password="pass")
        own_request = AssetRequest.objects.create(
            requester=self.user,
            request_type=AssetRequest.RequestType.ASSIGN,
            requested_name="笔记本电脑",
        )
        AssetRequest.objects.create(
            requester=other,
            request_type=AssetRequest.RequestType.LOAN,
            requested_name="显示器",
            expected_return_at=date.today() + timedelta(days=2),
        )
        response = self.client.get("/api/v1/requests/?mine=1")
        self.assertEqual([row["id"] for row in response.data], [own_request.id])


class ReportAssetDetailTests(TestCase):
    def setUp(self):
        self.admin = User.objects.create_superuser("report_admin", password="pass")
        self.client = APIClient()
        self.client.force_authenticate(self.admin)
        self.uncategorized = AssetCategory.objects.create(name="待分类", code="UC")
        self.notebook = AssetCategory.objects.create(name="笔记本电脑", code="NB")
        self.location = Location.objects.create(name="IT 库房", code="IT-WH", kind="warehouse")
        self.department = Department.objects.create(name="人力资源部", code="HR")
        self.owner = User.objects.create_user("report_owner", password="pass")
        EmployeeProfile.objects.create(
            user=self.owner,
            employee_no="REPORT-001",
            department=self.department,
        )
        self.asset = Asset.objects.create(
            asset_tag="IT-UC-2026-001",
            category=self.uncategorized,
            assigned_to=self.owner,
            custom_data={"import_warnings": ["缺少资产类型", "责任人待确认"]},
        )

    def test_quality_and_department_metrics_open_asset_details(self):
        report = self.client.get("/api/v1/reports/")
        department_row = report.data["by_department"][0]
        self.assertEqual(department_row["custodian_department_id"], self.department.id)

        quality = self.client.get("/api/v1/reports/assets/?kind=import_warnings")
        self.assertEqual(quality.status_code, 200)
        self.assertEqual(quality.data["results"][0]["id"], self.asset.id)
        self.assertEqual(quality.data["results"][0]["import_warnings"], ["缺少资产类型", "责任人待确认"])

        department = self.client.get(f"/api/v1/reports/assets/?kind=department&department_id={self.department.id}")
        self.assertEqual(department.data["title"], "人力资源部")
        self.assertEqual(department.data["count"], 1)

    def test_batch_completes_category_location_serial_and_import_warning(self):
        category = self.client.post(
            "/api/v1/reports/assets/",
            {"kind": "missing_category", "asset_ids": [self.asset.id], "category_id": self.notebook.id},
            format="json",
        )
        self.assertEqual(category.status_code, 200)
        self.asset.refresh_from_db()
        self.assertEqual(self.asset.category, self.notebook)

        location = self.client.post(
            "/api/v1/reports/assets/",
            {"kind": "missing_location", "asset_ids": [self.asset.id], "location_id": self.location.id},
            format="json",
        )
        self.assertEqual(location.status_code, 200)
        serial = self.client.post(
            "/api/v1/reports/assets/",
            {"kind": "missing_serial", "asset_ids": [self.asset.id], "serial_numbers": {str(self.asset.id): "SN-001"}},
            format="json",
        )
        self.assertEqual(serial.status_code, 200)
        resolved = self.client.post(
            "/api/v1/reports/assets/",
            {"kind": "import_warnings", "asset_ids": [self.asset.id]},
            format="json",
        )
        self.assertEqual(resolved.status_code, 200)
        self.asset.refresh_from_db()
        self.assertEqual(self.asset.current_location, self.location)
        self.assertEqual(self.asset.serial_number, "SN-001")
        self.assertNotIn("import_warnings", self.asset.custom_data)
        self.assertEqual(self.asset.events.count(), 4)

    def test_department_detail_rejects_batch_assignment(self):
        target = Department.objects.create(name="行政部", code="ADMIN")
        response = self.client.post(
            "/api/v1/reports/assets/",
            {"kind": "department", "asset_ids": [self.asset.id], "department_id": target.id},
            format="json",
        )
        self.assertEqual(response.status_code, 400)
        self.assertIn("责任人自动确定", response.data["message"])
        self.asset.refresh_from_db()
        self.assertEqual(self.asset.custodian_department, self.department)
        self.assertFalse(self.asset.events.exists())


class AssetRequestWorkflowTests(TestCase):
    def setUp(self):
        self.admin = User.objects.create_superuser("admin", password="pass")
        self.employee = User.objects.create_user("employee", password="pass", first_name="员工")
        self.category = AssetCategory.objects.create(name="笔记本电脑", code="LT")
        self.asset = Asset.objects.create(
            asset_tag="IT-LT-2026-101",
            name="办公笔记本",
            category=self.category,
            brand="Lenovo",
            model_name="ThinkBook 14",
            status=Asset.Status.AVAILABLE,
        )
        self.client = APIClient()

    def test_employee_requests_name_and_manager_assigns_specific_asset(self):
        self.client.force_authenticate(self.employee)
        created = self.client.post(
            "/api/v1/requests/",
            {
                "request_type": "loan",
                "requested_name": "笔记本电脑",
                "needed_at": date.today(),
                "expected_return_at": date.today() + timedelta(days=7),
                "reason": "出差",
            },
            format="json",
        )
        self.assertEqual(created.status_code, 201)
        self.assertIsNone(created.data["assigned_asset"])

        self.client.force_authenticate(self.admin)
        candidates = self.client.get(f"/api/v1/requests/{created.data['id']}/candidates/")
        self.assertEqual([row["id"] for row in candidates.data], [self.asset.id])
        fulfilled = self.client.post(
            f"/api/v1/requests/{created.data['id']}/fulfill/",
            {"asset_id": self.asset.id},
            format="json",
        )
        self.assertEqual(fulfilled.status_code, 200)

    def test_my_loaned_assets_uses_current_custodian_without_request_history(self):
        self.asset.status = Asset.Status.LOANED
        self.asset.assigned_to = self.employee
        self.asset.expected_return_at = date.today() + timedelta(days=5)
        self.asset.save()
        Asset.objects.create(
            asset_tag="IT-LT-2026-102",
            name="其他人的电脑",
            category=self.category,
            status=Asset.Status.LOANED,
            assigned_to=self.admin,
        )
        self.client.force_authenticate(self.employee)

        response = self.client.get("/api/v1/requests/my-loaned-assets/")

        self.assertEqual(response.status_code, 200)
        self.assertEqual([row["id"] for row in response.data], [self.asset.id])
        self.assertEqual(response.data[0]["assignee_name"], "员工")
        self.assertFalse(AssetRequest.objects.exists())

    def test_inventory_can_be_received_but_not_borrowed(self):
        item = InventoryItem.objects.create(
            sku="MOUSE-001",
            name="无线鼠标",
            kind=InventoryItem.Kind.ACCESSORY,
            unit="个",
            quantity=8,
        )
        self.client.force_authenticate(self.employee)
        options = self.client.get("/api/v1/requests/device-options/")
        inventory_option = next(row for row in options.data if row["key"] == f"inventory:{item.id}")
        self.assertEqual(inventory_option["available_count"], 8)

        rejected_loan = self.client.post(
            "/api/v1/requests/",
            {
                "request_type": "loan",
                "requested_item_type": "inventory",
                "requested_name": item.name,
                "inventory_item": item.id,
                "requested_quantity": 2,
                "needed_at": date.today(),
                "expected_return_at": date.today() + timedelta(days=2),
                "reason": "临时使用",
            },
            format="json",
        )
        self.assertEqual(rejected_loan.status_code, 400)

        created = self.client.post(
            "/api/v1/requests/",
            {
                "request_type": "assign",
                "requested_item_type": "inventory",
                "requested_name": item.name,
                "inventory_item": item.id,
                "requested_quantity": 2,
                "needed_at": date.today(),
                "reason": "办公使用",
            },
            format="json",
        )
        self.assertEqual(created.status_code, 201)
        self.assertEqual(created.data["requested_quantity"], 2)

        self.client.force_authenticate(self.admin)
        fulfilled = self.client.post(
            f"/api/v1/requests/{created.data['id']}/fulfill/",
            {},
            format="json",
        )
        self.assertEqual(fulfilled.status_code, 200)
        item.refresh_from_db()
        self.assertEqual(item.quantity, 6)
        transaction = InventoryTransaction.objects.get(pk=fulfilled.data["issued_inventory_transaction"])
        self.assertEqual(transaction.recipient, self.employee)
        self.assertEqual(transaction.quantity, 2)

    def test_assign_request_requires_needed_date_but_reason_is_optional(self):
        self.client.force_authenticate(self.employee)
        response = self.client.post(
            "/api/v1/requests/",
            {"request_type": "assign", "requested_name": "笔记本电脑"},
            format="json",
        )
        self.assertEqual(response.status_code, 400)
        self.assertIn("needed_at", response.data["errors"])
        self.assertNotIn("reason", response.data["errors"])

        created = self.client.post(
            "/api/v1/requests/",
            {
                "request_type": "assign",
                "requested_name": "笔记本电脑",
                "needed_at": date.today(),
                "reason": "",
            },
            format="json",
        )
        self.assertEqual(created.status_code, 201)
        self.assertEqual(created.data["reason"], "")

    def test_loan_request_still_requires_reason(self):
        self.client.force_authenticate(self.employee)
        response = self.client.post(
            "/api/v1/requests/",
            {
                "request_type": "loan",
                "requested_name": "笔记本电脑",
                "needed_at": date.today(),
                "expected_return_at": date.today() + timedelta(days=1),
                "reason": "",
            },
            format="json",
        )
        self.assertEqual(response.status_code, 400)
        self.assertIn("reason", response.data["errors"])

    def test_employee_cannot_assign_asset_directly(self):
        self.client.force_authenticate(self.employee)
        response = self.client.post(
            f"/api/v1/assets/{self.asset.id}/actions/",
            {"action": "assign", "target_user_id": self.employee.id},
            format="json",
        )
        self.assertEqual(response.status_code, 403)

    def test_non_requestable_asset_is_excluded_from_request_options(self):
        self.asset.is_requestable = False
        self.asset.save(update_fields=["is_requestable", "updated_at"])
        self.client.force_authenticate(self.employee)
        options = self.client.get("/api/v1/requests/device-options/")
        self.assertFalse(any(row["name"] == "笔记本电脑" for row in options.data))
        response = self.client.post(
            "/api/v1/requests/",
            {
                "request_type": "assign",
                "requested_item_type": "asset",
                "requested_name": "笔记本电脑",
                "needed_at": date.today(),
                "reason": "办公使用",
            },
            format="json",
        )
        self.assertEqual(response.status_code, 400)


class OIDCAndManagerSettingsTests(TestCase):
    @override_settings(DJANGO_SUPERUSER_USERNAME="owner")
    def test_configured_oidc_owner_is_promoted_and_local_admin_is_hidden(self):
        owner = sync_oidc_user(
            {
                "preferred_username": "owner",
                "email": "owner@example.com",
                "name": "系统负责人",
            }
        )
        User.objects.create_superuser("admin", password="pass")
        User.objects.create_user("zhangsan", password="pass", first_name="张三")
        self.assertTrue(owner.is_superuser)
        self.assertTrue(owner.is_staff)
        self.assertEqual(owner.employee_profile.employee_no, "owner")

        client = APIClient()
        client.force_authenticate(owner)
        response = client.get("/api/v1/settings/managers/")
        self.assertEqual(response.status_code, 200)
        usernames = [row["username"] for row in response.data["users"]]
        self.assertIn("owner", usernames)
        self.assertNotIn("admin", usernames)

    def test_oidc_email_is_metadata_and_cannot_be_used_as_username(self):
        with self.assertRaisesMessage(ValueError, "OIDC 账号未提供有效的登录名"):
            sync_oidc_user({"email": "zhangsan@example.com"})

    def test_local_admin_login_uses_database_credentials(self):
        User.objects.create_superuser("admin", password="test-password")
        client = APIClient()
        response = client.post(
            "/api/v1/auth/local/login/",
            {"username": "admin", "password": "test-password"},
            format="json",
        )
        self.assertEqual(response.status_code, 200)
        self.assertIn("token", response.data)


@override_settings(
    EMAIL_NOTIFICATIONS_ENABLED=True,
    EMAIL_HOST="smtp.test",
    EMAIL_BACKEND="django.core.mail.backends.locmem.EmailBackend",
    DEFAULT_FROM_EMAIL="affairs-os@example.com",
    EMAIL_SUBJECT_PREFIX="[行政资产管理] ",
    FRONTEND_URL="https://assets.example.com",
)
class EmailNotificationTests(TestCase):
    def setUp(self):
        self.requester = User.objects.create_user(
            "email_requester",
            password="pass",
            first_name="申请人",
            email="requester@example.com",
        )
        self.manager = User.objects.create_user(
            "inventory_email_manager",
            password="pass",
            first_name="库存管理员",
            email="manager@example.com",
        )
        AssetManagerRole.objects.create(user=self.manager, scopes=["inventory"])
        self.category = AssetCategory.objects.create(name="笔记本电脑", code="NB")
        self.asset = Asset.objects.create(
            asset_tag="IT-NB-2026-301",
            category=self.category,
            brand="Lenovo",
            model_name="ThinkBook",
            status=Asset.Status.AVAILABLE,
        )

    def test_request_emails_only_go_to_managers(self):
        requester_client = APIClient()
        requester_client.force_authenticate(self.requester)
        created = requester_client.post(
            "/api/v1/requests/",
            {
                "request_type": "loan",
                "requested_name": "笔记本电脑",
                "needed_at": date.today(),
                "expected_return_at": date.today() + timedelta(days=5),
                "reason": "出差",
            },
            format="json",
        )
        self.assertEqual(created.status_code, 201)
        self.assertEqual(
            set(EmailNotification.objects.values_list("event_type", flat=True)),
            {"request_pending"},
        )
        self.assertEqual(
            set(EmailNotification.objects.values_list("recipient_email", flat=True)),
            {"manager@example.com"},
        )

        manager_client = APIClient()
        manager_client.force_authenticate(self.manager)
        fulfilled = manager_client.post(
            f"/api/v1/requests/{created.data['id']}/fulfill/",
            {"asset_id": self.asset.id},
            format="json",
        )
        self.assertEqual(fulfilled.status_code, 200)
        self.assertFalse(EmailNotification.objects.filter(recipient_email="requester@example.com").exists())

    def test_email_task_marks_delivery_sent(self):
        notification = EmailNotification.objects.create(
            event_key="test-delivery",
            event_type="test",
            recipient_user=self.requester,
            recipient_email=self.requester.email,
            subject="[行政资产管理] 测试通知",
            body="这是一封测试邮件。",
        )
        result = send_email_notification(notification.pk)
        self.assertEqual(result, "sent")
        notification.refresh_from_db()
        self.assertEqual(notification.status, EmailNotification.Status.SENT)
        self.assertEqual(len(mail.outbox), 1)
        self.assertEqual(mail.outbox[0].to, ["requester@example.com"])

    def test_direct_asset_and_inventory_actions_do_not_email_user(self):
        perform_asset_action(
            asset=self.asset,
            action="assign",
            actor=self.manager,
            target_user=self.requester,
        )
        item = InventoryItem.objects.create(
            sku="NO-MAIL-001",
            name="鼠标垫",
            kind=InventoryItem.Kind.CONSUMABLE,
            quantity=5,
        )
        client = APIClient()
        client.force_authenticate(self.manager)
        issued = client.post(
            f"/api/v1/inventory/{item.id}/transactions/",
            {"action": "issue", "quantity": 1, "recipient_id": self.requester.id},
            format="json",
        )
        self.assertEqual(issued.status_code, 200)
        self.assertFalse(EmailNotification.objects.exists())

    def test_daily_task_queues_overdue_but_not_low_stock_reminders(self):
        self.asset.status = Asset.Status.LOANED
        self.asset.assigned_to = self.requester
        self.asset.expected_return_at = date.today() - timedelta(days=2)
        self.asset.save()
        InventoryItem.objects.create(
            sku="LOW-MAIL-001",
            name="无线鼠标",
            quantity=1,
            minimum_quantity=5,
        )

        first = send_daily_operational_notifications()
        second = send_daily_operational_notifications()
        self.assertEqual(
            first,
            {"overdue": 1, "low_stock": 0, "vehicle_due": 0, "contract_due": 0, "user_inactive": 0},
        )
        self.assertEqual(second, first)
        self.assertEqual(
            EmailNotification.objects.filter(event_type="loan_overdue").count(),
            1,
        )
        self.assertEqual(
            EmailNotification.objects.filter(event_type="loan_overdue_summary").count(),
            1,
        )
        self.assertEqual(
            EmailNotification.objects.filter(event_type="low_stock_summary").count(),
            0,
        )
        user_notifications = EmailNotification.objects.filter(recipient_email="requester@example.com")
        self.assertEqual(list(user_notifications.values_list("event_type", flat=True)), ["loan_overdue"])
        self.assertTrue(user_notifications.get().subject.startswith("[行政资产管理] "))

    def test_daily_task_queues_due_today_reminder(self):
        self.asset.status = Asset.Status.LOANED
        self.asset.assigned_to = self.requester
        self.asset.expected_return_at = timezone.localdate()
        self.asset.save()

        result = send_daily_operational_notifications()
        self.assertEqual(result["overdue"], 0)
        requester_notification = EmailNotification.objects.get(
            event_type="loan_due_today",
            recipient_email="requester@example.com",
        )
        manager_notification = EmailNotification.objects.get(
            event_type="loan_due_today_summary",
            recipient_email="manager@example.com",
        )
        self.assertIn("今天到期", requester_notification.subject)
        self.assertIn("请按时归还或联系管理员办理续借", requester_notification.body)
        self.assertIn("今天共 1 件到期", manager_notification.subject)
        self.assertFalse(EmailNotification.objects.filter(event_type="loan_overdue").exists())

    def test_daily_task_distinguishes_due_today_and_overdue(self):
        self.asset.status = Asset.Status.LOANED
        self.asset.assigned_to = self.requester
        self.asset.expected_return_at = timezone.localdate()
        self.asset.save()
        overdue = Asset.objects.create(
            asset_tag="IT-NB-2026-302",
            category=self.category,
            brand="Lenovo",
            model_name="ThinkBook",
            status=Asset.Status.LOANED,
            assigned_to=self.requester,
            expected_return_at=timezone.localdate() - timedelta(days=1),
        )

        send_daily_operational_notifications()
        event_types = set(
            EmailNotification.objects.filter(recipient_email="requester@example.com")
            .values_list("event_type", flat=True)
        )
        self.assertEqual(event_types, {"loan_due_today", "loan_overdue"})
        due_subject = EmailNotification.objects.get(
            event_type="loan_due_today",
            recipient_email="requester@example.com",
        ).subject
        overdue_subject = EmailNotification.objects.get(
            event_type="loan_overdue",
            recipient_email="requester@example.com",
        ).subject
        self.assertIn("今天到期", due_subject)
        self.assertIn("已经超期", overdue_subject)

    def test_contract_due_email_only_goes_to_owner_and_superuser(self):
        owner = User.objects.create_user(
            "contract-email-owner",
            password="pass",
            email="contract-owner@example.com",
        )
        other_manager = User.objects.create_user(
            "other-contract-email-manager",
            password="pass",
            email="other-contract-manager@example.com",
        )
        superuser = User.objects.create_superuser(
            "contract-email-superuser",
            password="pass",
            email="contract-superuser@example.com",
        )
        AssetManagerRole.objects.create(user=owner, scopes=["contracts"])
        AssetManagerRole.objects.create(user=other_manager, scopes=["contracts"])
        Contract.objects.create(
            contract_no="HT-MAIL-OWNER-001",
            name="负责人专属合同",
            owner=owner,
            status=Contract.Status.ACTIVE,
            end_date=timezone.localdate() + timedelta(days=10),
        )

        send_daily_operational_notifications()

        recipients = set(
            EmailNotification.objects.filter(event_type="contract_expiry")
            .values_list("recipient_email", flat=True)
        )
        self.assertEqual(
            recipients,
            {"contract-owner@example.com", "contract-superuser@example.com"},
        )

    def test_user_deactivation_sends_manager_email_once(self):
        self.asset.status = Asset.Status.ASSIGNED
        self.asset.assigned_to = self.requester
        self.asset.save()

        self.requester.is_active = False
        self.requester.save()

        notification = EmailNotification.objects.get(event_type="user_deactivated")
        self.assertEqual(notification.recipient_email, "manager@example.com")
        self.assertIn("员工离职提醒：申请人", notification.subject)
        self.assertIn("IT-NB-2026-301", notification.body)
        self.assertIn("名下合同 0 份", notification.body)
        self.assertIn("请及时办理交接", notification.body)

        self.requester.save()
        self.assertEqual(EmailNotification.objects.filter(event_type="user_deactivated").count(), 1)

    def test_user_reenable_does_not_send_deactivation_email(self):
        self.requester.is_active = True
        self.requester.save()
        self.assertFalse(EmailNotification.objects.filter(event_type="user_deactivated").exists())

    def test_creating_user_does_not_send_deactivation_email(self):
        User.objects.create_user("brand_new_user", password="pass", email="new@example.com")
        self.assertFalse(EmailNotification.objects.filter(event_type="user_deactivated").exists())

    def test_daily_task_repeats_inactive_user_reminder_until_resolved(self):
        self.asset.status = Asset.Status.ASSIGNED
        self.asset.assigned_to = self.requester
        self.asset.save()
        self.requester.is_active = False
        self.requester.save()

        with patch(
            "assets.tasks.timezone.localdate",
            side_effect=[date(2026, 8, 7), date(2026, 8, 8)],
        ):
            first_day = send_daily_operational_notifications()
            second_day = send_daily_operational_notifications()
        self.assertEqual(first_day["user_inactive"], 1)
        self.assertEqual(second_day["user_inactive"], 1)
        notifications = EmailNotification.objects.filter(
            event_type="user_deactivated",
            recipient_email="manager@example.com",
        ).order_by("created_at")
        # 停用时即时提醒 1 封 + 每天 1 封（两天）
        self.assertEqual(notifications.count(), 3)
        self.assertIn("离职交接提醒：申请人（未完成）", notifications.last().subject)

        # 名下资产处理后，每日任务不再提醒
        self.asset.status = Asset.Status.AVAILABLE
        self.asset.assigned_to = None
        self.asset.save()
        with patch("assets.tasks.timezone.localdate", return_value=date(2026, 8, 9)):
            resolved = send_daily_operational_notifications()
        self.assertEqual(resolved["user_inactive"], 0)
        self.assertEqual(notifications.count(), 3)

    def test_daily_task_skips_inactive_user_without_open_items(self):
        self.requester.is_active = False
        self.requester.save()
        with patch("assets.tasks.timezone.localdate", return_value=date(2026, 8, 7)):
            result = send_daily_operational_notifications()
        self.assertEqual(result["user_inactive"], 0)
        # 只有停用时的即时提醒，没有每日提醒
        self.assertEqual(
            EmailNotification.objects.filter(event_type="user_deactivated").count(),
            1,
        )

    def test_loan_extend_sends_email_to_borrower(self):
        self.asset.status = Asset.Status.LOANED
        self.asset.assigned_to = self.requester
        self.asset.expected_return_at = date.today() + timedelta(days=7)
        self.asset.save()

        result = perform_asset_action(
            asset=self.asset,
            action="extend",
            actor=self.manager,
            expected_return_at=date.today() + timedelta(days=21),
        )
        self.assertEqual(result.event.action, AssetEvent.Action.EXTENDED)
        notification = EmailNotification.objects.get(
            event_type="loan_extended",
            recipient_email="requester@example.com",
        )
        self.assertIn("借用延期提醒", notification.subject)
        self.assertIn(str(date.today() + timedelta(days=21)), notification.body)
        self.assertIn(str(date.today() + timedelta(days=7)), notification.body)


@override_settings(EMAIL_NOTIFICATIONS_ENABLED=False)
class AdministrativePhaseTests(TestCase):
    def setUp(self):
        self.admin = User.objects.create_user("phase-admin", password="pass", is_superuser=True, is_staff=True)
        self.employee = User.objects.create_user("phase-employee", password="pass")
        self.department = Department.objects.create(name="行政部", code="ADM")
        EmployeeProfile.objects.create(user=self.employee, employee_no="E-PHASE", department=self.department)
        self.category = ExpenseCategory.objects.get(code="PURCHASE")
        self.client = APIClient()

    def test_vehicle_dispatch_full_workflow(self):
        vehicle = Vehicle.objects.create(plate_number="粤BTEST01", name="测试车辆", seats=5, department=self.department)
        self.client.force_authenticate(self.employee)
        created = self.client.post(
            "/api/v1/vehicle-dispatches/",
            {
                "purpose": "客户接待", "destination": "机场", "passenger_count": 3,
                "planned_departure_at": (timezone.now() + timedelta(days=1)).isoformat(),
                "planned_return_at": (timezone.now() + timedelta(days=1, hours=3)).isoformat(),
            }, format="json",
        )
        self.assertEqual(created.status_code, 201)
        self.assertTrue(created.data["request_no"].startswith("PC-"))
        self.client.force_authenticate(self.admin)
        dispatched = self.client.post(
            f"/api/v1/vehicle-dispatches/{created.data['id']}/dispatch/",
            {"vehicle_id": vehicle.id, "driver_name": "测试司机"}, format="json",
        )
        self.assertEqual(dispatched.status_code, 200)
        departed = self.client.post(f"/api/v1/vehicle-dispatches/{created.data['id']}/depart/", {"mileage": 1000}, format="json")
        self.assertEqual(departed.data["status"], VehicleDispatch.Status.IN_PROGRESS)
        completed = self.client.post(f"/api/v1/vehicle-dispatches/{created.data['id']}/complete/", {"mileage": 1080}, format="json")
        self.assertEqual(completed.data["status"], VehicleDispatch.Status.COMPLETED)
        vehicle.refresh_from_db()
        self.assertEqual(vehicle.current_mileage, 1080)
        self.assertEqual(vehicle.status, Vehicle.Status.AVAILABLE)

    @patch("assets.views.nextcloud_storage.delete")
    @patch("assets.views.nextcloud_storage.upload")
    def test_contract_source_file_upload_and_delete(self, storage_upload, storage_delete):
        self.client.force_authenticate(self.admin)
        contract = Contract.objects.create(
            contract_no="HT-2026-001",
            name="办公服务合同",
            category=self.category,
            department=self.department,
            owner=self.employee,
            amount="1000.00",
        )
        upload = SimpleUploadedFile(
            "合同源文件.docx",
            b"contract-source",
            content_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        )
        created = self.client.post(
            f"/api/v1/contracts/{contract.pk}/files/",
            {"file": upload, "document_type": "invoice"},
            format="multipart",
        )
        self.assertEqual(created.status_code, 201)
        self.assertEqual(created.data["document_type_label"], "发票")
        storage_upload.assert_called_once()
        attachment = ContractAttachment.objects.get(pk=created.data["id"])
        self.assertTrue(attachment.remote_path.startswith("/AffairsOS/contracts/"))

        listed = self.client.get("/api/v1/contracts/")
        self.assertEqual(listed.data[0]["attachments"][0]["id"], attachment.id)

        deleted = self.client.delete(
            f"/api/v1/contracts/{contract.pk}/files/{attachment.pk}/"
        )
        self.assertEqual(deleted.status_code, 204)
        storage_delete.assert_called_once_with(attachment.remote_path)
        self.assertFalse(ContractAttachment.objects.filter(pk=attachment.pk).exists())

    def test_contract_dates_and_amount_can_be_edited_directly_or_with_change_history(self):
        self.client.force_authenticate(self.admin)
        contract = Contract.objects.create(
            contract_no="HT-CHANGE-001", name="年度服务合同", category=self.category,
            start_date=date(2026, 1, 1), end_date=date(2026, 12, 31), amount="12000.00",
            status=Contract.Status.ACTIVE,
        )
        edited = self.client.patch(
            f"/api/v1/contracts/{contract.pk}/",
            {"start_date": "2026-02-01", "end_date": "2027-01-31", "amount": "15000.00"},
            format="json",
        )
        self.assertEqual(edited.status_code, 200)
        contract.refresh_from_db()
        self.assertEqual(contract.start_date, date(2026, 2, 1))
        self.assertEqual(contract.end_date, date(2027, 1, 31))
        self.assertEqual(contract.amount, Decimal("15000.00"))
        changed = self.client.post(
            f"/api/v1/contracts/{contract.pk}/changes/",
            {
                "change_type": "extension", "changed_on": "2026-11-01",
                "new_end_date": "2027-12-31", "notes": "双方签署延期协议",
            }, format="json",
        )
        self.assertEqual(changed.status_code, 201)
        contract.refresh_from_db()
        self.assertEqual(contract.end_date, date(2027, 12, 31))
        history = ContractChange.objects.get(pk=changed.data["id"])
        self.assertEqual(history.old_end_date, date(2027, 1, 31))
        self.assertEqual(history.new_end_date, date(2027, 12, 31))

    def test_contract_direct_edit_rejects_reversed_dates(self):
        self.client.force_authenticate(self.admin)
        contract = Contract.objects.create(
            contract_no="HT-DATE-001", name="日期校验合同",
            start_date=date(2026, 1, 1), end_date=date(2026, 12, 31), amount="1000.00",
        )
        invalid = self.client.patch(
            f"/api/v1/contracts/{contract.pk}/",
            {"start_date": "2027-01-01", "end_date": "2026-12-31"},
            format="json",
        )
        self.assertEqual(invalid.status_code, 400)
        self.assertIn("结束日期不能早于开始日期。", invalid.data["errors"]["end_date"][0])

    def test_contract_renewal_creates_linked_record_and_completes_previous(self):
        self.client.force_authenticate(self.admin)
        contract_type = ContractType.objects.create(code="TEST-SERVICE", name="测试服务合同")
        previous = Contract.objects.create(
            contract_no="HT-2026-RENEW", name="办公服务合同", contract_type=contract_type,
            start_date=date(2026, 1, 1), end_date=date(2026, 12, 31), amount="36000.00",
            status=Contract.Status.ACTIVE,
        )
        renewed = self.client.post(
            f"/api/v1/contracts/{previous.pk}/renew/",
            {
                "contract_no": "HT-2027-RENEW", "name": "办公服务合同",
                "contract_type": contract_type.id, "start_date": "2027-01-01",
                "end_date": "2027-12-31", "amount": "38000.00", "status": "active",
            }, format="json",
        )
        self.assertEqual(renewed.status_code, 201)
        previous.refresh_from_db()
        successor = Contract.objects.get(pk=renewed.data["id"])
        self.assertEqual(previous.status, Contract.Status.COMPLETED)
        self.assertEqual(successor.previous_contract, previous)
        third = self.client.post(
            f"/api/v1/contracts/{successor.pk}/renew/",
            {
                "contract_no": "HT-2028-RENEW", "name": "办公服务合同",
                "contract_type": contract_type.id, "start_date": "2028-01-01",
                "end_date": "2028-12-31", "amount": "40000.00", "status": "active",
            }, format="json",
        )
        self.assertEqual(third.status_code, 201)
        latest = Contract.objects.get(pk=third.data["id"])
        listed = self.client.get("/api/v1/contracts/?q=办公服务合同")
        self.assertEqual([row["contract_no"] for row in listed.data], [latest.contract_no])
        history = self.client.get(f"/api/v1/contracts/{latest.pk}/history/")
        self.assertEqual(
            [row["contract_no"] for row in history.data],
            [previous.contract_no, successor.contract_no, latest.contract_no],
        )

    def test_contract_change_error_keeps_contract_untouched(self):
        self.client.force_authenticate(self.admin)
        contract = Contract.objects.create(
            contract_no="HT-ERR-001", name="错误场景合同", category=self.category,
            start_date=date(2026, 1, 1), end_date=date(2026, 12, 31), amount="12000.00",
            status=Contract.Status.ACTIVE,
        )
        failed = self.client.post(
            f"/api/v1/contracts/{contract.pk}/changes/",
            {
                "change_type": "extension", "changed_on": "2026-11-01",
                "new_end_date": "2025-12-31", "notes": "结束日期早于开始日期",
            }, format="json",
        )
        self.assertEqual(failed.status_code, 400)
        self.assertIn(
            "变更后的结束日期不能早于开始日期。",
            failed.data["errors"]["new_end_date"][0],
        )
        contract.refresh_from_db()
        self.assertEqual(contract.end_date, date(2026, 12, 31))
        self.assertFalse(ContractChange.objects.filter(contract=contract).exists())

    def test_supplement_change_creates_supplement_contract_and_totals_amount(self):
        self.client.force_authenticate(self.admin)
        parent = Contract.objects.create(
            contract_no="HT-SUP-001", name="年度保洁服务合同", category=self.category,
            department=self.department, owner=self.employee,
            start_date=date(2026, 1, 1), end_date=date(2026, 12, 31), amount="12000.00",
            status=Contract.Status.ACTIVE,
        )
        first = self.client.post(
            f"/api/v1/contracts/{parent.pk}/changes/",
            {
                "change_type": "supplement", "changed_on": "2026-03-01",
                "new_start_date": "2026-03-01", "new_end_date": "2026-06-30",
                "new_amount": "5000.00", "notes": "增加绿化养护服务",
            }, format="json",
        )
        self.assertEqual(first.status_code, 201)
        parent.refresh_from_db()
        self.assertEqual(parent.amount, Decimal("12000.00"))
        self.assertEqual(parent.end_date, date(2026, 12, 31))
        supplement = Contract.objects.get(supplement_of=parent)
        self.assertEqual(supplement.contract_no, "HT-SUP-001-S01")
        self.assertEqual(supplement.name, "年度保洁服务合同（补充协议 1）")
        self.assertEqual(supplement.amount, Decimal("5000.00"))
        self.assertEqual(supplement.start_date, date(2026, 3, 1))
        self.assertEqual(supplement.end_date, date(2026, 6, 30))
        self.assertEqual(supplement.status, Contract.Status.ACTIVE)
        change = ContractChange.objects.get(
            contract=parent,
            change_type=ContractChange.ChangeType.SUPPLEMENT,
        )
        self.assertEqual(change.new_amount, Decimal("5000.00"))
        self.assertEqual(change.old_amount, Decimal("12000.00"))

        second = self.client.post(
            f"/api/v1/contracts/{parent.pk}/changes/",
            {
                "change_type": "supplement", "changed_on": "2026-09-01",
                "new_start_date": "2026-09-01", "new_end_date": "2026-12-31",
                "new_amount": "3000.00", "notes": "增加门禁维护服务",
            }, format="json",
        )
        self.assertEqual(second.status_code, 201)
        second_supplement = Contract.objects.get(contract_no="HT-SUP-001-S02")

        listed = self.client.get("/api/v1/contracts/")
        listed_ids = [row["id"] for row in listed.data]
        self.assertIn(parent.id, listed_ids)
        self.assertNotIn(second_supplement.id, listed_ids)
        parent_row = next(row for row in listed.data if row["id"] == parent.id)
        self.assertEqual(parent_row["total_amount"], "20000.00")
        self.assertEqual(len(parent_row["supplement_contracts"]), 2)

        history = self.client.get(f"/api/v1/contracts/{parent.pk}/history/")
        current = history.data[-1]
        self.assertEqual(current["id"], parent.id)
        self.assertEqual(len(current["supplement_contracts"]), 2)
        self.assertEqual(current["total_amount"], "20000.00")

    def test_supplement_change_validation(self):
        self.client.force_authenticate(self.admin)
        parent = Contract.objects.create(
            contract_no="HT-SUP-VALID", name="验证合同", category=self.category,
            start_date=date(2026, 1, 1), end_date=date(2026, 12, 31), amount="10000.00",
            status=Contract.Status.ACTIVE,
        )
        no_amount = self.client.post(
            f"/api/v1/contracts/{parent.pk}/changes/",
            {
                "change_type": "supplement", "changed_on": "2026-03-01",
                "new_start_date": "2026-03-01", "new_end_date": "2026-06-30",
                "notes": "缺少金额",
            }, format="json",
        )
        self.assertEqual(no_amount.status_code, 400)
        self.assertIn("补充协议必须填写补充金额。", no_amount.data["errors"]["new_amount"][0])

        bad_dates = self.client.post(
            f"/api/v1/contracts/{parent.pk}/changes/",
            {
                "change_type": "supplement", "changed_on": "2026-03-01",
                "new_start_date": "2026-06-01", "new_end_date": "2026-03-01",
                "new_amount": "5000.00", "notes": "日期颠倒",
            }, format="json",
        )
        self.assertEqual(bad_dates.status_code, 400)
        self.assertIn(
            "补充协议结束日期不能早于开始日期。",
            bad_dates.data["errors"]["new_end_date"][0],
        )

    def test_supplement_change_rejected_on_supplement_contract(self):
        self.client.force_authenticate(self.admin)
        parent = Contract.objects.create(
            contract_no="HT-SUP-NEST", name="嵌套合同", category=self.category,
            start_date=date(2026, 1, 1), end_date=date(2026, 12, 31), amount="10000.00",
            status=Contract.Status.ACTIVE,
        )
        supplement = Contract.objects.create(
            contract_no="HT-SUP-NEST-S01", name="嵌套合同（补充协议 1）",
            category=self.category, amount="2000.00", supplement_of=parent,
            start_date=date(2026, 3, 1), end_date=date(2026, 6, 30),
            status=Contract.Status.ACTIVE,
        )
        failed = self.client.post(
            f"/api/v1/contracts/{supplement.pk}/changes/",
            {
                "change_type": "supplement", "changed_on": "2026-04-01",
                "new_start_date": "2026-04-01", "new_end_date": "2026-05-31",
                "new_amount": "1000.00", "notes": "不应允许",
            }, format="json",
        )
        self.assertEqual(failed.status_code, 400)
        self.assertIn("补充协议请登记在母合同上", failed.data["errors"]["change_type"][0])

    def test_superuser_delete_contract_includes_supplements(self):
        self.client.force_authenticate(self.admin)
        parent = Contract.objects.create(
            contract_no="HT-SUP-DEL", name="删除保护合同", category=self.category,
            amount="10000.00", status=Contract.Status.ACTIVE,
        )
        supplement = Contract.objects.create(
            contract_no="HT-SUP-DEL-S01", name="删除保护合同（补充协议 1）",
            category=self.category, amount="2000.00", supplement_of=parent,
            status=Contract.Status.ACTIVE,
        )
        deleted = self.client.delete(f"/api/v1/contracts/{parent.pk}/")
        self.assertEqual(deleted.status_code, 204)
        self.assertFalse(Contract.objects.filter(pk__in=[parent.pk, supplement.pk]).exists())

    def test_non_superuser_cannot_delete_contract(self):
        AssetManagerRole.objects.create(user=self.employee, scopes=["contracts"])
        self.client.force_authenticate(self.employee)
        contract = Contract.objects.create(
            contract_no="HT-NODEL-001",
            name="保留合同",
            category=self.category,
            owner=self.employee,
            amount="1000.00",
            status=Contract.Status.ACTIVE,
        )
        deleted = self.client.delete(f"/api/v1/contracts/{contract.pk}/")
        self.assertEqual(deleted.status_code, 405)
        self.assertTrue(Contract.objects.filter(pk=contract.pk).exists())

    def test_contract_manager_only_sees_and_operates_own_contracts(self):
        other_manager = User.objects.create_user("other-contract-manager", password="pass")
        AssetManagerRole.objects.create(user=self.employee, scopes=["contracts"])
        AssetManagerRole.objects.create(user=other_manager, scopes=["contracts"])
        own = Contract.objects.create(
            contract_no="HT-OWN-001",
            name="本人合同",
            owner=self.employee,
            amount="1000.00",
            status=Contract.Status.ACTIVE,
            end_date=timezone.localdate() + timedelta(days=10),
        )
        other = Contract.objects.create(
            contract_no="HT-OTHER-001",
            name="他人合同",
            owner=other_manager,
            amount="2000.00",
            status=Contract.Status.ACTIVE,
            end_date=timezone.localdate() + timedelta(days=10),
        )
        Contract.objects.create(
            contract_no="HT-NO-OWNER-001",
            name="未分配合同",
            amount="3000.00",
        )
        self.client.force_authenticate(self.employee)

        listed = self.client.get("/api/v1/contracts/")
        own_detail = self.client.get(f"/api/v1/contracts/{own.pk}/")
        other_detail = self.client.get(f"/api/v1/contracts/{other.pk}/")
        other_history = self.client.get(f"/api/v1/contracts/{other.pk}/history/")
        dashboard = self.client.get("/api/v1/dashboard/")

        self.assertEqual([row["id"] for row in listed.data], [own.id])
        self.assertEqual(own_detail.status_code, 200)
        self.assertEqual(other_detail.status_code, 404)
        self.assertEqual(other_history.status_code, 404)
        self.assertEqual(dashboard.data["admin_tasks"]["expiring_contracts"], 1)

    def test_contract_manager_is_forced_as_owner_on_create_update_and_renew(self):
        other_manager = User.objects.create_user("other-contract-owner", password="pass")
        AssetManagerRole.objects.create(user=self.employee, scopes=["contracts"])
        self.client.force_authenticate(self.employee)

        created = self.client.post(
            "/api/v1/contracts/",
            {
                "contract_no": "HT-FORCED-OWNER-001",
                "name": "自动负责人合同",
                "owner": other_manager.id,
                "amount": "1000.00",
                "status": Contract.Status.ACTIVE,
            },
            format="json",
        )
        self.assertEqual(created.status_code, 201)
        contract = Contract.objects.get(pk=created.data["id"])
        self.assertEqual(contract.owner, self.employee)

        updated = self.client.patch(
            f"/api/v1/contracts/{contract.pk}/",
            {"owner": other_manager.id, "notes": "尝试改给他人"},
            format="json",
        )
        self.assertEqual(updated.status_code, 200)
        contract.refresh_from_db()
        self.assertEqual(contract.owner, self.employee)

        renewed = self.client.post(
            f"/api/v1/contracts/{contract.pk}/renew/",
            {
                "contract_no": "HT-FORCED-OWNER-002",
                "name": "续签合同",
                "owner": other_manager.id,
                "amount": "1200.00",
                "status": Contract.Status.ACTIVE,
            },
            format="json",
        )
        self.assertEqual(renewed.status_code, 201)
        self.assertEqual(Contract.objects.get(pk=renewed.data["id"]).owner, self.employee)

    def test_superuser_can_see_all_contracts_and_assign_owner(self):
        other_manager = User.objects.create_user("visible-contract-owner", password="pass")
        first = Contract.objects.create(
            contract_no="HT-ALL-001",
            name="第一份合同",
            owner=self.employee,
        )
        second = Contract.objects.create(
            contract_no="HT-ALL-002",
            name="第二份合同",
            owner=other_manager,
        )
        self.client.force_authenticate(self.admin)

        listed = self.client.get("/api/v1/contracts/")
        reassigned = self.client.patch(
            f"/api/v1/contracts/{first.pk}/",
            {"owner": other_manager.id},
            format="json",
        )

        self.assertEqual({row["id"] for row in listed.data}, {first.id, second.id})
        self.assertEqual(reassigned.status_code, 200)
        first.refresh_from_db()
        self.assertEqual(first.owner, other_manager)

    def test_superuser_can_delete_asset_with_history(self):
        self.client.force_authenticate(self.admin)
        category = AssetCategory.objects.create(name="删除测试电脑", code="DEL-NB")
        asset = Asset.objects.create(
            asset_tag="IT-DEL-001",
            category=category,
            name="待删除资产",
            status=Asset.Status.AVAILABLE,
        )
        AssetEvent.objects.create(
            asset=asset,
            action=AssetEvent.Action.CREATED,
            from_status="",
            to_status=Asset.Status.AVAILABLE,
        )
        deleted = self.client.delete(f"/api/v1/assets/{asset.pk}/")
        self.assertEqual(deleted.status_code, 204)
        self.assertFalse(Asset.objects.filter(pk=asset.pk).exists())
        self.assertFalse(AssetEvent.objects.filter(asset_id=asset.pk).exists())

    def test_non_superuser_cannot_delete_asset(self):
        AssetManagerRole.objects.create(user=self.employee, scopes=["assets"])
        self.client.force_authenticate(self.employee)
        category = AssetCategory.objects.create(name="保留测试电脑", code="NODEL-NB")
        asset = Asset.objects.create(
            asset_tag="IT-NODEL-001",
            category=category,
            name="保留资产",
            status=Asset.Status.AVAILABLE,
        )
        deleted = self.client.delete(f"/api/v1/assets/{asset.pk}/")
        self.assertEqual(deleted.status_code, 405)
        self.assertTrue(Asset.objects.filter(pk=asset.pk).exists())

    def test_superuser_delete_asset_blocked_when_referenced(self):
        self.client.force_authenticate(self.admin)
        category = AssetCategory.objects.create(name="引用测试电脑", code="REF-NB")
        asset = Asset.objects.create(
            asset_tag="IT-REF-001",
            category=category,
            name="被引用资产",
            status=Asset.Status.AVAILABLE,
        )
        AssetRequest.objects.create(
            requester=self.employee,
            request_type=AssetRequest.RequestType.ASSIGN,
            requested_item_type=AssetRequest.ItemType.ASSET,
            requested_name="笔记本",
            reason="测试",
            status=AssetRequest.Status.FULFILLED,
            assigned_asset=asset,
        )
        deleted = self.client.delete(f"/api/v1/assets/{asset.pk}/")
        self.assertEqual(deleted.status_code, 400)
        self.assertIn("无法删除", deleted.data["message"])
        self.assertTrue(Asset.objects.filter(pk=asset.pk).exists())

    def test_superuser_can_delete_unused_basic_data(self):
        self.client.force_authenticate(self.admin)
        location = Location.objects.create(name="临时库房", code="TMP-LOC")
        deleted = self.client.delete(f"/api/v1/locations/{location.pk}/")
        self.assertEqual(deleted.status_code, 204)
        self.assertFalse(Location.objects.filter(pk=location.pk).exists())

    def test_non_superuser_cannot_delete_basic_data(self):
        AssetManagerRole.objects.create(user=self.employee, scopes=["settings"])
        self.client.force_authenticate(self.employee)
        location = Location.objects.create(name="临时库房二", code="TMP-LOC2")
        deleted = self.client.delete(f"/api/v1/locations/{location.pk}/")
        self.assertEqual(deleted.status_code, 405)
        self.assertTrue(Location.objects.filter(pk=location.pk).exists())

    def test_superuser_delete_basic_data_blocked_when_referenced(self):
        self.client.force_authenticate(self.admin)
        category = ExpenseCategory.objects.create(
            code="TMP-EXP",
            name="临时费用类别",
        )
        Contract.objects.create(
            contract_no="HT-REF-001",
            name="引用合同",
            category=category,
        )
        deleted = self.client.delete(f"/api/v1/expense-categories/{category.pk}/")
        self.assertEqual(deleted.status_code, 400)
        self.assertTrue(ExpenseCategory.objects.filter(pk=category.pk).exists())

    def test_superuser_can_delete_inventory_item_without_transactions(self):
        self.client.force_authenticate(self.admin)
        item = InventoryItem.objects.create(
            sku="DEL-001",
            name="待删物品",
            kind=InventoryItem.Kind.CONSUMABLE,
            quantity=3,
        )
        deleted = self.client.delete(f"/api/v1/inventory/{item.pk}/")
        self.assertEqual(deleted.status_code, 204)
        self.assertFalse(InventoryItem.objects.filter(pk=item.pk).exists())

    def test_non_superuser_cannot_delete_inventory_item(self):
        AssetManagerRole.objects.create(user=self.employee, scopes=["inventory"])
        self.client.force_authenticate(self.employee)
        item = InventoryItem.objects.create(
            sku="NODEL-001",
            name="保留物品",
            kind=InventoryItem.Kind.CONSUMABLE,
            quantity=3,
        )
        deleted = self.client.delete(f"/api/v1/inventory/{item.pk}/")
        self.assertEqual(deleted.status_code, 405)
        self.assertTrue(InventoryItem.objects.filter(pk=item.pk).exists())

    def test_module_switch_disables_module_for_managers(self):
        self.client.force_authenticate(self.admin)
        modules = self.client.get("/api/v1/settings/modules/")
        self.assertEqual(modules.status_code, 200)
        codes = {item["code"] for item in modules.data}
        self.assertIn("assets", codes)
        self.assertNotIn("settings", codes)

        toggled = self.client.patch(
            "/api/v1/settings/modules/",
            {"code": "assets", "enabled": False},
            format="json",
        )
        self.assertEqual(toggled.status_code, 200)
        self.assertFalse(toggled.data["enabled"])

        AssetManagerRole.objects.create(user=self.employee, scopes=["assets"])
        self.client.force_authenticate(self.employee)
        blocked = self.client.get("/api/v1/assets/")
        self.assertEqual(blocked.status_code, 403)

        self.client.force_authenticate(self.admin)
        lookups = self.client.get("/api/v1/lookups/")
        self.assertNotIn("assets", lookups.data["enabled_modules"])
        manager_settings = self.client.get("/api/v1/settings/managers/")
        self.assertNotIn("assets", {item["value"] for item in manager_settings.data["modules"]})

        self.client.patch(
            "/api/v1/settings/modules/",
            {"code": "assets", "enabled": True},
            format="json",
        )

    def test_module_switch_requires_superuser(self):
        self.client.force_authenticate(self.employee)
        blocked = self.client.get("/api/v1/settings/modules/")
        self.assertEqual(blocked.status_code, 403)

    def test_settings_module_cannot_be_toggled(self):
        self.client.force_authenticate(self.admin)
        blocked = self.client.patch(
            "/api/v1/settings/modules/",
            {"code": "settings", "enabled": False},
            format="json",
        )
        self.assertEqual(blocked.status_code, 400)

    def test_superuser_can_delete_vehicle(self):
        self.client.force_authenticate(self.admin)
        vehicle = Vehicle.objects.create(
            plate_number="粤BTEST01",
            name="测试车辆",
            seats=5,
            department=self.department,
        )
        deleted = self.client.delete(f"/api/v1/vehicles/{vehicle.pk}/")
        self.assertEqual(deleted.status_code, 204)
        self.assertFalse(Vehicle.objects.filter(pk=vehicle.pk).exists())

    def test_non_superuser_cannot_delete_vehicle(self):
        AssetManagerRole.objects.create(user=self.employee, scopes=["vehicles"])
        self.client.force_authenticate(self.employee)
        vehicle = Vehicle.objects.create(
            plate_number="粤BTEST02",
            name="保留车辆",
            seats=5,
            department=self.department,
        )
        deleted = self.client.delete(f"/api/v1/vehicles/{vehicle.pk}/")
        self.assertEqual(deleted.status_code, 405)
        self.assertTrue(Vehicle.objects.filter(pk=vehicle.pk).exists())

    def test_superuser_delete_vehicle_blocked_when_referenced(self):
        self.client.force_authenticate(self.admin)
        vehicle = Vehicle.objects.create(
            plate_number="粤BTEST03",
            name="引用车辆",
            seats=5,
            department=self.department,
        )
        VehicleExpense.objects.create(
            vehicle=vehicle,
            expense_type=VehicleExpense.ExpenseType.MAINTENANCE,
            occurred_on=date(2026, 8, 1),
            amount="500.00",
        )
        deleted = self.client.delete(f"/api/v1/vehicles/{vehicle.pk}/")
        self.assertEqual(deleted.status_code, 400)
        self.assertIn("无法删除", deleted.data["message"])
        self.assertTrue(Vehicle.objects.filter(pk=vehicle.pk).exists())

    def test_superuser_can_delete_vehicle_expense_with_ledger_record(self):
        self.client.force_authenticate(self.admin)
        vehicle = Vehicle.objects.create(
            plate_number="粤BTEST04",
            name="费用车辆",
            seats=5,
            department=self.department,
        )
        created = self.client.post(
            "/api/v1/vehicle-expenses/",
            {
                "vehicle": vehicle.pk,
                "expense_type": "maintenance",
                "occurred_on": "2026-08-01",
                "amount": "500.00",
                "notes": "保养",
            },
            format="json",
        )
        self.assertEqual(created.status_code, 201)
        record = VehicleExpense.objects.get(pk=created.data["id"])
        self.assertIsNotNone(record.expense)
        deleted = self.client.delete(f"/api/v1/vehicle-expenses/{record.pk}/")
        self.assertEqual(deleted.status_code, 204)
        self.assertFalse(VehicleExpense.objects.filter(pk=record.pk).exists())
        self.assertFalse(AdministrativeExpense.objects.filter(pk=record.expense_id).exists())

    def test_non_superuser_cannot_delete_vehicle_expense(self):
        AssetManagerRole.objects.create(user=self.employee, scopes=["vehicles"])
        self.client.force_authenticate(self.employee)
        vehicle = Vehicle.objects.create(
            plate_number="粤BTEST05",
            name="保留费用",
            seats=5,
            department=self.department,
        )
        record = VehicleExpense.objects.create(
            vehicle=vehicle,
            expense_type=VehicleExpense.ExpenseType.MAINTENANCE,
            occurred_on=date(2026, 8, 1),
            amount="100.00",
        )
        deleted = self.client.delete(f"/api/v1/vehicle-expenses/{record.pk}/")
        self.assertEqual(deleted.status_code, 405)
        self.assertTrue(VehicleExpense.objects.filter(pk=record.pk).exists())

    def test_edit_vehicle_expense_syncs_ledger_record(self):
        self.client.force_authenticate(self.admin)
        vehicle = Vehicle.objects.create(
            plate_number="粤BTEST06",
            name="同步费用",
            seats=5,
            department=self.department,
        )
        created = self.client.post(
            "/api/v1/vehicle-expenses/",
            {
                "vehicle": vehicle.pk,
                "expense_type": "maintenance",
                "occurred_on": "2026-08-01",
                "amount": "500.00",
            },
            format="json",
        )
        self.assertEqual(created.status_code, 201)
        updated = self.client.patch(
            f"/api/v1/vehicle-expenses/{created.data['id']}/",
            {"amount": "800.00", "occurred_on": "2026-08-02"},
            format="json",
        )
        self.assertEqual(updated.status_code, 200)
        record = VehicleExpense.objects.get(pk=created.data["id"])
        self.assertEqual(record.amount, Decimal("800.00"))
        self.assertEqual(record.expense.amount, Decimal("800.00"))
        self.assertEqual(record.expense.occurred_on, date(2026, 8, 2))
        self.assertEqual(record.expense.fiscal_year, 2026)

    def test_contract_search_and_type_filter(self):
        self.client.force_authenticate(self.admin)
        service_type = ContractType.objects.create(code="FILTER-SERVICE", name="过滤服务合同")
        lease_type = ContractType.objects.create(code="FILTER-LEASE", name="过滤租赁合同")
        Contract.objects.create(contract_no="SEARCH-001", name="网络运维服务", contract_type=service_type)
        Contract.objects.create(contract_no="SEARCH-002", name="办公室租赁", contract_type=lease_type)
        searched = self.client.get("/api/v1/contracts/?q=网络运维")
        self.assertEqual([row["contract_no"] for row in searched.data], ["SEARCH-001"])
        filtered = self.client.get(f"/api/v1/contracts/?contract_type={lease_type.id}")
        self.assertEqual([row["contract_no"] for row in filtered.data], ["SEARCH-002"])

    def test_dashboard_due_counts_match_filtered_contracts_and_vehicle_insurance(self):
        self.client.force_authenticate(self.admin)
        Contract.objects.create(
            contract_no="DUE-001", name="近期合同", owner=self.employee,
            status=Contract.Status.ACTIVE, end_date=timezone.localdate() + timedelta(days=10),
        )
        Contract.objects.create(
            contract_no="DONE-001", name="历史合同", owner=self.employee,
            status=Contract.Status.COMPLETED, end_date=timezone.localdate() - timedelta(days=10),
        )
        due_vehicle = Vehicle.objects.create(
            plate_number="粤BDUE01", name="保险待办车辆",
            insurance_expires_at=timezone.localdate() + timedelta(days=5),
        )
        Vehicle.objects.create(
            plate_number="粤BRET01", name="已处置车辆", status=Vehicle.Status.RETIRED,
            insurance_expires_at=timezone.localdate() - timedelta(days=5),
        )

        dashboard = self.client.get("/api/v1/dashboard/")
        contracts = self.client.get("/api/v1/contracts/?due=1")
        vehicles = self.client.get("/api/v1/vehicles/?insurance_due=1")

        self.assertEqual(dashboard.data["admin_tasks"]["expiring_contracts"], len(contracts.data))
        self.assertEqual(dashboard.data["admin_tasks"]["vehicle_insurance_due"], len(vehicles.data))
        self.assertEqual([row["contract_no"] for row in contracts.data], ["DUE-001"])
        self.assertEqual([row["id"] for row in vehicles.data], [due_vehicle.id])

    @patch("assets.views.nextcloud_storage.upload")
    def test_supplier_license_upload_marks_license_registered(self, storage_upload):
        self.client.force_authenticate(self.admin)
        supplier = Supplier.objects.create(code="SUP-LICENSE", name="证照供应商")
        upload = SimpleUploadedFile("营业执照.png", b"png-data", content_type="image/png")

        response = self.client.post(
            f"/api/v1/suppliers/{supplier.id}/files/",
            {"file": upload, "document_type": "business_license"},
            format="multipart",
        )

        self.assertEqual(response.status_code, 201)
        self.assertTrue(SupplierAttachment.objects.filter(supplier=supplier).exists())
        supplier.refresh_from_db()
        self.assertEqual(supplier.business_license_status, "registered")
        storage_upload.assert_called_once()

    def test_office_manager_only_sees_owned_contracts_in_office_detail(self):
        other = User.objects.create_user("office-other", password="pass")
        AssetManagerRole.objects.create(user=self.employee, scopes=["offices"])
        office = Office.objects.create(code="OFFICE-T01", name="测试办事处")
        own = Contract.objects.create(
            contract_no="OFFICE-OWN", name="本人租约", office=office, owner=self.employee,
        )
        Contract.objects.create(
            contract_no="OFFICE-OTHER", name="他人租约", office=office, owner=other,
        )
        self.client.force_authenticate(self.employee)

        response = self.client.get(f"/api/v1/offices/{office.id}/")

        self.assertEqual(response.status_code, 200)
        self.assertEqual([item["id"] for item in response.data["contracts"]], [own.id])

    def test_office_residents_link_to_employees_and_warn_on_multiple_places(self):
        self.employee.first_name = "张三"
        self.employee.last_name = ""
        self.employee.save(update_fields=["first_name", "last_name"])
        self.client.force_authenticate(self.admin)
        first = self.client.post(
            "/api/v1/offices/",
            {
                "code": "OFFICE-RES-01", "name": "宁波一号住宿点", "city": "宁波",
                "address": "江北区测试路1号", "resident_users": [self.employee.id],
                "resident_capacity": 3, "resident_count": 1,
            }, format="json",
        )
        self.assertEqual(first.status_code, 201)
        second = self.client.post(
            "/api/v1/offices/",
            {
                "code": "OFFICE-RES-02", "name": "宁波二号住宿点", "city": "宁波",
                "address": "鄞州区测试路2号", "resident_users": [self.employee.id],
                "resident_capacity": 2, "resident_count": 1,
            }, format="json",
        )
        self.assertEqual(second.status_code, 201)
        self.assertEqual(second.data["resident_user_details"][0]["display_name"], "张三")
        self.assertEqual(second.data["resident_capacity"], 2)
        self.assertEqual(second.data["resident_count"], 1)
        self.assertIn("张三 还居住在：宁波·宁波一号住宿点", second.data["resident_warnings"])

    def test_daily_task_marks_unhandled_contract_expired_without_email(self):
        contract = Contract.objects.create(
            contract_no="HT-EXPIRED-001", name="已过期合同",
            end_date=date.today() - timedelta(days=1), status=Contract.Status.ACTIVE,
        )
        send_daily_operational_notifications()
        contract.refresh_from_db()
        self.assertEqual(contract.status, Contract.Status.EXPIRED)

    def test_vehicle_expense_is_written_to_annual_ledger(self):
        vehicle = Vehicle.objects.create(plate_number="粤BTEST02", name="费用测试车", department=self.department)
        self.client.force_authenticate(self.admin)
        response = self.client.post(
            "/api/v1/vehicle-expenses/",
            {"vehicle": vehicle.id, "expense_type": "maintenance", "occurred_on": date.today(), "amount": "680.00", "odometer": 5000},
            format="json",
        )
        self.assertEqual(response.status_code, 201)
        expense = AdministrativeExpense.objects.get(source_type="vehicle", source_id=response.data["id"])
        self.assertEqual(expense.amount, 680)
        self.assertEqual(expense.fiscal_year, date.today().year)

    def test_purchase_approval_and_order_create_budget_stages(self):
        supplier = Supplier.objects.create(code="SUP-01", name="测试供应商")
        self.client.force_authenticate(self.employee)
        created = self.client.post(
            "/api/v1/purchase-requests/",
            {"reason": "补充办公用品", "category": self.category.id, "items": [{"name": "打印纸", "quantity": 10, "unit": "箱", "estimated_unit_price": "120.00", "specification": "A4"}]},
            format="json",
        )
        self.assertEqual(created.status_code, 201)
        self.client.force_authenticate(self.admin)
        approved = self.client.post(f"/api/v1/purchase-requests/{created.data['id']}/approve/", {}, format="json")
        self.assertEqual(approved.status_code, 200)
        approved_expense = AdministrativeExpense.objects.get(source_type="purchase_request", source_id=created.data["id"])
        self.assertEqual(approved_expense.amount_type, AdministrativeExpense.AmountType.APPROVED)
        order = self.client.post(
            "/api/v1/purchase-orders/",
            {"request": created.data["id"], "supplier": supplier.id, "status": "ordered", "ordered_on": date.today(), "items": [{"name": "打印纸", "quantity": 10, "unit": "箱", "unit_price": "110.00", "specification": "A4"}]},
            format="json",
        )
        self.assertEqual(order.status_code, 201)
        committed = AdministrativeExpense.objects.get(source_type="purchase_order", source_id=order.data["id"])
        self.assertEqual(committed.amount_type, AdministrativeExpense.AmountType.COMMITTED)
        self.assertEqual(committed.amount, 1100)

    def test_employee_only_sees_own_requests(self):
        other = User.objects.create_user("phase-other", password="pass")
        VehicleDispatch.objects.create(
            request_no="PC-OTHER", requester=other, purpose="其他", destination="园区",
            planned_departure_at=timezone.now() + timedelta(days=1), planned_return_at=timezone.now() + timedelta(days=1, hours=1),
        )
        self.client.force_authenticate(self.employee)
        response = self.client.get("/api/v1/vehicle-dispatches/")
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data, [])


class PopulateContractAmountsTests(TestCase):
    def test_only_unqualified_amount_descriptions_are_populated(self):
        pure = Contract.objects.create(
            contract_no="AMOUNT-PURE", name="明确总额", amount=0,
            amount_description="1,350元",
        )
        rate = Contract.objects.create(
            contract_no="AMOUNT-RATE", name="月度单价", amount=0,
            amount_description="固定金额：5654元/月",
        )
        existing = Contract.objects.create(
            contract_no="AMOUNT-EXISTING", name="已有金额", amount="999.00",
            amount_description="1200元",
        )

        dry_output = io.StringIO()
        call_command("populate_contract_amounts", dry_run=True, stdout=dry_output)
        pure.refresh_from_db()
        self.assertEqual(pure.amount, Decimal("0.00"))
        self.assertIn("补录 1 份合同", dry_output.getvalue())

        call_command("populate_contract_amounts")
        pure.refresh_from_db()
        rate.refresh_from_db()
        existing.refresh_from_db()
        self.assertEqual(pure.amount, Decimal("1350.00"))
        self.assertEqual(rate.amount, Decimal("0.00"))
        self.assertEqual(existing.amount, Decimal("999.00"))
